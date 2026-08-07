from flask import Flask, render_template, request, jsonify, redirect, url_for, session, Response, stream_with_context
from functools import wraps
from datetime import datetime, date, timedelta
import copy
import csv
import io
import json
import logging
import os
import time
import re
import smtplib
import uuid
from collections import Counter, defaultdict
from email.message import EmailMessage
from difflib import SequenceMatcher
from dateutil.relativedelta import relativedelta
from dateutil import parser as dateutil_parser
from calendar import monthrange
from zoneinfo import ZoneInfo
from apscheduler.schedulers.background import BackgroundScheduler
from apscheduler.triggers.cron import CronTrigger
from werkzeug.middleware.proxy_fix import ProxyFix
from pypdf import PdfReader

try:
    from dotenv import load_dotenv
except ImportError:  # pragma: no cover
    def load_dotenv(*args, **kwargs):  # type: ignore[misc]
        return False

# Load variables from a local .env file so config works regardless of how the
# app is launched (gunicorn, `flask run`, `python app.py`, tests, etc.).
# gunicorn imports the module directly and never invokes the Flask CLI, so we
# cannot rely on Flask's automatic .env loading here.
load_dotenv()

try:
    from openai import OpenAI, APITimeoutError
except ImportError:  # pragma: no cover
    OpenAI = None  # type: ignore[misc, assignment]

    class APITimeoutError(Exception):  # type: ignore[no-redef]
        pass

logger = logging.getLogger(__name__)

# Bank statement LLM import: max upload size, max chars sent to the model (approximate context limit).
# Environment:
#   OPENAI_API_KEY — required for LLM extraction / classification (tabular CSV/XLSX can skip extraction).
#   OPENAI_MODEL — optional, default gpt-4o-mini.
#   OPENAI_BASE_URL — optional; use for OpenAI-compatible API endpoints.
#   OPENAI_EXTRACTION_TIMEOUT — per LLM chunk; default 90s (must stay below gunicorn --timeout).
STATEMENT_MAX_BYTES = 2 * 1024 * 1024
STATEMENT_MAX_CHARS_FOR_LLM = 100_000
# When full-statement LLM extraction is needed, send text in chunks so each call finishes
# under the worker timeout (Revolut-sized months were timing out at ~100k chars / 120s).
STATEMENT_LLM_CHUNK_CHARS = 28_000
# Minimum rows for a local tabular parse to count as success (else fall through to LLM).
SPENDING_TABULAR_MIN_ROWS = 1
# Revolut-style State values we keep; others (PENDING/FAILED/…) are skipped.
_SPENDING_TABULAR_KEEP_STATES = frozenset({'completed', 'reverted', 'completed_reverted'})
# Spending preview API: cap intermediate-representation strings so responses stay practical.
SPENDING_PIPELINE_EXTRACT_PREVIEW = 14_000
SPENDING_PIPELINE_HINTS_PREVIEW = 18_000
SPENDING_PIPELINE_DIRECTION_HINT_ROWS = 150
SPENDING_PDF_COLUMN_HINTS_BANNER = (
    '\n\n--- COLUMN HINTS (from PDF layout: use money_in vs money_out per row; '
    'do not infer from description) ---\n'
)
# Layout-first extraction: parse LINE_HINT rows locally; small LLM call only for skip-line regexes.
# Fallback compares parsed rows to *eligible* hint lines (non-skipped rows with money), not total hints.
SPENDING_LAYOUT_MIN_HINT_ROWS = 5
SPENDING_LAYOUT_FALLBACK_MIN_PARSED = 3
SPENDING_LAYOUT_FALLBACK_RATIO = 0.18
SPENDING_LAYOUT_EXCERPT_MAX_LINES = 90
SPENDING_LAYOUT_EXCERPT_MAX_CHARS = 14_000
SPENDING_LAYOUT_SAMPLE_HINTS = 12
SPENDING_LAYOUT_MAPPING_LLM_MAX_CHARS = 9000
_SPENDING_LAYOUT_DEFAULT_SKIP_REGEXES = [
    r'(?i)^\s*balance\b',
    r'(?i)brought\s+forward',
    r'(?i)opening\s+balance',
    r'(?i)closing\s+balance',
    r'(?i)^\s*continued\s*$',
    r'(?i)^\s*date\s+description',
    r'(?i)\bpaid\s+in\b.*\bpaid\s+out\b',
    r'(?i)^\s*payment\s+type\s+',
    r'(?i)^\s*sheet\s+\d',
]
_LINE_HINT_ROW_RE = re.compile(
    r'^LINE_HINT money_in=(\d+(?:\.\d+)?|0) money_out=(\d+(?:\.\d+)?|0) \| (.+)$'
)
# Import preview: flag when existing transaction amount is within this many pounds of bank or share.
STATEMENT_AUDIT_MATCH_TOLERANCE = 5.0
# Internal transfer auto/manual pairing: amount difference allowed between legs.
# Tighter for small amounts (1% of larger leg, min 5p, max 50p when larger leg < £50) to avoid
# false auto-matches. Larger transfers still allow up to £5 (fees/FX) via min(5, 1% of leg).
SPENDING_TRANSFER_MIN_AMOUNT_DIFF = 0.05
SPENDING_TRANSFER_LARGE_LEG_PCT = 0.01
SPENDING_TRANSFER_SMALL_LEG_MAX = 50.0
SPENDING_TRANSFER_SMALL_LEG_ABS_CAP = 0.5
SPENDING_TRANSFER_LARGE_LEG_ABS_CAP = 5.0
# Internal transfer matching (e.g. HSBC to Revolut): max calendar gap between legs (day count).
SPENDING_TRANSFER_MAX_DAY_GAP = 1
SPENDING_TRANSFER_UNMATCHED_PREVIEW_MAX = 12
# Anomaly: ignore tiny category spend even if 1.5× prior average (reduces noise).
SPENDING_ANOMALY_MIN_GBP = 20.0
SPENDING_ANOMALY_MIN_OUTGOING_PCT = 0.01

# SMTP for monthly bill-upload reminders (optional). If SMTP_HOST is unset, reminders are skipped.
#   SMTP_HOST, SMTP_PORT (default 587), SMTP_USER, SMTP_PASSWORD
#   SMTP_USE_TLS — default true (STARTTLS)
#   MAIL_FROM — From address
#   FINANCE_TRACKER_NOTIFY_EMAIL — default To when loan has no bill_reminder_email
#   PUBLIC_BASE_URL — e.g. https://tracker.example.com (no trailing slash) for links in emails
BILL_REMINDER_HOUR = 9
BASELINE_SIM_THRESHOLD = 0.75
BASELINE_NOTE_MAX_LEN = 300
SPENDING_ALLOWED_CATEGORIES = [
    'housing',
    'groceries',
    'transport',
    'utilities',
    'subscriptions',
    'health',
    'dining',
    'entertainment',
    'travel',
    'shopping',
    'debt',
    'savings',
    'other',
    'unclassified',
]
SPENDING_CATEGORY_SET = set(SPENDING_ALLOWED_CATEGORIES)
# Daily Budget: bill-like categories are reserved monthly (not daily discretionary).
DAILY_BUDGET_BILL_CATEGORIES = frozenset({'housing', 'utilities', 'subscriptions', 'debt'})
DAILY_BUDGET_IGNORE_CATEGORIES = frozenset({'savings'})
DAILY_BUDGET_MODES = frozenset({'fixed', 'envelope', 'carry_surplus'})
DAILY_BUDGET_UNDERSPEND_PRIORITIES = frozenset({'debt_first', 'goals_first'})
DAILY_BUDGET_MANUAL_MATCH_RATIO = 0.72
# Manual vs statement: allow slight amount drift (rounding / tips).
DAILY_BUDGET_MANUAL_MATCH_AMOUNT_TOL = 0.15
# Statement date may post up to N days after manual (banks often lag); not the reverse.
DAILY_BUDGET_MANUAL_MATCH_DATE_SLACK_DAYS = 3
# Preview suggestions for near-misses (looser than auto-match; UI-only exclude).
DAILY_BUDGET_MANUAL_SUGGEST_DATE_SLACK_DAYS = 7
DAILY_BUDGET_MANUAL_SUGGEST_AMOUNT_TOL = 5.0
DAILY_BUDGET_MANUAL_SUGGEST_LIMIT = 3
# Statement row vs expected monthly bill_items (label + amount).
SPENDING_EXPECTED_BILL_SIM_THRESHOLD = 0.75
DAILY_ENTRY_CATEGORIES = [
    c for c in SPENDING_ALLOWED_CATEGORIES if c not in ('unclassified',)
]


def _sanitize_note(raw) -> str:
    if raw is None:
        return ''
    return str(raw).strip()[:BASELINE_NOTE_MAX_LEN]


def _parse_baseline_item_row(row: dict) -> dict | None:
    """Single baseline row from client/save payload; returns None if invalid."""
    if not isinstance(row, dict):
        return None
    try:
        ab = float(row.get('amount_bank'))
    except (TypeError, ValueError):
        return None
    if ab <= 0:
        return None
    share = row.get('amount_share')
    if share is None and row.get('amount_default') is not None:
        share = row.get('amount_default')
    if share is None:
        share = round(ab / 2.0, 2)
    else:
        try:
            share = round(float(share), 2)
        except (TypeError, ValueError):
            return None
    if share <= 0:
        return None
    desc = str(row.get('description') or '').strip()[:500]
    if not desc:
        desc = 'Bill'
    cat = row.get('category')
    if cat is not None:
        cat = str(cat).strip()[:120]
        if not cat:
            cat = None
    else:
        cat = None
    return {
        'id': str(row.get('id') or uuid.uuid4()),
        'description': desc,
        'amount_bank': round(ab, 2),
        'amount_share': share,
        'category': cat,
        'note': _sanitize_note(row.get('note')),
    }


def _baseline_item_from_stored(b: dict) -> dict:
    """Normalize stored baseline dict for API/merge output."""
    try:
        ab = float(b.get('amount_bank', 0))
    except (TypeError, ValueError):
        ab = 0.0
    try:
        sh = float(b.get('amount_share', ab / 2.0))
    except (TypeError, ValueError):
        sh = round(ab / 2.0, 2) if ab else 0.0
    desc = str(b.get('description') or '').strip()[:500] or 'Bill'
    cat = b.get('category')
    if cat is not None:
        cat = str(cat)[:120]
    else:
        cat = None
    return {
        'id': str(b.get('id') or uuid.uuid4()),
        'description': desc,
        'amount_bank': round(ab, 2),
        'amount_share': round(sh, 2),
        'category': cat,
        'note': _sanitize_note(b.get('note')),
    }


def merge_baseline_with_candidates(existing: list | None, candidates: list) -> list:
    """
    Combine a new statement's normalized candidates with saved baseline rows.
    Each candidate updates at most one unmatched baseline row (description similarity + bank amount).
    Unmatched existing rows are kept at the end. New candidates with no match become new rows (new id, empty note).
    """
    existing = existing or []
    if not isinstance(candidates, list):
        return [_baseline_item_from_stored(b) for b in existing]

    matched_ids: set[str] = set()
    out: list = []

    for c in candidates:
        if not isinstance(c, dict):
            continue
        try:
            cb = float(c.get('amount_bank', 0))
        except (TypeError, ValueError):
            continue
        if cb <= 0:
            continue
        cd = str(c.get('description') or '')
        best_b = None
        best_sim = -1.0
        for b in existing:
            bid = str(b.get('id') or '')
            if not bid or bid in matched_ids:
                continue
            try:
                ab = float(b.get('amount_bank', 0))
            except (TypeError, ValueError):
                continue
            if not _amounts_close_for_compare(ab, cb):
                continue
            sim = _description_similarity(str(b.get('description', '')), cd)
            if sim >= BASELINE_SIM_THRESHOLD and sim > best_sim:
                best_sim = sim
                best_b = b

        share = c.get('amount_default')
        if share is None:
            try:
                share = round(float(c.get('amount_bank')) / 2.0, 2)
            except (TypeError, ValueError):
                share = round(cb / 2.0, 2)
        else:
            try:
                share = round(float(share), 2)
            except (TypeError, ValueError):
                share = round(cb / 2.0, 2)
        cat = c.get('category')
        if cat is not None:
            cat = str(cat)[:120]
        else:
            cat = None
        desc = str(c.get('description') or '').strip()[:500] or 'Bill'

        if best_b is not None:
            bid = str(best_b.get('id'))
            matched_ids.add(bid)
            out.append({
                'id': bid,
                'description': desc,
                'amount_bank': round(cb, 2),
                'amount_share': share,
                'category': cat,
                'note': _sanitize_note(best_b.get('note')),
            })
        else:
            out.append({
                'id': str(uuid.uuid4()),
                'description': desc,
                'amount_bank': round(cb, 2),
                'amount_share': share,
                'category': cat,
                'note': '',
            })

    for b in existing:
        bid = str(b.get('id') or '')
        if bid and bid not in matched_ids:
            out.append(_baseline_item_from_stored(b))

    return out


def _env_bool(key, default=False):
    v = os.getenv(key)
    if v is None:
        return default
    return v.strip().lower() in ('1', 'true', 'yes', 'on')


app = Flask(__name__)
app.config['SECRET_KEY'] = os.getenv('SECRET_KEY', 'your-secret-key-here')
app.config['SESSION_COOKIE_NAME'] = 'finance_tracker_session'
app.config['SESSION_COOKIE_HTTPONLY'] = True
app.config['SESSION_COOKIE_SAMESITE'] = 'Lax'
# Behind nginx/Caddy with TLS, set SESSION_COOKIE_SECURE=1 so browsers send the cookie on HTTPS.
app.config['SESSION_COOKIE_SECURE'] = _env_bool('SESSION_COOKIE_SECURE', False)
app.config['PERMANENT_SESSION_LIFETIME'] = timedelta(days=31)

# Honour X-Forwarded-* from a reverse proxy so request.scheme / host match what the browser uses.
if _env_bool('TRUST_PROXY', True):
    app.wsgi_app = ProxyFix(app.wsgi_app, x_for=1, x_proto=1, x_host=1, x_port=1, x_prefix=1)

# Avoid sticky browser caches of CSS/JS after deploys.
app.config['SEND_FILE_MAX_AGE_DEFAULT'] = 0
app.config['TEMPLATES_AUTO_RELOAD'] = True


def _static_asset_version(filename: str) -> str:
    """Cache-bust token from static file mtime (changes on each deploy edit)."""
    try:
        path = os.path.join(app.static_folder or 'static', filename)
        return str(int(os.path.getmtime(path)))
    except OSError:
        return '0'


@app.context_processor
def inject_asset_helpers():
    def asset_url(filename: str) -> str:
        return url_for('static', filename=filename, v=_static_asset_version(filename))

    return {'asset_url': asset_url}


@app.after_request
def _no_cache_static(response):
    if request.path.startswith('/static/'):
        response.headers['Cache-Control'] = 'no-cache, no-store, must-revalidate'
        response.headers['Pragma'] = 'no-cache'
        response.headers['Expires'] = '0'
    return response


def load_users():
    """
    Load users from FINANCE_TRACKER_USERS env var:
    "username:password,another:secret"
    Falls back to defaults for local development.
    """
    raw_users = os.getenv('FINANCE_TRACKER_USERS', '').strip()
    if not raw_users:
        return {
            'admin': 'admin123',
            'user': 'user123'
        }

    users = {}
    for pair in raw_users.split(','):
        if ':' not in pair:
            continue
        username, password = pair.split(':', 1)
        username = username.strip()
        password = password.strip()
        if username and password:
            users[username] = password

    return users or {
        'admin': 'admin123',
        'user': 'user123'
    }

USERS = load_users()

def login_required(f):
    @wraps(f)
    def decorated_function(*args, **kwargs):
        if 'username' not in session:
            return redirect(url_for('login'))
        return f(*args, **kwargs)
    return decorated_function

scheduler = BackgroundScheduler()
scheduler.start()

def load_data():
    default_data = {
        'loans': {},  # Dictionary to store multiple loans
        'users': {},
    }
    try:
        with open('data.json', 'r') as f:
            content = f.read().strip()
            if not content:  # Handle empty file
                save_data(default_data)
                return default_data
            
            # Load existing data and merge with defaults to handle missing fields
            existing_data = json.loads(content)
            if 'loans' not in existing_data:
                # Convert old format to new format
                old_data = existing_data
                existing_data = default_data
                if 'loan_amount' in old_data:
                    loan_id = 'loan_1'  # Default ID for the first loan
                    existing_data['loans'][loan_id] = {
                        'name': 'My Loan',  # Default name
                        'loan_amount': old_data.get('loan_amount', 0),
                        'interest_rate': old_data.get('interest_rate', 0),
                        'interest_day': old_data.get('interest_day', 1),
                        'transactions': old_data.get('transactions', [])
                    }
            
            changed = _ensure_data_shape(existing_data)
            if changed:
                save_data(existing_data)
            return existing_data
            
    except (FileNotFoundError, json.JSONDecodeError):
        save_data(default_data)
        return default_data

def save_data(data):
    with open('data.json', 'w') as f:
        json.dump(data, f, indent=2)


def _ensure_data_shape(data: dict) -> bool:
    """Runtime-safe migration for legacy data.json structures."""
    changed = False
    if 'loans' not in data or not isinstance(data.get('loans'), dict):
        data['loans'] = {}
        changed = True
    if 'users' not in data or not isinstance(data.get('users'), dict):
        data['users'] = {}
        changed = True
    return changed


def _normalize_bank_source(raw) -> str | None:
    """Optional bank/account label for a statement import (UI column: Source)."""
    s = str(raw or '').strip()[:80]
    return s or None


def _collect_bank_sources(spending: dict) -> list[str]:
    """Unique prior bank_source values for autocomplete (statements + transactions)."""
    seen: dict[str, str] = {}
    for bucket_key in ('statements', 'transactions'):
        for row in spending.get(bucket_key) or []:
            if not isinstance(row, dict):
                continue
            label = _normalize_bank_source(row.get('bank_source'))
            if not label:
                continue
            key = label.casefold()
            if key not in seen:
                seen[key] = label
    return sorted(seen.values(), key=lambda x: x.casefold())


def _ensure_user_spending(data: dict, username: str) -> tuple[dict, bool]:
    changed = False
    users = data.setdefault('users', {})
    if username not in users or not isinstance(users.get(username), dict):
        users[username] = {}
        changed = True
    user_bucket = users[username]
    spending = user_bucket.get('spending')
    if not isinstance(spending, dict):
        spending = {}
        user_bucket['spending'] = spending
        changed = True
    defaults = {
        'statements': [],
        'transactions': [],
        'monthly_insights': {},
        'classification_overrides': {},
        'classification_cache': {},
        'daily_budget': {},
    }
    for key, default in defaults.items():
        if key not in spending or not isinstance(spending.get(key), type(default)):
            spending[key] = default.copy() if isinstance(default, dict) else list(default)
            changed = True
    db, db_changed = _ensure_daily_budget(spending)
    if db_changed:
        changed = True
    return spending, changed


def _ensure_daily_budget(spending: dict) -> tuple[dict, bool]:
    changed = False
    bucket = spending.get('daily_budget')
    if not isinstance(bucket, dict):
        bucket = {}
        spending['daily_budget'] = bucket
        changed = True
    plan = bucket.get('plan')
    if not isinstance(plan, dict):
        plan = {}
        bucket['plan'] = plan
        changed = True
    plan_defaults = {
        'income_monthly': 0.0,
        'bills_monthly': 0.0,
        'savings_percent': 20.0,
        'daily_mode': 'envelope',
        # Day of month salary usually arrives (1–31). Daily pacing uses the pay
        # period from that day through the day before the next payday.
        'pay_day': 1,
        # ISO date when the user began daily tracking. Mid-period starts pro-rate the
        # remaining pool so empty pre-join days are not treated as £0 underspend.
        'tracking_from': None,
        'source_month': None,
        'bill_items': [],
        # When daily leftover is skimmed: repay overspend debt first, or fill goals first.
        'underspend_priority': 'debt_first',
        'updated_at': None,
    }
    for key, default in plan_defaults.items():
        if key not in plan:
            plan[key] = list(default) if isinstance(default, list) else default
            changed = True
    if plan.get('daily_mode') not in DAILY_BUDGET_MODES:
        plan['daily_mode'] = 'envelope'
        changed = True
    if plan.get('underspend_priority') not in DAILY_BUDGET_UNDERSPEND_PRIORITIES:
        plan['underspend_priority'] = 'debt_first'
        changed = True
    try:
        pay_day = int(plan.get('pay_day'))
    except (TypeError, ValueError):
        pay_day = 1
    if pay_day < 1 or pay_day > 31:
        pay_day = 1
    if plan.get('pay_day') != pay_day:
        plan['pay_day'] = pay_day
        changed = True
    if not isinstance(plan.get('bill_items'), list):
        plan['bill_items'] = []
        changed = True
    goals = bucket.get('goals')
    if not isinstance(goals, list):
        bucket['goals'] = []
        changed = True
    if 'overspend_debt' not in bucket:
        bucket['overspend_debt'] = None
        changed = True
    decisions = bucket.get('overspend_decisions')
    if not isinstance(decisions, dict):
        bucket['overspend_decisions'] = {}
        changed = True
    return bucket, changed


def _normalize_label(s: str) -> str:
    s = (s or '').strip().lower()
    s = re.sub(r'[^a-z0-9 ]+', ' ', s)
    s = re.sub(r'\s+', ' ', s).strip()
    return s


def _normalize_spending_direction(raw_direction, description: str, amount: float) -> str:
    d = str(raw_direction or '').strip().lower()
    if d in ('incoming', 'outgoing'):
        return d
    desc = _normalize_label(description)
    incoming_words = ('salary', 'payroll', 'refund', 'interest', 'benefit', 'bonus', 'deposit', 'received')
    outgoing_words = ('debit', 'payment', 'card', 'purchase', 'dd', 'standing order', 'transfer out')
    if any(w in desc for w in incoming_words):
        return 'incoming'
    if any(w in desc for w in outgoing_words):
        return 'outgoing'
    return 'outgoing' if amount >= 0 else 'incoming'


def _spending_optional_money(val) -> float | None:
    if val is None or val == '':
        return None
    if isinstance(val, (int, float)) and not isinstance(val, bool):
        return float(val) if float(val) > 0 else None
    s = str(val).strip().replace(',', '')
    s = re.sub(r'^[$£€]\s*', '', s)
    if not s:
        return None
    try:
        x = float(s)
    except (TypeError, ValueError):
        return None
    return x if x > 0 else None


def _direction_and_amount_from_money_columns(row: dict) -> tuple[str | None, float | None]:
    """
    Prefer explicit money_in / money_out from the LLM (matches separate Paid in / Paid out columns).
    Aliases: paid_in/paid_out, credit/debit amounts.
    """
    mi = _spending_optional_money(row.get('money_in'))
    mo = _spending_optional_money(row.get('money_out'))
    if mi is None:
        mi = _spending_optional_money(row.get('paid_in'))
    if mo is None:
        mo = _spending_optional_money(row.get('paid_out'))

    if mi is not None and mo is None:
        return 'incoming', mi
    if mo is not None and mi is None:
        return 'outgoing', mo
    if mi is not None and mo is not None:
        if mi > 0 and mo <= 0:
            return 'incoming', mi
        if mo > 0 and mi <= 0:
            return 'outgoing', mo
        # Both positive: rare; prefer LLM direction if it matches one column
        d = str(row.get('direction') or '').strip().lower()
        if d == 'incoming':
            return 'incoming', mi
        if d == 'outgoing':
            return 'outgoing', mo
        return 'outgoing', mo
    return None, None


def _spending_fingerprint(
    report_month: str,
    d_str: str,
    amount: float,
    direction: str,
    description: str,
) -> str:
    desc = _normalize_label(description)
    rm = (report_month or '').strip()[:7]
    return f'{rm}|{d_str}|{direction}|{round(float(amount), 2):.2f}|{desc}'


def _apply_spending_preview_duplicate_marks(
    report_month: str, rows: list, spending: dict
) -> tuple[int, int, list[str]]:
    """
    Set preview_duplicate / preview_duplicate_reason on each preview row, using the
    same fingerprint scheme as import. First occurrence in the upload is kept; later
    matching rows in the file, or any row matching the ledger, are marked duplicate.

    Also sets preview_review_reason for non-duplicate outgoing rows:
    - expected_bill — matches an included daily-budget bill_item (not “missed”)
    - missed — on statement but not in the unmatched manual list (review these)

    Returns (ledger_duplicate_count, upload_duplicate_count, ledger_fingerprints_for_ui).
    The fingerprint list is restricted to the current report month for a smaller JSON payload.
    """
    tx_store = spending.get('transactions') or []
    ledger_fps = {str(t.get('fingerprint')) for t in tx_store if t.get('fingerprint')}
    rm = (report_month or '').strip()[:7]
    seen: set[str] = set()
    claimed_manual_ids: set[str] = set()
    used_bill_idxs: set[int] = set()
    bill_items = _spending_included_bill_items(spending)
    led = 0
    dup_upload = 0
    for r in rows:
        r['preview_review_reason'] = None
        r['preview_manual_suggestions'] = []
        fp = _spending_fingerprint(
            rm,
            str(r.get('date') or ''),
            float(r.get('amount') or 0),
            str(r.get('direction') or 'outgoing'),
            str(r.get('description') or ''),
        )
        if fp in ledger_fps:
            r['preview_duplicate'] = True
            r['preview_duplicate_reason'] = 'ledger'
            led += 1
        elif fp in seen:
            r['preview_duplicate'] = True
            r['preview_duplicate_reason'] = 'upload'
            dup_upload += 1
        else:
            direction = str(r.get('direction') or 'outgoing')
            manual = _daily_budget_fuzzy_match_manual(
                spending,
                date_str=str(r.get('date') or '')[:10],
                amount=float(r.get('amount') or 0),
                description=str(r.get('description') or ''),
                direction=direction,
                exclude_ids=claimed_manual_ids,
            )
            if manual is not None:
                mid = str(manual.get('id') or '')
                if mid:
                    claimed_manual_ids.add(mid)
                r['preview_duplicate'] = True
                r['preview_duplicate_reason'] = 'manual'
                led += 1
            else:
                r['preview_duplicate'] = False
                r['preview_duplicate_reason'] = None
                seen.add(fp)
                if direction == 'outgoing':
                    bill_j = _match_expected_bill_item(
                        str(r.get('description') or ''),
                        float(r.get('amount') or 0),
                        bill_items,
                        used_idxs=used_bill_idxs,
                    )
                    if bill_j is not None:
                        used_bill_idxs.add(bill_j)
                        r['preview_review_reason'] = 'expected_bill'
                    else:
                        r['preview_review_reason'] = 'missed'
                        r['preview_manual_suggestions'] = _daily_budget_suggest_manual_matches(
                            spending,
                            date_str=str(r.get('date') or '')[:10],
                            amount=float(r.get('amount') or 0),
                            description=str(r.get('description') or ''),
                            direction=direction,
                            exclude_ids=claimed_manual_ids,
                        )
    month_prefix = f'{rm}|' if len(rm) == 7 else None
    client_fps = sorted(fp for fp in ledger_fps if month_prefix and fp.startswith(month_prefix))
    return led, dup_upload, client_fps


def _report_month_for_spending_tx(t: dict) -> str:
    """Bucket key for monthly insights: prefer explicit report_month from import."""
    return str(t.get('report_month') or t.get('month') or '').strip()


def _parse_optional_amount(raw) -> float | None:
    """Parse a query amount; blank/invalid → None (ignore filter)."""
    if raw is None:
        return None
    s = str(raw).strip().replace(',', '')
    if not s:
        return None
    try:
        return float(s)
    except (TypeError, ValueError):
        return None


def _spending_tx_search_haystack(tx: dict) -> str:
    """Lowercase blob for free-text `q` matching across common ledger fields."""
    parts = [
        str(tx.get('date') or ''),
        str(tx.get('description') or ''),
        str(tx.get('category') or ''),
        str(tx.get('bank_source') or ''),
        str(tx.get('direction') or ''),
        str(tx.get('source') or ''),
        str(_report_month_for_spending_tx(tx) or ''),
    ]
    try:
        amt = float(tx.get('amount'))
        parts.append(f'{amt:.2f}')
        parts.append(str(tx.get('amount')))
    except (TypeError, ValueError):
        parts.append(str(tx.get('amount') or ''))
    return ' '.join(parts).lower()


def _search_spending_transactions(
    transactions: list,
    *,
    q: str = '',
    date_from: str | None = None,
    date_to: str | None = None,
    category: str | None = None,
    direction: str | None = None,
    bank_source: str | None = None,
    min_amount: float | None = None,
    max_amount: float | None = None,
) -> list[dict]:
    """
    Filter the flat spending ledger (all months). Free-text `q` is a case-insensitive
    substring over description, date, category, bank_source, direction, source, month, amount.
    Results are newest-first by ledger date, then description, then id.
    """
    q_norm = (q or '').strip().lower()
    d_from = _parse_iso_date(date_from or '')
    d_to = _parse_iso_date(date_to or '')
    cat = (category or '').strip().lower()
    direction_norm = (direction or '').strip().lower()
    if direction_norm and direction_norm not in ('incoming', 'outgoing'):
        direction_norm = ''
    src_raw = (bank_source or '').strip()
    src_none = src_raw == '__none__'
    src_norm = '' if src_none else src_raw.casefold()

    matched: list[dict] = []
    for tx in transactions or []:
        if not isinstance(tx, dict):
            continue
        tx_date = _parse_iso_date(str(tx.get('date') or ''))
        if d_from is not None and (tx_date is None or tx_date < d_from):
            continue
        if d_to is not None and (tx_date is None or tx_date > d_to):
            continue
        if direction_norm and str(tx.get('direction') or '').lower() != direction_norm:
            continue
        if cat:
            tx_cat = str(tx.get('category') or 'unclassified').strip().lower() or 'unclassified'
            if tx_cat != cat:
                continue
        if src_none:
            if _normalize_bank_source(tx.get('bank_source')):
                continue
        elif src_norm:
            label = _normalize_bank_source(tx.get('bank_source'))
            if not label or label.casefold() != src_norm:
                continue
        try:
            amount = float(tx.get('amount'))
        except (TypeError, ValueError):
            amount = None
        if min_amount is not None:
            if amount is None or amount < min_amount:
                continue
        if max_amount is not None:
            if amount is None or amount > max_amount:
                continue
        if q_norm and q_norm not in _spending_tx_search_haystack(tx):
            continue
        matched.append(tx)

    matched.sort(
        key=lambda t: (
            str(t.get('date') or ''),
            str(t.get('description') or ''),
            str(t.get('id') or ''),
        ),
        reverse=True,
    )
    return matched


def _first_day_of_month(month_key: str) -> date | None:
    mk = (month_key or '').strip()
    if len(mk) == 7:
        try:
            return datetime.strptime(mk + '-01', '%Y-%m-%d').date()
        except ValueError:
            return None
    return _parse_iso_date(mk[:10])


def _last_day_of_month(month_key: str) -> date | None:
    d0 = _first_day_of_month(month_key)
    if d0 is None:
        return None
    last = monthrange(d0.year, d0.month)[1]
    return date(d0.year, d0.month, last)


def _parse_spending_period_from_values(
    report_month_raw: str | None,
    period_start_raw: str | None,
    period_end_raw: str | None,
) -> tuple[dict | None, str | None]:
    """
    Resolve reporting month + inclusive date range for spending imports.
    If only report_month (YYYY-MM) is given, range is full calendar month.
    """
    report_month = (report_month_raw or '').strip()[:7]
    ps = (period_start_raw or '').strip()[:10]
    pe = (period_end_raw or '').strip()[:10]

    if not report_month and not ps and not pe:
        today = date.today()
        report_month = today.strftime('%Y-%m')
        d_start = date(today.year, today.month, 1)
        last = monthrange(today.year, today.month)[1]
        d_end = date(today.year, today.month, last)
    elif report_month and not ps and not pe:
        d0 = _first_day_of_month(report_month)
        if d0 is None:
            return None, 'Invalid report_month (use YYYY-MM).'
        d_start = d0
        d_end = _last_day_of_month(report_month)
        if d_end is None:
            return None, 'Invalid report_month (use YYYY-MM).'
    else:
        d_start = _parse_iso_date(ps) if ps else None
        d_end = _parse_iso_date(pe) if pe else None
        if d_start is None or d_end is None:
            return None, 'period_start and period_end are required as YYYY-MM-DD (or set report_month only).'
        if d_start > d_end:
            return None, 'period_start must be on or before period_end.'
        if (d_end - d_start).days > 400:
            return None, 'Reporting period is too long (max 400 days).'
        if not report_month:
            report_month = d_start.strftime('%Y-%m')

    if d_start > d_end:
        return None, 'period_start must be on or before period_end.'
    if (d_end - d_start).days > 400:
        return None, 'Reporting period is too long (max 400 days).'

    return {
        'report_month': report_month,
        'period_start': d_start.strftime('%Y-%m-%d'),
        'period_end': d_end.strftime('%Y-%m-%d'),
        'period_start_date': d_start,
        'period_end_date': d_end,
    }, None


def _filter_spending_rows_by_period(rows: list, d_start: date, d_end: date) -> tuple[list, int, int]:
    """
    Keep rows whose primary transaction date falls in [d_start, d_end] inclusive.

    Rows whose started_date falls outside the period (but primary/completed date
    is inside) are kept and tagged as date-boundary cases for preview UX.
    Returns (kept_rows, dropped_count, boundary_count).
    """
    kept = []
    dropped = 0
    boundary = 0
    for row in rows:
        d = _parse_iso_date(str(row.get('date') or ''))
        if d is None:
            dropped += 1
            continue
        if d < d_start or d > d_end:
            dropped += 1
            continue
        started = _parse_iso_date(str(row.get('started_date') or ''))
        if started is not None and (started < d_start or started > d_end):
            row['date_boundary'] = True
            row['date_boundary_reason'] = 'started_outside'
            boundary += 1
        else:
            row['date_boundary'] = False
            row.pop('date_boundary_reason', None)
        kept.append(row)
    return kept, dropped, boundary


def _resolve_spending_transaction_dates(row: dict) -> tuple[date | None, date | None, date | None]:
    """
    Resolve (primary, started, completed) dates from an extracted transaction row.

    Prefer completed/settled date as the ledger date when both start and completed
    exist (e.g. Revolut Started Date vs Completed Date). Fall back to explicit
    date/statement_date, then started.
    """
    if not isinstance(row, dict):
        return None, None, None
    started = _parse_bank_statement_date(
        str(
            row.get('started_date')
            or row.get('start_date')
            or row.get('Started Date')
            or row.get('started')
            or ''
        )
    )
    completed = _parse_bank_statement_date(
        str(
            row.get('completed_date')
            or row.get('completion_date')
            or row.get('Completed Date')
            or row.get('completed')
            or row.get('settled_date')
            or ''
        )
    )
    explicit = _parse_bank_statement_date(
        str(row.get('date') or row.get('statement_date') or '')
    )
    if completed is not None:
        primary = completed
    elif explicit is not None:
        primary = explicit
    else:
        primary = started
    return primary, started, completed


def _strip_llm_markdown_fence(s: str) -> str:
    t = (s or '').strip()
    if not t.startswith('```'):
        return t
    lines = t.split('\n')
    if lines and lines[0].startswith('```'):
        lines = lines[1:]
    if lines and lines[-1].strip().startswith('```'):
        lines = lines[:-1]
    return '\n'.join(lines).strip()


def _extract_first_json_object(s: str) -> str:
    """If the model wrapped JSON in prose or fences, keep the outermost {...} block."""
    start = s.find('{')
    if start == -1:
        return s
    depth = 0
    in_str = False
    escape = False
    for i in range(start, len(s)):
        c = s[i]
        if in_str:
            if escape:
                escape = False
            elif c == '\\':
                escape = True
            elif c == '"':
                in_str = False
            continue
        if c == '"':
            in_str = True
        elif c == '{':
            depth += 1
        elif c == '}':
            depth -= 1
            if depth == 0:
                return s[start : i + 1]
    return s[start:]


def _escape_raw_controls_in_json_strings(s: str) -> str:
    """
    Models sometimes emit raw newlines/tabs inside JSON string values, which
    standard json.loads rejects. Escape those while respecting backslash escapes.
    """
    out: list[str] = []
    i = 0
    n = len(s)
    in_str = False
    escape = False
    while i < n:
        c = s[i]
        if not in_str:
            if c == '"':
                in_str = True
            out.append(c)
            i += 1
            continue
        if escape:
            out.append(c)
            escape = False
            i += 1
            continue
        if c == '\\':
            out.append(c)
            escape = True
            i += 1
            continue
        if c == '"':
            in_str = False
            out.append(c)
            i += 1
            continue
        if c == '\n':
            out.append('\\n')
            i += 1
            continue
        if c == '\r':
            out.append('\\r')
            i += 1
            continue
        if c == '\t':
            out.append('\\t')
            i += 1
            continue
        o = ord(c)
        if o < 32:
            out.append(f'\\u{o:04x}')
            i += 1
            continue
        out.append(c)
        i += 1
    return ''.join(out)


def _parse_llm_json_object(raw: str, *, context: str = 'llm') -> tuple[dict, str | None]:
    """
    Parse JSON from an LLM message: tolerate markdown fences, leading prose,
    illegal control characters inside strings, and minor syntax issues (via json-repair).
    Returns (dict, None) on success, or ({}, error_message) on total failure.
    """
    s = (raw or '').strip()
    if not s:
        return {}, None
    s = _strip_llm_markdown_fence(s).lstrip('\ufeff').strip()

    seen: set[str] = set()
    fragments: list[str] = []
    for frag in (s, _extract_first_json_object(s)):
        frag = frag.strip()
        if frag and frag not in seen:
            seen.add(frag)
            fragments.append(frag)
    for frag in list(fragments):
        esc = _escape_raw_controls_in_json_strings(frag)
        if esc != frag and esc not in seen:
            seen.add(esc)
            fragments.append(esc)

    last_err: str | None = None
    for frag in fragments:
        try:
            obj = json.loads(frag)
            if isinstance(obj, dict):
                return obj, None
        except json.JSONDecodeError as e:
            last_err = str(e)

    try:
        from json_repair import repair_json  # type: ignore[import-untyped]
    except ImportError:
        repair_json = None

    if repair_json:
        for frag in fragments:
            try:
                obj = repair_json(frag, return_objects=True)
                if isinstance(obj, dict):
                    return obj, None
            except Exception as e:
                last_err = str(e)

    logger.warning('%s: could not parse model JSON (%s)', context, last_err)
    return {}, last_err or 'invalid JSON'


def _normalize_tabular_header(cell: str) -> str:
    return re.sub(r'[\s_]+', ' ', (cell or '').strip().lower())


def _tabular_header_fingerprint(header_row: list[str]) -> str:
    return '\x1f'.join(_normalize_tabular_header(h) for h in header_row)


# Process-local cache: fingerprint -> {'columns': dict[str,int], 'amount_sign': str, 'source': str}
_TABULAR_HEADER_MAP_CACHE: dict[str, dict] = {}


def _tabular_header_aliases() -> dict[str, tuple[str, ...]]:
    """Optional zero-cost fast path only — unknown banks use a tiny header LLM call."""
    return {
        'description': (
            'description', 'details', 'narrative', 'merchant', 'reference',
            'counterparty', 'payee', 'transaction description', 'name',
        ),
        'amount': (
            'amount', 'value', 'sum', 'transaction amount', 'txn amount',
        ),
        'money_in': (
            'money in', 'paid in', 'credit', 'credits', 'inflow', 'deposit',
            'money in (£)', 'paid in (£)',
        ),
        'money_out': (
            'money out', 'paid out', 'debit', 'debits', 'outflow', 'withdrawal',
            'money out (£)', 'paid out (£)',
        ),
        'date': (
            'date', 'booking date', 'transaction date', 'value date',
            'posting date', 'txn date', 'book date',
        ),
        'started_date': (
            'started date', 'start date', 'started', 'created date',
        ),
        'completed_date': (
            'completed date', 'completion date', 'completed', 'settled date',
            'settled', 'finished date',
        ),
        'state': ('state', 'status'),
        'direction': ('direction', 'flow', 'credit/debit', 'dr/cr'),
        'type': ('type', 'transaction type', 'product type'),
    }


def _validate_tabular_column_map(col: dict[str, int], header_len: int) -> dict[str, int] | None:
    cleaned: dict[str, int] = {}
    for field, idx in col.items():
        if field not in {
            'description', 'amount', 'money_in', 'money_out', 'date',
            'started_date', 'completed_date', 'state', 'direction', 'type',
        }:
            continue
        try:
            i = int(idx)
        except (TypeError, ValueError):
            continue
        if i < 0 or i >= header_len:
            continue
        cleaned[field] = i
    has_desc = 'description' in cleaned
    has_amount = 'amount' in cleaned or ('money_in' in cleaned and 'money_out' in cleaned)
    has_date = 'date' in cleaned or 'started_date' in cleaned or 'completed_date' in cleaned
    if not (has_desc and has_amount and has_date):
        return None
    return cleaned


def _map_tabular_headers_alias(header_row: list[str]) -> dict[str, int] | None:
    """Deterministic alias match (fast path). Returns None when headers are unfamiliar."""
    aliases = _tabular_header_aliases()
    normalized = [_normalize_tabular_header(h) for h in header_row]
    col: dict[str, int] = {}
    for idx, h in enumerate(normalized):
        if not h:
            continue
        for field, names in aliases.items():
            if field in col:
                continue
            if h in names:
                col[field] = idx
                break
    return _validate_tabular_column_map(col, len(header_row))


def _normalize_amount_sign(raw) -> str:
    s = str(raw or '').strip().lower().replace('-', '_').replace(' ', '_')
    if s in {
        'negative_is_outgoing', 'negative_outgoing', 'signed_negative_out',
        'revolut', 'accounting',
    }:
        return 'negative_is_outgoing'
    if s in {
        'positive_is_outgoing', 'positive_outgoing', 'signed_positive_out',
    }:
        return 'positive_is_outgoing'
    if s in {'absolute', 'unsigned', 'abs', 'money_columns'}:
        return 'absolute'
    return 'negative_is_outgoing'


def _llm_map_tabular_headers(header_row: list[str]) -> dict | None:
    """
    Tiny LLM call: header titles only → column indices for local row parsing.
    Returns {'columns': dict[str,int], 'amount_sign': str} or None.
    """
    client = _get_openai_client()
    if not client:
        return None
    headers = [(h if h is not None else '') for h in header_row]
    if sum(1 for h in headers if str(h).strip()) < 3:
        return None

    indexed = [{'index': i, 'title': str(h)} for i, h in enumerate(headers)]
    model = os.getenv('OPENAI_MODEL', 'gpt-4o-mini').strip() or 'gpt-4o-mini'
    system = (
        'You map bank CSV/Excel column headers to canonical fields for transaction import. '
        'Return only valid JSON: '
        '{"is_transaction_table":boolean,'
        '"columns":{"description":int|null,"amount":int|null,"money_in":int|null,'
        '"money_out":int|null,"date":int|null,"started_date":int|null,'
        '"completed_date":int|null,"state":int|null,"direction":int|null},'
        '"amount_sign":"negative_is_outgoing"|"positive_is_outgoing"|"absolute"}. '
        'Use the provided header index integers (0-based). Set unused fields to null. '
        'Rules: (1) description = merchant/payee/narrative/details. '
        '(2) Prefer separate money_in/money_out (Paid in/Paid out) when both exist; else use amount. '
        '(3) date = single booking/value date; started_date/completed_date when both start and settle exist. '
        '(4) amount_sign: negative_is_outgoing when negatives are debits (e.g. Revolut); '
        'positive_is_outgoing when positives are debits; absolute when amounts are unsigned '
        'or when money_in/money_out columns are used. '
        '(5) is_transaction_table=false for non-transaction sheets (balances-only, metadata).'
    )
    user_msg = 'Header columns:\n' + json.dumps(indexed, ensure_ascii=False)
    try:
        completion = client.chat.completions.create(
            model=model,
            messages=[
                {'role': 'system', 'content': system},
                {'role': 'user', 'content': user_msg},
            ],
            response_format={'type': 'json_object'},
            temperature=0,
        )
    except Exception as e:
        logger.warning('tabular header LLM mapping failed: %s', e)
        return None
    raw = completion.choices[0].message.content or '{}'
    data, jerr = _parse_llm_json_object(raw, context='tabular_header_map')
    if jerr or not isinstance(data, dict):
        logger.warning('tabular header LLM JSON failed: %s', jerr)
        return None
    if data.get('is_transaction_table') is False:
        return None
    cols_raw = data.get('columns')
    if not isinstance(cols_raw, dict):
        return None
    # Drop nulls; keep int indices.
    tentative: dict[str, int] = {}
    for field, val in cols_raw.items():
        if val is None or val == '':
            continue
        tentative[str(field)] = val
    cleaned = _validate_tabular_column_map(tentative, len(headers))
    if cleaned is None:
        return None
    return {
        'columns': cleaned,
        'amount_sign': _normalize_amount_sign(data.get('amount_sign')),
    }


def _resolve_tabular_headers(header_row: list[str], *, allow_llm: bool = True) -> dict | None:
    """
    Resolve header titles → column map.
    Prefer process cache, then deterministic aliases (no API), then a tiny header-only LLM call.
    Returns {'columns', 'amount_sign', 'source'} or None.
    """
    fp = _tabular_header_fingerprint(header_row)
    cached = _TABULAR_HEADER_MAP_CACHE.get(fp)
    if isinstance(cached, dict) and isinstance(cached.get('columns'), dict):
        return dict(cached)

    alias = _map_tabular_headers_alias(header_row)
    if alias is not None:
        # Signed amount exports (Revolut-like) use negative=outgoing by convention.
        amount_sign = 'negative_is_outgoing' if 'amount' in alias else 'absolute'
        resolved = {'columns': alias, 'amount_sign': amount_sign, 'source': 'alias'}
        _TABULAR_HEADER_MAP_CACHE[fp] = resolved
        return dict(resolved)

    if not allow_llm:
        return None

    llm = _llm_map_tabular_headers(header_row)
    if llm is None:
        return None
    resolved = {
        'columns': llm['columns'],
        'amount_sign': llm.get('amount_sign') or 'negative_is_outgoing',
        'source': 'llm_headers',
    }
    _TABULAR_HEADER_MAP_CACHE[fp] = resolved
    return dict(resolved)


def _parse_signed_amount_cell(val) -> float | None:
    if val is None or val == '':
        return None
    if isinstance(val, (int, float)) and not isinstance(val, bool):
        return float(val)
    s = str(val).strip().replace(',', '').replace('\u00a0', '').replace(' ', '')
    s = re.sub(r'^[$£€]', '', s)
    if not s or s in {'.', '-', '+', '(', ')'}:
        return None
    if s.startswith('(') and s.endswith(')'):
        s = '-' + s[1:-1]
    try:
        return float(s)
    except ValueError:
        return None


def _row_looks_numeric_heavy(row: list[str]) -> bool:
    """True when most cells look like amounts — unlikely to be a header row."""
    non_empty = [c for c in row if str(c or '').strip()]
    if len(non_empty) < 3:
        return True
    numeric = 0
    for c in non_empty:
        if _parse_signed_amount_cell(c) is not None:
            numeric += 1
            continue
        # ISO-ish dates alone don't make a header numeric-heavy.
        if re.match(r'^\d{4}-\d{2}-\d{2}', str(c).strip()):
            numeric += 1
    return numeric >= max(2, (len(non_empty) + 1) // 2)


def _direction_from_cell(raw: str) -> str | None:
    d = _normalize_tabular_header(raw).replace('/', ' ')
    if not d:
        return None
    if d in {'in', 'incoming', 'credit', 'cr', 'deposit', 'money in', 'paid in'}:
        return 'incoming'
    if d in {'out', 'outgoing', 'debit', 'dr', 'withdrawal', 'money out', 'paid out'}:
        return 'outgoing'
    if 'credit' in d or d.endswith(' cr'):
        return 'incoming'
    if 'debit' in d or d.endswith(' dr'):
        return 'outgoing'
    return None


def _tabular_row_to_raw_tx(
    cells: list[str],
    col: dict[str, int],
    *,
    amount_sign: str = 'negative_is_outgoing',
) -> dict | None:
    def cell(field: str) -> str:
        i = col.get(field)
        if i is None or i >= len(cells):
            return ''
        return (cells[i] or '').strip()

    state = cell('state')
    if state:
        st = _normalize_tabular_header(state).replace(' ', '_')
        if st not in _SPENDING_TABULAR_KEEP_STATES:
            return None

    desc = cell('description')
    if not desc:
        return None

    direction = None
    amount = None
    money_in = None
    money_out = None

    if 'amount' in col:
        signed = _parse_signed_amount_cell(cell('amount'))
        if signed is None or signed == 0:
            return None
        dir_cell = _direction_from_cell(cell('direction'))
        sign = _normalize_amount_sign(amount_sign)
        if dir_cell is not None and sign == 'absolute':
            direction = dir_cell
            amount = abs(signed)
        elif sign == 'positive_is_outgoing':
            if signed > 0:
                direction, amount = 'outgoing', signed
            else:
                direction, amount = 'incoming', abs(signed)
        else:
            # negative_is_outgoing (default) and absolute without direction column
            if signed < 0:
                direction, amount = 'outgoing', abs(signed)
            else:
                direction, amount = 'incoming', signed
            if dir_cell is not None:
                direction = dir_cell
                amount = abs(signed)
    else:
        money_in = _spending_optional_money(cell('money_in'))
        money_out = _spending_optional_money(cell('money_out'))
        if money_in is not None and money_out is None:
            direction, amount = 'incoming', money_in
        elif money_out is not None and money_in is None:
            direction, amount = 'outgoing', money_out
        else:
            return None

    started_raw = cell('started_date')
    completed_raw = cell('completed_date')
    date_raw = cell('date')
    # Prefer completed as ledger date when both exist (Revolut).
    if completed_raw:
        primary_raw = completed_raw
    elif date_raw:
        primary_raw = date_raw
    else:
        primary_raw = started_raw
    if not primary_raw:
        return None

    row = {
        'date': primary_raw,
        'description': desc,
        'amount': amount,
        'direction': direction,
    }
    if started_raw:
        row['started_date'] = started_raw
    if completed_raw:
        row['completed_date'] = completed_raw
    if money_in is not None:
        row['money_in'] = money_in
    if money_out is not None:
        row['money_out'] = money_out
    return row


def _iter_csv_like_rows(statement_text: str):
    """Yield lists of cell strings from CSV-like statement text (Excel convert / CSV)."""
    # Drop multi-sheet banners from _extract_spreadsheet_text.
    cleaned_lines: list[str] = []
    for line in (statement_text or '').splitlines():
        if line.startswith('--- Sheet:'):
            continue
        cleaned_lines.append(line)
    cleaned = '\n'.join(cleaned_lines).strip()
    if not cleaned:
        return
    try:
        reader = csv.reader(io.StringIO(cleaned))
        for row in reader:
            if row is None:
                continue
            yield [(c if c is not None else '') for c in row]
    except csv.Error:
        return


def _split_tabular_statement(statement_text: str) -> tuple[list[str], list[list[str]]] | None:
    """
    Split CSV-like text into (header_row, data_rows) when it looks like a table export.
    Does not map columns — that happens via alias or a header-only LLM call.
    """
    header = None
    data_rows: list[list[str]] = []
    for row in _iter_csv_like_rows(statement_text):
        if not any((c or '').strip() for c in row):
            continue
        if header is None:
            if _row_looks_numeric_heavy(row):
                # First row already looks like data — not a confident tabular export.
                return None
            header = row
            continue
        data_rows.append(row)
    if header is None or not data_rows:
        return None
    if sum(1 for c in header if str(c or '').strip()) < 3:
        return None
    return header, data_rows


def _parse_tabular_data_rows(
    header: list[str],
    data_rows: list[list[str]],
    col: dict[str, int],
    *,
    amount_sign: str = 'negative_is_outgoing',
) -> tuple[list[dict], dict]:
    header_fp = _tabular_header_fingerprint(header)
    raw: list[dict] = []
    skipped_state = 0
    skipped_other = 0
    for cells in data_rows:
        # Skip repeated header rows (multi-sheet concat).
        if _tabular_header_fingerprint(cells) == header_fp:
            continue
        if not _row_looks_numeric_heavy(cells) and _map_tabular_headers_alias(cells) is not None:
            continue
        tx = _tabular_row_to_raw_tx(cells, col, amount_sign=amount_sign)
        if tx is None:
            state_i = col.get('state')
            if state_i is not None and state_i < len(cells) and (cells[state_i] or '').strip():
                st = _normalize_tabular_header(cells[state_i]).replace(' ', '_')
                if st and st not in _SPENDING_TABULAR_KEEP_STATES:
                    skipped_state += 1
                    continue
            skipped_other += 1
            continue
        raw.append(tx)
    stats = {'skipped_state': skipped_state, 'skipped_other': skipped_other}
    return raw, stats


def _try_parse_tabular_spending_transactions(
    statement_text: str,
    *,
    allow_llm: bool = True,
) -> tuple[list, dict] | None:
    """
    Parse bank CSV / Excel exports locally after mapping headers.
    Header mapping: alias fast path, else tiny LLM call on header titles only (not row data).
    Returns (raw_rows, meta) or None when not a confident tabular export.
    """
    split = _split_tabular_statement(statement_text)
    if split is None:
        return None
    header, data_rows = split
    resolved = _resolve_tabular_headers(header, allow_llm=allow_llm)
    if resolved is None:
        return None
    col = resolved['columns']
    amount_sign = resolved.get('amount_sign') or 'negative_is_outgoing'
    raw, stats = _parse_tabular_data_rows(header, data_rows, col, amount_sign=amount_sign)
    if len(raw) < SPENDING_TABULAR_MIN_ROWS:
        return None

    profile = 'revolut_like' if (
        'started_date' in col and 'completed_date' in col and 'amount' in col
    ) else 'signed_amount' if 'amount' in col else 'money_columns'

    meta = {
        'mode': 'tabular',
        'reason': 'structured_csv_headers',
        'profile': profile,
        'header': [_normalize_tabular_header(c) for c in header],
        'columns': sorted(col.keys()),
        'header_map_source': resolved.get('source'),
        'amount_sign': amount_sign,
        'row_count': len(raw),
        'skipped_state': stats['skipped_state'],
        'skipped_other': stats['skipped_other'],
    }
    return raw, meta


def _iter_tabular_spending_extraction(statement_text: str):
    """
    Yield progress for header mapping, then {'type':'result',...} or nothing if not tabular.
    """
    split = _split_tabular_statement(statement_text)
    if split is None:
        return
    header, data_rows = split

    # Alias / cache first (no progress needed).
    fp = _tabular_header_fingerprint(header)
    cached = _TABULAR_HEADER_MAP_CACHE.get(fp)
    if isinstance(cached, dict) and isinstance(cached.get('columns'), dict):
        resolved = dict(cached)
    else:
        alias = _map_tabular_headers_alias(header)
        if alias is not None:
            amount_sign = 'negative_is_outgoing' if 'amount' in alias else 'absolute'
            resolved = {'columns': alias, 'amount_sign': amount_sign, 'source': 'alias'}
            _TABULAR_HEADER_MAP_CACHE[fp] = resolved
        else:
            yield {
                'type': 'progress',
                'step': 'tabular_headers',
                'message': 'Mapping spreadsheet/CSV column headers with a small model call…',
            }
            llm = _llm_map_tabular_headers(header)
            if llm is None:
                return
            resolved = {
                'columns': llm['columns'],
                'amount_sign': llm.get('amount_sign') or 'negative_is_outgoing',
                'source': 'llm_headers',
            }
            _TABULAR_HEADER_MAP_CACHE[fp] = resolved

    col = resolved['columns']
    amount_sign = resolved.get('amount_sign') or 'negative_is_outgoing'
    raw, stats = _parse_tabular_data_rows(header, data_rows, col, amount_sign=amount_sign)
    if len(raw) < SPENDING_TABULAR_MIN_ROWS:
        return

    profile = 'revolut_like' if (
        'started_date' in col and 'completed_date' in col and 'amount' in col
    ) else 'signed_amount' if 'amount' in col else 'money_columns'

    yield {
        'type': 'progress',
        'step': 'tabular_parse',
        'message': (
            f'Parsed {len(raw)} transactions from spreadsheet/CSV locally '
            f'({profile}, headers via {resolved.get("source")})…'
        ),
    }
    yield {
        'type': 'result',
        'rows': raw,
        'meta': {
            'mode': 'tabular',
            'reason': 'structured_csv_headers',
            'profile': profile,
            'header': [_normalize_tabular_header(c) for c in header],
            'columns': sorted(col.keys()),
            'header_map_source': resolved.get('source'),
            'amount_sign': amount_sign,
            'row_count': len(raw),
            'skipped_state': stats['skipped_state'],
            'skipped_other': stats['skipped_other'],
        },
    }

def _split_statement_text_into_llm_chunks(text: str, max_chars: int = STATEMENT_LLM_CHUNK_CHARS) -> list[str]:
    """Split long statement text into line-aligned chunks for sequential LLM extraction."""
    t = text or ''
    if len(t) <= max_chars:
        return [t] if t else []
    lines = t.splitlines(keepends=True)
    chunks: list[str] = []
    buf: list[str] = []
    size = 0
    for line in lines:
        line_len = len(line)
        if buf and size + line_len > max_chars:
            chunks.append(''.join(buf))
            buf = [line]
            size = line_len
        else:
            buf.append(line)
            size += line_len
    if buf:
        chunks.append(''.join(buf))
    return chunks


def _extract_spending_transactions_llm_chunk(
    client,
    *,
    model: str,
    system: str,
    chunk: str,
    period_hint: str | None,
    part_label: str | None,
    note_truncated: bool,
    extraction_timeout: float,
) -> list:
    user_msg = 'Bank statement text:\n\n' + chunk
    if part_label:
        user_msg += (
            f'\n\n({part_label} '
            'Extract every transaction in this part only; do not invent rows from other parts.)'
        )
    if period_hint:
        user_msg += '\n\n' + period_hint
    if note_truncated:
        user_msg += '\n\n(Note: text was truncated.)'

    try:
        completion = client.with_options(
            timeout=extraction_timeout
        ).chat.completions.create(
            model=model,
            messages=[
                {'role': 'system', 'content': system},
                {'role': 'user', 'content': user_msg},
            ],
            response_format={'type': 'json_object'},
            temperature=0.1,
        )
    except APITimeoutError as e:
        raise RuntimeError(
            'The statement was too large to process within the time limit. '
            'Try a CSV or Excel (.xlsx) export from your bank (structured exports parse instantly), '
            'or narrow the date range and import in smaller chunks.'
        ) from e
    raw = completion.choices[0].message.content or '{}'
    data, jerr = _parse_llm_json_object(raw, context='spending_transactions')
    if jerr:
        raise RuntimeError(
            'Could not parse the model response as JSON. This sometimes happens with very long statements '
            'or stray characters in a description. Try a CSV or Excel (.xlsx) export, narrow the date range, or retry. '
            f'Detail: {jerr}'
        )
    txs = data.get('transactions')
    if txs is None:
        txs = data.get('items') or []
    return txs if isinstance(txs, list) else []


def _spending_extraction_system_prompt() -> str:
    return (
        'You extract ALL individual bank transactions from the statement text into structured JSON. '
        'Return only valid JSON as {"transactions":[{"date":"YYYY-MM-DD","started_date":"YYYY-MM-DD"|null,'
        '"completed_date":"YYYY-MM-DD"|null,"description":"string","amount":number,'
        '"direction":"incoming|outgoing","money_in":number|null,"money_out":number|null}]}. '
        'Rules: (1) amount is always a positive absolute number (same as the money in or money out value for that row). '
        '(2) Many statements use TWO columns: Paid in / Money in / Credits vs Paid out / Money out / Debits. '
        'For those rows, set money_in to the amount in the IN column and money_out to the amount in the OUT column (use null for the empty column). '
        'Direction MUST follow the column: if money_in is set, direction is "incoming"; if money_out is set, direction is "outgoing". '
        'Do NOT infer direction from the merchant description when column hints or two-column layout is present. '
        '(3) If the statement does not use two columns, set money_in and money_out to null and use amount + direction as usual. '
        '(4) Be exhaustive: include every posted transaction line. Do not merge rows; do not omit small items. '
        '(5) Skip only running balances, headers, page numbers, and non-transaction summaries. '
        '(6) Use ISO dates YYYY-MM-DD. If the statement shows DD/MM/YYYY, convert to ISO. '
        '(7) Lines beginning with LINE_HINT are layout hints — use them to fill money_in/money_out and direction correctly. '
        '(8) When a statement has both a start/started date and a completed/settled date (e.g. Revolut '
        '"Started Date" and "Completed Date"), set started_date and completed_date accordingly, and set '
        '"date" to the completed/settled date. If only one date exists, set date to that value and leave '
        'started_date/completed_date null.'
    )


def _prepare_spending_llm_chunks(statement_text: str) -> tuple[list[str], bool]:
    text = statement_text or ''
    truncated = False
    if len(text) > STATEMENT_MAX_CHARS_FOR_LLM:
        text = text[:STATEMENT_MAX_CHARS_FOR_LLM]
        truncated = True
    return _split_statement_text_into_llm_chunks(text), truncated


def _extract_spending_transactions_llm(statement_text: str, period_hint: str | None = None) -> list:
    rows = []
    for ev in _iter_extract_spending_transactions_llm(statement_text, period_hint):
        if ev.get('type') == 'rows':
            rows = ev.get('rows') or []
    return rows


def _iter_extract_spending_transactions_llm(statement_text: str, period_hint: str | None = None):
    """
    Yield progress between chunks so streaming preview / gunicorn see activity.
    Final yield: {'type':'rows','rows':list,'chunk_count':int}.
    """
    client = _get_openai_client()
    if not client:
        raise RuntimeError('OPENAI_API_KEY is not set')

    model = os.getenv('OPENAI_MODEL', 'gpt-4o-mini').strip() or 'gpt-4o-mini'
    chunks, truncated = _prepare_spending_llm_chunks(statement_text)
    if not chunks:
        yield {'type': 'rows', 'rows': [], 'chunk_count': 0}
        return

    system = _spending_extraction_system_prompt()
    extraction_timeout = _openai_extraction_timeout_seconds()
    all_txs: list = []
    n = len(chunks)
    for idx, chunk in enumerate(chunks):
        if n > 1:
            yield {
                'type': 'progress',
                'step': 'llm_extract',
                'message': f'Extracting transactions with the model (chunk {idx + 1} of {n})…',
            }
        part_label = (
            f'This is part {idx + 1} of {n} of a longer statement.'
            if n > 1
            else None
        )
        all_txs.extend(
            _extract_spending_transactions_llm_chunk(
                client,
                model=model,
                system=system,
                chunk=chunk,
                period_hint=period_hint,
                part_label=part_label,
                note_truncated=truncated and idx == n - 1,
                extraction_timeout=extraction_timeout,
            )
        )
    yield {'type': 'rows', 'rows': all_txs, 'chunk_count': n}


def _split_spending_statement_text_for_model(text: str) -> tuple[str, str]:
    """Split model input into plain extracted text and LINE_HINT block (if present)."""
    banner = SPENDING_PDF_COLUMN_HINTS_BANNER
    if banner in text:
        base, _, hints = text.partition(banner)
        return base.rstrip(), hints.lstrip()
    return text, ''


def _parse_line_hint_rows(hints_block: str) -> list[dict]:
    rows: list[dict] = []
    for line in (hints_block or '').splitlines():
        line = line.strip()
        if not line.startswith('LINE_HINT'):
            continue
        m = _LINE_HINT_ROW_RE.match(line)
        if not m:
            continue
        try:
            mi = round(float(m.group(1)), 2)
            mo = round(float(m.group(2)), 2)
        except ValueError:
            continue
        lt = (m.group(3) or '').strip()
        if not lt:
            continue
        rows.append({'money_in': mi, 'money_out': mo, 'line_text': lt})
    return rows


def _compile_spending_skip_line_regexes(patterns: list[str]) -> list[re.Pattern]:
    compiled: list[re.Pattern] = []
    for p in patterns:
        s = str(p).strip()
        if not s:
            continue
        try:
            compiled.append(re.compile(s))
        except re.error:
            logger.warning('skipping invalid layout skip regex: %s', s[:120])
    return compiled


def _merge_spending_layout_skip_regexes(llm_patterns: list | None) -> list[str]:
    merged: list[str] = []
    seen: set[str] = set()
    for p in _SPENDING_LAYOUT_DEFAULT_SKIP_REGEXES:
        if p not in seen:
            seen.add(p)
            merged.append(p)
    if isinstance(llm_patterns, list):
        for p in llm_patterns:
            s = str(p).strip()
            if s and s not in seen and len(merged) < 28:
                seen.add(s)
                merged.append(s)
    return merged


def _llm_spending_layout_skip_regexes(base_text: str, hint_rows: list[dict]) -> list[str]:
    """Small model call: extra regexes for non-transaction lines (balances, headers)."""
    client = _get_openai_client()
    if not client:
        return _merge_spending_layout_skip_regexes([])

    lines = base_text.splitlines()[:SPENDING_LAYOUT_EXCERPT_MAX_LINES]
    excerpt = '\n'.join(lines)
    if len(excerpt) > SPENDING_LAYOUT_EXCERPT_MAX_CHARS:
        excerpt = excerpt[:SPENDING_LAYOUT_EXCERPT_MAX_CHARS]

    samples = []
    for h in hint_rows[:SPENDING_LAYOUT_SAMPLE_HINTS]:
        samples.append(
            f'LINE_HINT money_in={h["money_in"]} money_out={h["money_out"]} | {h["line_text"]}'
        )
    sample_block = '\n'.join(samples)
    user_body = f'Statement excerpt:\n{excerpt}\n\nSample LINE_HINT rows:\n{sample_block}\n'
    if len(user_body) > SPENDING_LAYOUT_MAPPING_LLM_MAX_CHARS:
        user_body = user_body[:SPENDING_LAYOUT_MAPPING_LLM_MAX_CHARS]

    model = os.getenv('OPENAI_MODEL', 'gpt-4o-mini').strip() or 'gpt-4o-mini'
    system = (
        'You help parse UK/Irish bank statements. The app already knows debit/credit amounts per row from PDF '
        'geometry (LINE_HINT). Your ONLY job is to list extra Python regex patterns for lines that are NOT real '
        'transactions: running balances, section headers, page markers, "brought forward", etc. '
        'Return only valid JSON: {"skip_line_regexes":["regex",...]}. '
        'Each regex is tested against the text AFTER " | " in a LINE_HINT row (the bank line, not the prefix). '
        'Use (?i) for case-insensitive patterns. Prefer anchored patterns (^...) when sensible. '
        'Be conservative: do NOT add broad patterns that could match merchant or payment descriptions '
        '(that would drop real transactions). When unsure, omit the pattern. '
        'Return an empty array if defaults suffice. At most 14 patterns; each under 180 characters.'
    )
    try:
        completion = client.chat.completions.create(
            model=model,
            messages=[
                {'role': 'system', 'content': system},
                {'role': 'user', 'content': user_body},
            ],
            response_format={'type': 'json_object'},
            temperature=0.1,
        )
        raw = completion.choices[0].message.content or '{}'
        data, jerr = _parse_llm_json_object(raw, context='spending_layout_skip')
        if jerr:
            logger.warning('layout skip LLM JSON failed: %s', jerr)
            return _merge_spending_layout_skip_regexes([])
        return _merge_spending_layout_skip_regexes(data.get('skip_line_regexes'))
    except Exception as e:
        logger.warning('layout skip LLM call failed: %s', e)
        return _merge_spending_layout_skip_regexes([])


def _extract_first_bank_date_from_line(line_text: str):
    """Find the first plausible transaction date and return (date, description_remainder)."""
    line_text = (line_text or '').strip()
    if not line_text:
        return None, ''
    tokens = line_text.split()
    for n in range(1, min(5, len(tokens) + 1)):
        cand = ' '.join(tokens[:n])
        d = _parse_bank_statement_date(cand)
        if d is not None:
            rest = ' '.join(tokens[n:]).strip()
            return d, rest
    for m in re.finditer(r'\b(\d{1,2})[/.-](\d{1,2})[/.-](\d{2,4})\b', line_text):
        sub = m.group(0)
        d = _parse_bank_statement_date(sub)
        if d is not None:
            rest = (line_text[: m.start()] + ' ' + line_text[m.end() :]).strip()
            rest = re.sub(r'\s+', ' ', rest)
            return d, rest
    for m in re.finditer(r'\b(\d{4})-(\d{2})-(\d{2})\b', line_text):
        sub = m.group(0)
        d = _parse_bank_statement_date(sub)
        if d is not None:
            rest = (line_text[: m.start()] + ' ' + line_text[m.end() :]).strip()
            rest = re.sub(r'\s+', ' ', rest)
            return d, rest
    return None, line_text


def _token_is_strict_statement_amount(tok: str) -> bool:
    """True if token looks like a posted currency amount (optional £/$/€, optional commas, .pp)."""
    t = (tok or '').strip()
    t = re.sub(r'^[$£€]\s*', '', t)
    t = t.replace(',', '')
    t = re.sub(r'(?:CR|DR)$', '', t, flags=re.I)
    return bool(re.match(r'^\d+\.\d{2}$', t))


def _description_before_amount_tokens(desc: str) -> str:
    """
    Keep payee/description text after the date, stopping at the first amount-like token.

    Statement rows are usually [date] [description…] [transaction amount(s)] [running balance].
    LINE_HINT already has the debit/credit amounts; the line text repeats them plus balance.
    A left boundary (first money token) matches that layout more reliably than stripping only
    from the right (e.g. odd trailing tokens, or future extra columns).
    """
    if not (desc or '').strip():
        return desc
    parts = desc.split()
    out: list[str] = []
    for tok in parts:
        if _token_is_strict_statement_amount(tok):
            break
        out.append(tok)
    return ' '.join(out).strip()


def _layout_eligible_hint_count(hint_rows: list[dict], skip_pattern_strings: list[str]) -> int:
    """Hint rows that have money and are not skipped — used for fallback yield thresholds."""
    compiled = _compile_spending_skip_line_regexes(skip_pattern_strings)
    n = 0
    for h in hint_rows:
        lt = h.get('line_text') or ''
        if any(p.search(lt) for p in compiled):
            continue
        try:
            mi = round(float(h.get('money_in', 0) or 0), 2)
            mo = round(float(h.get('money_out', 0) or 0), 2)
        except (TypeError, ValueError):
            continue
        if mi <= 0 and mo <= 0:
            continue
        n += 1
    return n


def _raw_transactions_from_line_hints(
    hint_rows: list[dict],
    skip_pattern_strings: list[str],
    *,
    carry_forward_date: bool = True,
) -> list[dict]:
    """Build raw transaction dicts compatible with _normalize_spending_transactions."""
    compiled = _compile_spending_skip_line_regexes(skip_pattern_strings)
    raw: list[dict] = []
    last_d: date | None = None
    for h in hint_rows:
        lt = h.get('line_text') or ''
        if any(p.search(lt) for p in compiled):
            continue
        try:
            mi = round(float(h.get('money_in', 0) or 0), 2)
            mo = round(float(h.get('money_out', 0) or 0), 2)
        except (TypeError, ValueError):
            continue
        if mi <= 0 and mo <= 0:
            continue
        explicit_d, desc = _extract_first_bank_date_from_line(lt)
        if explicit_d is not None:
            d = explicit_d
            last_d = explicit_d
        elif carry_forward_date and last_d is not None:
            d = last_d
            desc = lt.strip()
        else:
            continue
        if not desc:
            desc = 'Bank transaction'
        else:
            desc = _description_before_amount_tokens(desc)
        if not desc:
            desc = 'Bank transaction'
        ds = d.strftime('%Y-%m-%d')
        if mi > 0 and mo > 0:
            raw.append({
                'date': ds,
                'description': desc[:500],
                'money_in': mi,
                'money_out': None,
                'direction': 'incoming',
            })
            raw.append({
                'date': ds,
                'description': desc[:500],
                'money_in': None,
                'money_out': mo,
                'direction': 'outgoing',
            })
        elif mi > 0:
            raw.append({
                'date': ds,
                'description': desc[:500],
                'money_in': mi,
                'money_out': None,
                'direction': 'incoming',
            })
        else:
            raw.append({
                'date': ds,
                'description': desc[:500],
                'money_in': None,
                'money_out': mo,
                'direction': 'outgoing',
            })
    return raw


def _layout_extraction_should_fallback(hint_row_count: int, eligible_count: int, raw_rows: list) -> bool:
    if hint_row_count < SPENDING_LAYOUT_MIN_HINT_ROWS:
        return True
    p = len(raw_rows)
    if p == 0:
        return True
    if eligible_count <= 0:
        return True
    need = max(
        SPENDING_LAYOUT_FALLBACK_MIN_PARSED,
        int(eligible_count * SPENDING_LAYOUT_FALLBACK_RATIO),
    )
    # At least one output per eligible line when lines are single-sided; dual in/out yields 2 per line.
    need = min(need, eligible_count)
    return p < need


def iter_spending_transaction_extraction(statement_text: str, period_hint: str | None = None):
    """
    Yields {'type':'progress','step':str,'message':str} then
    {'type':'result','rows':list,'meta':dict}.
    """
    # Fast path: structured CSV / Excel — map headers (alias or tiny LLM), parse rows locally.
    tabular_result = None
    for ev in _iter_tabular_spending_extraction(statement_text):
        if ev.get('type') == 'progress':
            yield ev
        elif ev.get('type') == 'result':
            tabular_result = ev
    if tabular_result is not None:
        yield tabular_result
        return

    base, hints_block = _split_spending_statement_text_for_model(statement_text)
    hint_rows = _parse_line_hint_rows(hints_block)

    if len(hint_rows) < SPENDING_LAYOUT_MIN_HINT_ROWS:
        yield {
            'type': 'progress',
            'step': 'llm_extract',
            'message': 'Extracting transactions with the model (not enough layout-derived LINE_HINT rows)…',
        }
        rows = None
        chunk_count = 1
        for ev in _iter_extract_spending_transactions_llm(statement_text, period_hint):
            if ev.get('type') == 'progress':
                yield ev
            elif ev.get('type') == 'rows':
                rows = ev.get('rows') or []
                chunk_count = int(ev.get('chunk_count') or 1)
        yield {
            'type': 'result',
            'rows': rows or [],
            'meta': {
                'mode': 'llm_full',
                'reason': 'insufficient_layout_hints',
                'hint_row_count': len(hint_rows),
                'llm_chunk_count': chunk_count,
            },
        }
        return

    yield {
        'type': 'progress',
        'step': 'layout_mapping',
        'message': 'Inferring which lines to skip (compact model request on a short excerpt)…',
    }
    skip_patterns = _llm_spending_layout_skip_regexes(base, hint_rows)

    yield {
        'type': 'progress',
        'step': 'layout_parse',
        'message': 'Parsing dates and descriptions from LINE_HINT rows locally…',
    }
    rows = _raw_transactions_from_line_hints(hint_rows, skip_patterns)
    eligible = _layout_eligible_hint_count(hint_rows, skip_patterns)
    layout_retry_defaults = False
    if _layout_extraction_should_fallback(len(hint_rows), eligible, rows):
        yield {
            'type': 'progress',
            'step': 'layout_parse',
            'message': 'Retrying layout parse with default skip patterns only (model-added filters may have been too tight)…',
        }
        defaults_only = _merge_spending_layout_skip_regexes([])
        rows_retry = _raw_transactions_from_line_hints(hint_rows, defaults_only)
        elig_retry = _layout_eligible_hint_count(hint_rows, defaults_only)
        if len(rows_retry) > len(rows) or not _layout_extraction_should_fallback(
            len(hint_rows), elig_retry, rows_retry
        ):
            rows = rows_retry
            eligible = elig_retry
            skip_patterns = defaults_only
            layout_retry_defaults = True

    layout_attempt_count = len(rows)

    if _layout_extraction_should_fallback(len(hint_rows), eligible, rows):
        yield {
            'type': 'progress',
            'step': 'llm_extract',
            'message': 'Layout parse still thin after retries — falling back to full model extraction…',
        }
        rows = None
        chunk_count = 1
        for ev in _iter_extract_spending_transactions_llm(statement_text, period_hint):
            if ev.get('type') == 'progress':
                yield ev
            elif ev.get('type') == 'rows':
                rows = ev.get('rows') or []
                chunk_count = int(ev.get('chunk_count') or 1)
        yield {
            'type': 'result',
            'rows': rows or [],
            'meta': {
                'mode': 'llm_full',
                'reason': 'layout_low_yield',
                'hint_row_count': len(hint_rows),
                'eligible_hint_count': eligible,
                'layout_attempt_count': layout_attempt_count,
                'llm_chunk_count': chunk_count,
            },
        }
        return

    yield {
        'type': 'result',
        'rows': rows,
        'meta': {
            'mode': 'layout_hints',
            'hint_row_count': len(hint_rows),
            'eligible_hint_count': eligible,
            'raw_row_count': len(rows),
            'skip_pattern_count': len(skip_patterns),
            'retried_defaults_only': layout_retry_defaults,
        },
    }


def _normalize_spending_transactions(raw_list: list) -> list:
    out = []
    for row in raw_list:
        if not isinstance(row, dict):
            continue
        d, started, completed = _resolve_spending_transaction_dates(row)
        if d is None:
            continue
        desc = str(row.get('description') or '').strip()[:500]
        if not desc:
            desc = 'Bank transaction'
        col_dir, col_amt = _direction_and_amount_from_money_columns(row)
        if col_dir is not None and col_amt is not None:
            amount = round(col_amt, 2)
            direction = col_dir
        else:
            try:
                amount = abs(float(row.get('amount')))
            except (TypeError, ValueError):
                continue
            if amount <= 0:
                continue
            direction = _normalize_spending_direction(row.get('direction'), desc, amount)
        item = {
            'id': str(uuid.uuid4()),
            'date': d.strftime('%Y-%m-%d'),
            'month': d.strftime('%Y-%m'),
            'description': desc,
            'amount': round(amount, 2),
            'direction': direction,
            'category': None,
            'confidence': None,
            'rationale': '',
            'date_boundary': False,
        }
        if started is not None:
            item['started_date'] = started.strftime('%Y-%m-%d')
        if completed is not None:
            item['completed_date'] = completed.strftime('%Y-%m-%d')
        out.append(item)
    out.sort(key=lambda x: (x['date'], x['description']))
    return out


def _classify_outgoing_descriptions_llm(descriptions: list[str]) -> dict:
    if not descriptions:
        return {}
    client = _get_openai_client()
    if not client:
        return {}
    model = os.getenv('OPENAI_MODEL', 'gpt-4o-mini').strip() or 'gpt-4o-mini'
    system = (
        'Classify outgoing personal spending descriptions into one category from this set only: '
        + ', '.join(SPENDING_ALLOWED_CATEGORIES)
        + '. Use travel for hotels, accommodation, holiday rentals, and tourism (tours, attractions, museums, activities while travelling). '
        'Return only valid JSON as {"items":[{"description":"string","category":"string","confidence":number,"reason":"string"}]}. '
        'confidence must be between 0 and 1.'
    )
    user_msg = 'Descriptions:\n' + '\n'.join(f'- {d}' for d in descriptions[:150])
    try:
        completion = client.chat.completions.create(
            model=model,
            messages=[
                {'role': 'system', 'content': system},
                {'role': 'user', 'content': user_msg},
            ],
            response_format={'type': 'json_object'},
            temperature=0.1,
        )
    except Exception as e:
        # Timeout / connection / API error: don't crash the import. Callers
        # fall back to the 'other' category for anything left unclassified.
        logger.warning('outgoing classification LLM call failed: %s', e)
        return {}
    raw = completion.choices[0].message.content or '{}'
    parsed, jerr = _parse_llm_json_object(raw, context='spending_classify')
    if jerr:
        return {}
    items = parsed.get('items')
    if not isinstance(items, list):
        return {}
    out = {}
    for item in items:
        if not isinstance(item, dict):
            continue
        desc = str(item.get('description') or '').strip()
        if not desc:
            continue
        cat = str(item.get('category') or 'other').strip().lower()
        if cat not in SPENDING_CATEGORY_SET:
            cat = 'other'
        try:
            conf = float(item.get('confidence', 0))
        except (TypeError, ValueError):
            conf = 0.0
        conf = max(0.0, min(1.0, conf))
        out[_normalize_label(desc)] = {
            'category': cat,
            'confidence': round(conf, 3),
            'rationale': str(item.get('reason') or '').strip()[:240],
        }
    return out


def _apply_outgoing_classification(rows: list, spending: dict) -> bool:
    changed = False
    overrides = spending.get('classification_overrides') or {}
    cache = spending.get('classification_cache') or {}
    unknown_labels = []
    for row in rows:
        if row.get('direction') != 'outgoing':
            row['category'] = None
            continue
        label = _normalize_label(row.get('description', ''))
        if not label:
            row['category'] = 'other'
            row['confidence'] = 0.0
            row['rationale'] = 'fallback'
            continue
        if label in overrides:
            row['category'] = str(overrides[label]).strip().lower() or 'other'
            if row['category'] not in SPENDING_CATEGORY_SET:
                row['category'] = 'other'
            row['confidence'] = 1.0
            row['rationale'] = 'manual override'
            continue
        cached = cache.get(label)
        if isinstance(cached, dict) and cached.get('category') in SPENDING_CATEGORY_SET:
            row['category'] = cached['category']
            row['confidence'] = cached.get('confidence')
            row['rationale'] = cached.get('rationale') or ''
            continue
        unknown_labels.append(label)

    if unknown_labels:
        desc_map = {}
        for row in rows:
            if row.get('direction') != 'outgoing':
                continue
            label = _normalize_label(row.get('description', ''))
            if label and label not in desc_map:
                desc_map[label] = row.get('description', '')
        llm_result = _classify_outgoing_descriptions_llm([desc_map[k] for k in sorted(set(unknown_labels))])
        for label in set(unknown_labels):
            classified = llm_result.get(label)
            if not classified:
                classified = {'category': 'other', 'confidence': 0.0, 'rationale': 'fallback'}
            cache[label] = classified
            changed = True
        spending['classification_cache'] = cache

    for row in rows:
        if row.get('direction') != 'outgoing':
            continue
        label = _normalize_label(row.get('description', ''))
        if label in overrides:
            continue
        c = cache.get(label) if label else None
        if not isinstance(c, dict):
            c = {'category': 'other', 'confidence': 0.0, 'rationale': 'fallback'}
        cat = str(c.get('category') or 'other').strip().lower()
        if cat not in SPENDING_CATEGORY_SET:
            cat = 'other'
        row['category'] = cat
        row['confidence'] = c.get('confidence')
        row['rationale'] = c.get('rationale') or ''
    return changed


def _invalidate_automatic_classification_for_labels(spending: dict, labels: set[str]) -> bool:
    """Drop cached LLM categories for these labels so the next apply will re-query. Skips manual overrides."""
    if not labels:
        return False
    overrides = spending.get('classification_overrides') or {}
    cache = spending.get('classification_cache') or {}
    changed = False
    for label in labels:
        if not label or label in overrides:
            continue
        if label in cache:
            del cache[label]
            changed = True
    if changed:
        spending['classification_cache'] = cache
    return changed


def _sync_outgoing_transactions_from_cache_for_labels(spending: dict, labels: set[str]) -> None:
    """Apply classification_cache to all outgoing transactions whose description label is in ``labels`` (excluding overrides)."""
    if not labels:
        return
    overrides = spending.get('classification_overrides') or {}
    cache = spending.get('classification_cache') or {}
    for t in spending.get('transactions') or []:
        if t.get('direction') != 'outgoing':
            continue
        label = _normalize_label(t.get('description', ''))
        if not label or label not in labels or label in overrides:
            continue
        c = cache.get(label)
        if not isinstance(c, dict):
            c = {'category': 'other', 'confidence': 0.0, 'rationale': 'fallback'}
        cat = str(c.get('category') or 'other').strip().lower()
        if cat not in SPENDING_CATEGORY_SET:
            cat = 'other'
        t['category'] = cat
        t['confidence'] = c.get('confidence')
        t['rationale'] = c.get('rationale') or ''


def _rerun_spending_categorization_for_month(spending: dict, report_month: str) -> dict:
    """
    Re-run LLM categorisation for outgoing lines in ``report_month``.
    Preserves classification_overrides. Updates cache and all transactions sharing refreshed labels.
    """
    rm = str(report_month or '').strip()[:7]
    if len(rm) != 7 or rm[4] != '-':
        return {'ok': False, 'error': 'report_month (YYYY-MM) is required'}
    txs = spending.get('transactions') or []
    month_rows = [t for t in txs if _report_month_for_spending_tx(t) == rm]
    if not month_rows:
        return {'ok': False, 'error': 'No transactions for that month'}
    overrides = spending.get('classification_overrides') or {}
    labels: set[str] = set()
    for t in month_rows:
        if t.get('direction') != 'outgoing':
            continue
        label = _normalize_label(t.get('description', ''))
        if label:
            labels.add(label)
    _invalidate_automatic_classification_for_labels(spending, labels)
    _apply_outgoing_classification(month_rows, spending)
    refreshed = {l for l in labels if l not in overrides}
    _sync_outgoing_transactions_from_cache_for_labels(spending, refreshed)
    _recompute_monthly_insights(spending, None)
    return {'ok': True, 'labels_refreshed': len(refreshed)}


def _spending_is_paired_leg(t: dict) -> bool:
    return bool(t.get('internal_transfer') and t.get('transfer_pair_id'))


def _spending_excluded_from_insight_metrics(t: dict) -> bool:
    """Paired internal-transfer legs and user-flagged rows are omitted from income/outgoing KPIs."""
    if _spending_is_paired_leg(t):
        return True
    return bool(t.get('insights_excluded'))


def _spending_tx_parsed_date(tx: dict) -> date | None:
    return _parse_iso_date(str(tx.get('date') or ''))


def _spending_pair_amounts_match(a: float, b: float) -> bool:
    x = abs(float(a))
    y = abs(float(b))
    diff = abs(x - y)
    hi = max(x, y)
    if hi < 1e-12:
        return diff <= SPENDING_TRANSFER_MIN_AMOUNT_DIFF + 1e-9
    base = max(SPENDING_TRANSFER_MIN_AMOUNT_DIFF, SPENDING_TRANSFER_LARGE_LEG_PCT * hi)
    if hi < SPENDING_TRANSFER_SMALL_LEG_MAX:
        tol = min(SPENDING_TRANSFER_SMALL_LEG_ABS_CAP, base)
    else:
        tol = min(SPENDING_TRANSFER_LARGE_LEG_ABS_CAP, base)
    return diff <= tol + 1e-9


def _spending_pair_dates_within_window(d0: date | None, d1: date | None) -> bool:
    if d0 is None or d1 is None:
        return False
    return abs((d0 - d1).days) <= SPENDING_TRANSFER_MAX_DAY_GAP


def _clear_auto_transfer_pairs_in_month(txs: list, month_key: str) -> int:
    """
    Remove auto-applied internal transfer links for a reporting month. Manual pairs are kept.
    Returns the number of pair bonds cleared (one per pair, not per leg).
    """
    mk = (month_key or '').strip()[:7]
    pair_ids: set[str] = set()
    for t in txs:
        if str(_report_month_for_spending_tx(t)) != mk:
            continue
        if t.get('pairing_source') == 'auto' and t.get('transfer_pair_id'):
            pair_ids.add(str(t['transfer_pair_id']))
    if not pair_ids:
        return 0
    broken = 0
    for t in txs:
        pid = str(t.get('transfer_pair_id') or '')
        if pid in pair_ids:
            t.pop('transfer_pair_id', None)
            t['internal_transfer'] = False
            t.pop('pairing_source', None)
            broken += 1
    return broken // 2


def _suggest_spending_transfer_pairs(outs: list, ins: list) -> list[tuple[dict, dict]]:
    """
    1:1 matching of outgoing vs incoming: amounts within tolerance, dates within
    SPENDING_TRANSFER_MAX_DAY_GAP, each row at most once. Among valid edges, prefer
    **smaller amount gap** (tighter leg match) then **fewer days apart**, then stable ids
    so closer matches are chosen instead of the first by arbitrary sort order.
    """
    if not outs or not ins:
        return []
    candidates: list[tuple[dict, dict, float, int, str, str]] = []
    for o in outs:
        oid = str(o.get('id') or '')
        if not oid:
            continue
        od = _spending_tx_parsed_date(o)
        o_amt = float(o.get('amount', 0))
        for i in ins:
            iid = str(i.get('id') or '')
            if not iid:
                continue
            i_amt = float(i.get('amount', 0))
            if not _spending_pair_amounts_match(o_amt, i_amt):
                continue
            idt = _spending_tx_parsed_date(i)
            if not _spending_pair_dates_within_window(od, idt):
                continue
            amount_diff = abs(abs(o_amt) - abs(i_amt))
            day_gap = 0
            if od and idt:
                day_gap = abs((od - idt).days)
            # Lower keys are better: tighter amount, same-day / closer dates, deterministic tiebreak
            candidates.append((o, i, amount_diff, day_gap, oid, iid))
    candidates.sort(
        key=lambda r: (r[2], r[3], r[4], r[5]),
    )
    used_o: set[str] = set()
    used_in: set[str] = set()
    pairs: list[tuple[dict, dict]] = []
    for o, i, _ad, _dg, _oid, _iid in candidates:
        oid = str(o.get('id') or '')
        iid = str(i.get('id') or '')
        if not oid or not iid or oid in used_o or iid in used_in:
            continue
        used_o.add(oid)
        used_in.add(iid)
        pairs.append((o, i))
    return pairs


def _apply_spending_pair_to_rows(a: dict, b: dict, source: str) -> None:
    pid = str(uuid.uuid4())
    a['transfer_pair_id'] = pid
    b['transfer_pair_id'] = pid
    a['internal_transfer'] = True
    b['internal_transfer'] = True
    a['pairing_source'] = source
    b['pairing_source'] = source


def _unlink_spending_pair(txs: list, tx: dict) -> bool:
    """Clear pairing on tx and its peer. Returns True if a pair was removed."""
    pid = str(tx.get('transfer_pair_id') or '')
    if not pid:
        return False
    found = 0
    for t in txs:
        if str(t.get('transfer_pair_id') or '') == pid:
            t.pop('transfer_pair_id', None)
            t['internal_transfer'] = False
            t.pop('pairing_source', None)
            found += 1
    return found > 0


def _spending_manual_pair_error(a: dict | None, b: dict | None) -> str | None:
    """Return an error string if a and b cannot be manually linked; else None."""
    if not a or not b:
        return 'Transaction not found'
    if str(a.get('id')) == str(b.get('id')):
        return 'Cannot pair a transaction with itself'
    if str(_report_month_for_spending_tx(a)) != str(_report_month_for_spending_tx(b)):
        return 'Transactions must share the same reporting month'
    dirs = {a.get('direction'), b.get('direction')}
    if dirs != {'incoming', 'outgoing'}:
        return 'One transaction must be outgoing and one incoming'
    if a.get('transfer_pair_id') or b.get('transfer_pair_id'):
        return 'Unlink existing pair(s) before creating a new one'
    am1, am2 = float(a.get('amount', 0)), float(b.get('amount', 0))
    if not _spending_pair_amounts_match(am1, am2):
        return 'Amounts must match within configured tolerance'
    d0 = _spending_tx_parsed_date(a)
    d1 = _spending_tx_parsed_date(b)
    if not _spending_pair_dates_within_window(d0, d1):
        return f'Dates must be within {SPENDING_TRANSFER_MAX_DAY_GAP} day(s)'
    return None


def _apply_auto_pairing_to_tx_pool(txs: list, month_key: str) -> dict:
    """
    Re-run auto internal-transfer pairing for all rows in ``txs`` for ``month_key``.
    ``txs`` is typically the full ``spending['transactions']`` list; may also be
    a throwaway list containing only a month's rows for preview simulation.
    """
    _clear_auto_transfer_pairs_in_month(txs, month_key)
    mk = (month_key or '').strip()[:7]
    month_rows = [t for t in txs if str(_report_month_for_spending_tx(t)) == mk]
    unpaired = [t for t in month_rows if not t.get('transfer_pair_id')]
    outs = [t for t in unpaired if t.get('direction') == 'outgoing']
    ins = [t for t in unpaired if t.get('direction') == 'incoming']
    pair_tuples = _suggest_spending_transfer_pairs(outs, ins)
    for o, i in pair_tuples:
        _apply_spending_pair_to_rows(o, i, 'auto')
    unpaired_out = [t for t in month_rows if t.get('direction') == 'outgoing' and not t.get('transfer_pair_id')]
    unpaired_in = [t for t in month_rows if t.get('direction') == 'incoming' and not t.get('transfer_pair_id')]
    return {
        'applied_pairs': len(pair_tuples),
        'unmatched_outgoing_count': len(unpaired_out),
        'unmatched_incoming_count': len(unpaired_in),
        'unmatched_outgoing_total': round(sum(float(t.get('amount', 0)) for t in unpaired_out), 2),
        'unmatched_incoming_total': round(sum(float(t.get('amount', 0)) for t in unpaired_in), 2),
    }


def apply_auto_transfer_pairing_for_month(spending: dict, month_key: str) -> dict:
    """
    Re-run auto internal-transfer pairing for a reporting month. Clears prior auto pairs only,
    then matches unpaired incoming/outgoing. Manual pairs are preserved.
    """
    txs = spending.get('transactions') or []
    return _apply_auto_pairing_to_tx_pool(txs, month_key)


def simulate_spending_transfer_reconciliation_preview(
    spending: dict,
    report_month: str,
    candidate_rows: list,
) -> dict:
    """
    What-if pairing if ``candidate_rows`` are imported, merged with any existing
    stored transactions for the same reporting month. Does not modify stored data.
    """
    mk = (report_month or '').strip()[:7]
    store = spending.get('transactions') or []
    existing = [copy.deepcopy(t) for t in store if str(_report_month_for_spending_tx(t)) == mk]
    cands = [copy.deepcopy(t) for t in candidate_rows]
    for t in cands:
        t['report_month'] = mk
    pool = existing + cands
    stats = _apply_auto_pairing_to_tx_pool(pool, mk)
    month_rows = [t for t in pool if str(_report_month_for_spending_tx(t)) == mk]
    rec = _spending_transfer_reconciliation_for_month(month_rows)
    cand_ids = {str(t.get('id') or '') for t in cands if t.get('id')}
    pairing: dict = {}
    for t in month_rows:
        tid = str(t.get('id') or '')
        if not tid or tid not in cand_ids:
            continue
        pid = str(t.get('transfer_pair_id') or '')
        if not pid:
            pairing[tid] = {
                'paired': False,
                'peer_id': None,
                'peer_is_from_ledger': False,
                'peer_description': None,
            }
            continue
        other = next(
            (
                x
                for x in month_rows
                if str(x.get('id')) != tid
                and str(x.get('transfer_pair_id') or '') == pid
            ),
            None,
        )
        if not other:
            pairing[tid] = {
                'paired': False,
                'peer_id': None,
                'peer_is_from_ledger': False,
                'peer_description': None,
            }
            continue
        oid = str(other.get('id') or '')
        pairing[tid] = {
            'paired': True,
            'peer_id': oid,
            'peer_is_from_ledger': oid not in cand_ids,
            'peer_description': (str(other.get('description') or ''))[:120] or None,
        }
    return {
        'reconciliation': rec,
        'pairing': pairing,
        'ledger_row_count_in_month': len(existing),
        'auto_applied_pairs_in_simulation': stats['applied_pairs'],
    }


def _month_prev(month_key: str, delta: int = 1) -> str | None:
    try:
        d = datetime.strptime(month_key + '-01', '%Y-%m-%d').date()
    except ValueError:
        return None
    prev = d - relativedelta(months=delta)
    return prev.strftime('%Y-%m')


def _month_next(month_key: str, delta: int = 1) -> str | None:
    try:
        d = datetime.strptime(month_key + '-01', '%Y-%m-%d').date()
    except ValueError:
        return None
    nxt = d + relativedelta(months=delta)
    return nxt.strftime('%Y-%m')


SUBSCRIPTION_SIGNAL_WINDOW_MONTHS = 6
# Monthly totals / selected-occurrence spread: utilities and insurance often move more than 25%.
SUBSCRIPTION_MAX_AMOUNT_SPREAD_RATIO = 0.40  # (max-min)/mean across selected months
# Pairwise charge comparison inside a monthly chain (relative vs the larger amount).
SUBSCRIPTION_PAIR_AMOUNT_TOL = 0.35
# Day-of-month may drift a few days; also used for month-end wrap (30 vs 2).
SUBSCRIPTION_DAY_TOLERANCE = 5
# Near-monthly spacing between successive selected charges (handles posting-date drift).
SUBSCRIPTION_INTERVAL_MIN_DAYS = 24
SUBSCRIPTION_INTERVAL_MAX_DAYS = 40
# Allow one skipped month when day-of-month still lines up (e.g. late post after a gap).
SUBSCRIPTION_SKIP_INTERVAL_MAX_DAYS = 70
# Merging lookalike labels (e.g. "HALIFAX" / "HALIFAX DD", "SPOTIFY" / "SPOTIFY LDN")
SUBSCRIPTION_FUZZY_MIN_PREFIX_LEN = 4
SUBSCRIPTION_LABEL_MERGE_AMOUNT_TOL = 0.25  # (max-min)/max when two labels both have spend the same month

# Strip from normalized labels so "VISA SPOTIFY LONDON" and "SPOTIFY" can align (prefix is often not the merchant).
_SUBSCRIPTION_MERGE_STRIP_LEADING = frozenset({
    'atm', 'bac', 'bacs', 'card', 'cash', 'cdr', 'chg', 'cheque', 'chq', 'cont',
    'contactless', 'credit', 'crdt', 'dd', 'debit', 'dep', 'deposit', 'direct', 'ecom',
    'electron', 'faster', 'fee', 'fps', 'from', 'fpi', 'fpo', 'intnl', 'intl', 'mc', 'maestro',
    'mobile', 'online', 'pay', 'payment', 'pmnt', 'pos', 'pmt', 'purch', 'purchase', 'sepa',
    'standing', 'tfr', 'to', 'trf', 'transfer', 'vdp', 'visa', 'withdrawal',
})
_SUBSCRIPTION_MERGE_STRIP_TRAILING = frozenset({
    'crdt', 'credit', 'debit', 'electron', 'eu', 'eur', 'gbr', 'gb', 'intl', 'intnl', 'ldn',
    'lond', 'london', 'maestro', 'mc', 'us', 'usd', 'uk', 'vis', 'visa',
})


def _subscription_labels_trim_for_merge(s: str) -> str:
    """Drop common bank / location noise from the start and end of a normalized label."""
    toks = (s or '').strip().split()
    if not toks:
        return ''
    while toks and toks[0] in _SUBSCRIPTION_MERGE_STRIP_LEADING:
        toks = toks[1:]
    while toks and toks[-1] in _SUBSCRIPTION_MERGE_STRIP_TRAILING:
        toks = toks[:-1]
    return ' '.join(toks)


def _subscription_labels_fuzzy_match(a: str, b: str) -> bool:
    """
    True if ``a`` and ``b`` are the same "merchant" for subscription grouping:
    - exact match, or
    - longer is ``shorter + " …"`` (word-boundary prefix) when the shorter is long enough, or
    - same first word, length ≥ 4 (covers NETFLIX.COM vs NETFLIX LDN, etc.).

    Common bank prefixes (e.g. ``visa``, ``pos``, ``direct debit``) and trailing place/payment
    tokens (``london``, ``vis``) are stripped first so the merchant name lines up.

    Very short strings only match if equal, to avoid spurious links.
    """
    a = _subscription_labels_trim_for_merge((a or '').strip())
    b = _subscription_labels_trim_for_merge((b or '').strip())
    if not a or not b:
        return False
    if a == b:
        return True
    s, t = (a, b) if len(a) <= len(b) else (b, a)
    if len(s) < SUBSCRIPTION_FUZZY_MIN_PREFIX_LEN:
        return False
    if t == s or t.startswith(s + ' '):
        return True
    wa = a.split()
    wb = b.split()
    if wa and wb and len(wa[0]) >= SUBSCRIPTION_FUZZY_MIN_PREFIX_LEN and wa[0] == wb[0]:
        return True
    return False


def _subscription_pair_merge_amounts_ok(
    by_a: dict,
    by_b: dict,
    months_seq: list[str],
    tol: float = SUBSCRIPTION_LABEL_MERGE_AMOUNT_TOL,
) -> bool:
    """
    If two labels both have non-trivial spend in the same month, the two amounts
    must be within ``tol`` of each other (by relative error vs the larger) so we
    do not join unrelated same-prefix charges.
    """
    for m in months_seq:
        try:
            aa = float(by_a.get(m) or 0.0)
            bb = float(by_b.get(m) or 0.0)
        except (TypeError, ValueError):
            continue
        if aa <= 0.005 or bb <= 0.005:
            continue
        lo, hi = (aa, bb) if aa <= bb else (bb, aa)
        if hi <= 0:
            return False
        if (hi - lo) / hi > tol + 1e-9:
            return False
    return True


def _uf_find(parent: list, i: int) -> int:
    while parent[i] != i:
        parent[i] = parent[parent[i]]
        i = parent[i]
    return i


def _uf_union(parent: list, rank: list, i: int, j: int) -> None:
    ri, rj = _uf_find(parent, i), _uf_find(parent, j)
    if ri == rj:
        return
    if rank[ri] < rank[rj]:
        parent[ri] = rj
    else:
        parent[rj] = ri
        if rank[ri] == rank[rj]:
            rank[ri] += 1


def _merge_subscription_signal_label_groups(
    label_month_totals: dict,
    label_sample_desc: dict,
    months_seq: list[str],
    label_charges: dict | None = None,
) -> tuple[dict, dict, dict]:
    """
    Union labels that are fuzzy name matches and, when they overlap in a month,
    have similar total amounts. Rebuilds total-by-month, sample descriptions
    (longest original description is kept for display), and optional charge lists.
    """
    empty_charges: dict = {}
    labels = [k for k in label_month_totals.keys() if k]
    n = len(labels)
    if n <= 1:
        if n == 1 and label_charges is not None:
            only = labels[0]
            empty_charges[only] = list(label_charges.get(only) or [])
        elif label_charges:
            for lab, rows in label_charges.items():
                empty_charges[lab] = list(rows or [])
        return label_month_totals, label_sample_desc, empty_charges if label_charges is not None else {}
    parent = list(range(n))
    rank = [0] * n
    for i in range(n):
        for j in range(i + 1, n):
            li, lj = labels[i], labels[j]
            if not _subscription_labels_fuzzy_match(li, lj):
                continue
            if not _subscription_pair_merge_amounts_ok(
                label_month_totals[li],
                label_month_totals[lj],
                months_seq,
            ):
                continue
            _uf_union(parent, rank, i, j)

    groups: dict[int, set[str]] = {}
    for i, lab in enumerate(labels):
        r = _uf_find(parent, i)
        groups.setdefault(r, set()).add(lab)

    merged_totals: dict = {}
    merged_desc: dict = {}
    merged_charges: dict = {}
    for g in groups.values():
        if len(g) == 1:
            only = next(iter(g))
            merged_totals[only] = dict(label_month_totals[only])
            merged_desc[only] = label_sample_desc.get(only, only)
            if label_charges is not None:
                merged_charges[only] = list(label_charges.get(only) or [])
            continue
        by_month: dict[str, float] = defaultdict(float)
        for lab in g:
            for mk, v in (label_month_totals.get(lab) or {}).items():
                try:
                    by_month[mk] += float(v)
                except (TypeError, ValueError):
                    continue
        # Canonical key: longer normalized name first (most specific for sorting / tests)
        canonical = max(g, key=lambda x: (len(x), x))
        merged_totals[canonical] = dict(by_month)
        best = ''
        for lab in g:
            s = label_sample_desc.get(lab) or lab
            if len(s) > len(best):
                best = s
        merged_desc[canonical] = best
        if label_charges is not None:
            rows: list = []
            for lab in g:
                rows.extend(label_charges.get(lab) or [])
            merged_charges[canonical] = rows
    return merged_totals, merged_desc, merged_charges


def _months_window_ending(month_key: str, n: int) -> list[str]:
    """Chronological list of up to ``n`` calendar months ending at ``month_key`` (inclusive)."""
    out: list[str] = []
    cur = str(month_key or '').strip()[:7]
    if len(cur) != 7 or cur[4] != '-':
        return out
    for _ in range(max(1, n)):
        out.append(cur)
        p = _month_prev(cur, 1)
        if not p:
            break
        cur = p
    out.reverse()
    return out


def _longest_consecutive_month_streak(sorted_months: list[str]) -> int:
    if not sorted_months:
        return 0
    best = 1
    cur_run = 1
    for i in range(1, len(sorted_months)):
        if _month_next(sorted_months[i - 1], 1) == sorted_months[i]:
            cur_run += 1
            best = max(best, cur_run)
        else:
            cur_run = 1
    return best


def _subscription_amounts_compatible(
    a: float,
    b: float,
    tol: float = SUBSCRIPTION_PAIR_AMOUNT_TOL,
) -> bool:
    """True when two charge amounts are within ``tol`` relative to the larger."""
    try:
        aa = float(a)
        bb = float(b)
    except (TypeError, ValueError):
        return False
    if aa <= 0.005 or bb <= 0.005:
        return False
    lo, hi = (aa, bb) if aa <= bb else (bb, aa)
    return (hi - lo) / hi <= tol + 1e-9


def _subscription_dom_distance(day_a: int, day_b: int) -> int:
    """Min distance between day-of-month values on a 31-day circle (month-end wrap)."""
    try:
        a = int(day_a)
        b = int(day_b)
    except (TypeError, ValueError):
        return 99
    diff = abs(a - b)
    return min(diff, 31 - diff)


def _subscription_normalize_charges(raw_charges: list) -> list[dict]:
    """Parse charge dicts to ``{date, amount, report_month}`` sorted by date."""
    items: list[dict] = []
    for c in raw_charges or []:
        if not isinstance(c, dict):
            continue
        d = c.get('date')
        if not isinstance(d, date):
            d = _parse_iso_date(str(d or ''))
        if d is None:
            continue
        try:
            amt = float(c.get('amount') or 0)
        except (TypeError, ValueError):
            continue
        if amt <= 0.005:
            continue
        rm = str(c.get('report_month') or '').strip()[:7]
        if len(rm) != 7 or rm[4] != '-':
            rm = d.strftime('%Y-%m')
        items.append({'date': d, 'amount': amt, 'report_month': rm})
    items.sort(key=lambda x: (x['date'], x['amount']))
    return items


def _representative_monthly_bill_amount(raw_charges: list) -> float:
    """
    Amount to reserve for a monthly bill from same-merchant charges in one month.

    Normally sums (multiple distinct spends). When two+ similar unit charges are
    spaced like separate monthly cycles (e.g. delayed prior bill + current, both
    posting in the same calendar month), return one representative unit amount
    instead of combining them.
    """
    items = _subscription_normalize_charges(raw_charges)
    if not items:
        # Allow amount-only rows (no parseable date) — sum what we can.
        total = 0.0
        for c in raw_charges or []:
            if not isinstance(c, dict):
                continue
            try:
                amt = float(c.get('amount') or 0)
            except (TypeError, ValueError):
                continue
            if amt > 0.005:
                total += amt
        return round(total, 2)
    if len(items) == 1:
        return round(float(items[0]['amount']), 2)

    amounts = [float(c['amount']) for c in items]
    similar = all(
        _subscription_amounts_compatible(amounts[0], a) for a in amounts[1:]
    )
    span_days = (items[-1]['date'] - items[0]['date']).days
    if similar and span_days >= SUBSCRIPTION_INTERVAL_MIN_DAYS:
        # Delayed double-post in one calendar month — keep a single cycle amount.
        # Prefer the latest charge (usually the "on time" one for this month).
        return round(float(items[-1]['amount']), 2)
    return round(sum(amounts), 2)

def _subscription_longest_monthly_chain(charges: list[dict]) -> list[dict]:
    """
    Longest chain of similar-amount charges spaced like a monthly bill.

    Successive links are preferably 24–40 days apart. One skipped month is allowed
    (up to ~70 days) when the day-of-month still matches within
    ``SUBSCRIPTION_DAY_TOLERANCE``. Amounts may vary within
    ``SUBSCRIPTION_PAIR_AMOUNT_TOL``.
    """
    items = _subscription_normalize_charges(charges)
    n = len(items)
    if n < 2:
        return list(items)

    best: list[dict] = []

    def _chain_dom_score(chain: list[dict]) -> float:
        if len(chain) < 2:
            return 0.0
        days = [c['date'].day for c in chain]
        med = sorted(days)[len(days) // 2]
        return -sum(_subscription_dom_distance(d, med) for d in days) / len(days)

    for i in range(n):
        chain = [items[i]]
        for j in range(i + 1, n):
            tip = chain[-1]
            gap = (items[j]['date'] - tip['date']).days
            if gap < SUBSCRIPTION_INTERVAL_MIN_DAYS:
                continue
            if not _subscription_amounts_compatible(tip['amount'], items[j]['amount']):
                continue
            if gap <= SUBSCRIPTION_INTERVAL_MAX_DAYS:
                chain.append(items[j])
                continue
            # One skipped calendar month: still OK if day-of-month lines up.
            if (
                gap <= SUBSCRIPTION_SKIP_INTERVAL_MAX_DAYS
                and _subscription_dom_distance(tip['date'].day, items[j]['date'].day)
                <= SUBSCRIPTION_DAY_TOLERANCE
            ):
                chain.append(items[j])
                continue
            if gap > SUBSCRIPTION_SKIP_INTERVAL_MAX_DAYS:
                break
        if len(chain) > len(best) or (
            len(chain) == len(best) and _chain_dom_score(chain) > _chain_dom_score(best)
        ):
            best = chain
    return best


def _subscription_calendar_month_representatives(charges: list[dict]) -> list[dict]:
    """
    Fallback: one representative charge per calendar report month, preferring
    day-of-month clustering around the median day among amount-stable picks.
    """
    items = _subscription_normalize_charges(charges)
    if len(items) < 2:
        return list(items)

    by_month: dict[str, list[dict]] = defaultdict(list)
    for c in items:
        by_month[c['report_month']].append(c)

    # Seed median day from months that look singly-charged.
    seed_days: list[int] = []
    for mk, rows in by_month.items():
        if len(rows) == 1:
            seed_days.append(rows[0]['date'].day)
        else:
            # Prefer the charge closest to mid-month cluster later; seed with all days.
            seed_days.extend(r['date'].day for r in rows)
    if not seed_days:
        return []
    median_day = sorted(seed_days)[len(seed_days) // 2]

    selected: list[dict] = []
    for mk in sorted(by_month.keys()):
        rows = by_month[mk]
        # Prefer rows near median day; break ties by closeness to median amount of prior picks.
        def _score(r: dict) -> tuple:
            dom = _subscription_dom_distance(r['date'].day, median_day)
            return (dom, r['amount'])

        rows_sorted = sorted(rows, key=_score)
        pick = rows_sorted[0]
        if selected and not _subscription_amounts_compatible(selected[-1]['amount'], pick['amount']):
            # Try another row in the month with compatible amount
            alt = next(
                (
                    r for r in rows_sorted
                    if _subscription_amounts_compatible(selected[-1]['amount'], r['amount'])
                ),
                None,
            )
            if alt is None:
                continue
            pick = alt
        # Soft day filter once we have a chain tip
        if selected:
            if _subscription_dom_distance(selected[-1]['date'].day, pick['date'].day) > SUBSCRIPTION_DAY_TOLERANCE:
                # Still allow if interval is near-monthly
                gap = (pick['date'] - selected[-1]['date']).days
                if gap < SUBSCRIPTION_INTERVAL_MIN_DAYS or gap > SUBSCRIPTION_SKIP_INTERVAL_MAX_DAYS:
                    continue
        selected.append(pick)

    # Require overall day clustering for weak lists
    if len(selected) >= 2:
        days = [c['date'].day for c in selected]
        med = sorted(days)[len(days) // 2]
        consistent = [
            c for c in selected
            if _subscription_dom_distance(c['date'].day, med) <= SUBSCRIPTION_DAY_TOLERANCE
        ]
        if len(consistent) >= 2:
            return consistent
    return selected if len(selected) >= 2 else []


def _subscription_select_monthly_occurrences(charges: list[dict]) -> list[dict]:
    """
    Choose the best monthly-recurring subset for a merchant label.

    Prefers interval-based chains (flexible posting dates); falls back to
    calendar-month representatives with day-of-month clustering.
    """
    chain = _subscription_longest_monthly_chain(charges)
    calendar = _subscription_calendar_month_representatives(charges)
    if len(chain) >= len(calendar) and len(chain) >= 2:
        return chain
    if len(calendar) >= 2:
        return calendar
    return chain if len(chain) >= 2 else []


def _subscription_pattern_too_noisy(all_charges: list[dict], selected: list[dict]) -> bool:
    """
    True when the merchant has lots of non-selected spend (typical shopping),
    so a short similar-amount coincidence is unlikely to be a bill.
    """
    all_n = len(_subscription_normalize_charges(all_charges))
    sel_n = len(selected)
    if sel_n <= 0:
        return True
    # Many extra charges beyond the monthly picks → noisy merchant
    return all_n >= sel_n * 2 + 2


def _subscription_signals_for_month(spending: dict, month_key: str) -> list:
    """
    Cross-month recurring / subscription-style spend: similar normalized merchant labels
    (prefix / same first word) are merged when monthly amounts are compatible; then
    near-monthly charge chains (flexible day-of-month + amount drift) qualify.
    Excludes internal-transfer legs.
    """
    months_seq = _months_window_ending(month_key, SUBSCRIPTION_SIGNAL_WINDOW_MONTHS)
    if not months_seq:
        return []
    month_set = set(months_seq)
    label_month_totals: dict[str, dict[str, float]] = defaultdict(lambda: defaultdict(float))
    label_sample_desc: dict[str, str] = {}
    label_charges: dict[str, list] = defaultdict(list)

    for t in spending.get('transactions') or []:
        if t.get('direction') != 'outgoing':
            continue
        if _spending_excluded_from_insight_metrics(t):
            continue
        rm = _report_month_for_spending_tx(t)
        if rm not in month_set:
            continue
        label = _normalize_label(t.get('description', ''))
        if not label:
            continue
        try:
            amt = float(t.get('amount', 0))
        except (TypeError, ValueError):
            continue
        if amt <= 0.005:
            continue
        label_month_totals[label][rm] += amt
        if label not in label_sample_desc:
            label_sample_desc[label] = str(t.get('description', ''))[:120]
        parsed = _spending_tx_parsed_date(t)
        if parsed is not None:
            label_charges[label].append({
                'date': parsed,
                'amount': amt,
                'report_month': rm,
            })

    label_month_totals, label_sample_desc, label_charges = _merge_subscription_signal_label_groups(
        label_month_totals, label_sample_desc, months_seq, label_charges
    )

    out: list = []
    for label, by_month in label_month_totals.items():
        charges = label_charges.get(label) or []
        selected = _subscription_select_monthly_occurrences(charges)

        # Fall back to month-total presence when dates were missing entirely.
        if len(selected) < 2:
            present_months = sorted(m for m, v in by_month.items() if v > 0.005)
            if month_key not in present_months or len(present_months) < 2:
                continue
            totals = [by_month[m] for m in present_months]
            mean_t = sum(totals) / len(totals) if totals else 0.0
            if mean_t <= 0:
                continue
            spread = (max(totals) - min(totals)) / mean_t if len(totals) > 1 else 0.0
            streak = _longest_consecutive_month_streak(present_months)
            qualifies = (
                (streak >= 2 or len(present_months) >= 3)
                and spread <= SUBSCRIPTION_MAX_AMOUNT_SPREAD_RATIO
            )
            if not qualifies:
                continue
            selected_months = present_months
            selected_amounts = totals
        else:
            selected_months = sorted({
                c['report_month'] for c in selected
            } | {
                c['date'].strftime('%Y-%m') for c in selected
            })
            if month_key not in selected_months:
                continue
            selected_amounts = [c['amount'] for c in selected]
            mean_t = sum(selected_amounts) / len(selected_amounts)
            if mean_t <= 0:
                continue
            spread = (
                (max(selected_amounts) - min(selected_amounts)) / mean_t
                if len(selected_amounts) > 1 else 0.0
            )
            streak = _longest_consecutive_month_streak(selected_months)
            months_active = len(selected_months)
            noisy = _subscription_pattern_too_noisy(charges, selected)

            # Strong patterns: 3+ hits, or 2 consecutive with controlled spread.
            # Day-flexible interval chains already encode date tolerance.
            qualifies = spread <= SUBSCRIPTION_MAX_AMOUNT_SPREAD_RATIO
            if not qualifies:
                continue
            if months_active >= 3 or streak >= 3:
                ok = True
            elif streak >= 2 or len(selected) >= 2:
                # Weak (2-hit) pattern: require day-of-month agreement to cut random coincidences.
                days = [c['date'].day for c in selected]
                med = sorted(days)[len(days) // 2]
                dom_ok = all(
                    _subscription_dom_distance(d, med) <= SUBSCRIPTION_DAY_TOLERANCE
                    for d in days
                )
                ok = dom_ok and not noisy
            else:
                ok = False
            if not ok:
                continue
            present_months = selected_months

        in_window = len([mm for mm in months_seq if mm in present_months])
        # Prefer focal-month charge amount; else month total; else last selected.
        # When two similar charges land in the same month (delayed prior + current),
        # use one representative unit amount — not the combined total.
        last_amt = None
        dated_selected = [c for c in selected if isinstance(c, dict) and isinstance(c.get('date'), date)]
        focal_charges = [
            c for c in dated_selected
            if c.get('report_month') == month_key or c['date'].strftime('%Y-%m') == month_key
        ]
        if not focal_charges:
            # All raw charges for this label in the focal month (may include non-chain rows)
            focal_charges = [
                c for c in _subscription_normalize_charges(charges)
                if c.get('report_month') == month_key or c['date'].strftime('%Y-%m') == month_key
            ]
        if focal_charges:
            last_amt = _representative_monthly_bill_amount(focal_charges)
        elif by_month.get(month_key, 0) > 0.005:
            last_amt = round(float(by_month[month_key]), 2)
        elif selected_amounts:
            last_amt = round(float(selected_amounts[-1]), 2)
        else:
            continue
        if last_amt is None or last_amt < 0.01:
            continue

        prev_amt = None
        if month_key in present_months:
            mi = present_months.index(month_key)
            prev_m = present_months[mi - 1] if mi > 0 else None
        else:
            prev_m = present_months[-1] if present_months else None

        if prev_m:
            prev_rows = [
                c for c in dated_selected
                if c.get('report_month') == prev_m or c['date'].strftime('%Y-%m') == prev_m
            ]
            if prev_rows:
                prev_amt = round(float(prev_rows[-1]['amount']), 2)
            elif by_month.get(prev_m, 0) > 0.005:
                prev_amt = round(float(by_month[prev_m]), 2)
        elif len(dated_selected) >= 2:
            prev_amt = round(float(dated_selected[-2]['amount']), 2)

        if prev_amt is not None and prev_amt > 0:
            if last_amt > prev_amt * 1.05:
                trend = 'up'
            elif last_amt < prev_amt * 0.95:
                trend = 'down'
            else:
                trend = 'flat'
        elif prev_m is not None and prev_amt == 0:
            trend = 'up' if last_amt > 0 else 'flat'
        else:
            trend = 'insufficient_history'

        streak_out = _longest_consecutive_month_streak(present_months)
        out.append({
            'label': label,
            'display_description': label_sample_desc.get(label, label),
            'months_active': len(present_months),
            'months_in_window': in_window,
            'consecutive_streak': streak_out,
            'last_amount': last_amt,
            'total_last_month': last_amt,
            'amount_last_month': last_amt,
            'amount_avg_active_months': round(mean_t, 2),
            'amount_variability': round(spread, 3) if len(selected_amounts) > 1 else 0.0,
            'trend': trend,
        })
    out.sort(key=lambda x: (-x['months_active'], -x['amount_last_month']))
    return out[:20]


def _spending_transfer_reconciliation_for_month(month_rows: list) -> dict:
    """Reconciliation metadata for a month's transactions (paired vs unmatched legs)."""
    pair_ids: set[str] = set()
    for t in month_rows:
        if t.get('transfer_pair_id') and t.get('internal_transfer'):
            pair_ids.add(str(t['transfer_pair_id']))
    pair_count = len(pair_ids)
    paired_out = sum(
        float(t.get('amount', 0)) for t in month_rows
        if t.get('direction') == 'outgoing' and str(t.get('transfer_pair_id') or '') in pair_ids
    )
    paired_in = sum(
        float(t.get('amount', 0)) for t in month_rows
        if t.get('direction') == 'incoming' and str(t.get('transfer_pair_id') or '') in pair_ids
    )
    paired_out = round(paired_out, 2)
    paired_in = round(paired_in, 2)
    leg_mismatch = round(abs(paired_out - paired_in), 2)

    uo = [t for t in month_rows if t.get('direction') == 'outgoing' and not _spending_is_paired_leg(t)]
    ui = [t for t in month_rows if t.get('direction') == 'incoming' and not _spending_is_paired_leg(t)]

    def _preview_list(rows: list) -> list:
        out = []
        for t in sorted(
            rows,
            key=lambda x: (str(x.get('date') or ''), -float(x.get('amount', 0))),
        )[:SPENDING_TRANSFER_UNMATCHED_PREVIEW_MAX]:
            out.append({
                'id': t.get('id'),
                'date': t.get('date'),
                'amount': round(float(t.get('amount', 0)), 2),
                'description': (str(t.get('description') or ''))[:120],
            })
        return out

    return {
        'pair_count': pair_count,
        'internal_transfer_outgoing_total': paired_out,
        'paired_incoming_total': paired_in,
        'paired_leg_amount_mismatch': leg_mismatch,
        'unmatched_outgoing': {
            'count': len(uo),
            'total': round(sum(float(t.get('amount', 0)) for t in uo), 2),
            'items': _preview_list(uo),
        },
        'unmatched_incoming': {
            'count': len(ui),
            'total': round(sum(float(t.get('amount', 0)) for t in ui), 2),
            'items': _preview_list(ui),
        },
    }


def _insight_category_amount_map(insight: dict | None) -> dict[str, float]:
    if not isinstance(insight, dict):
        return {}
    m: dict[str, float] = {}
    for row in insight.get('category_breakdown') or []:
        if not isinstance(row, dict):
            continue
        c = str(row.get('category') or 'unclassified').strip().lower()
        try:
            m[c] = float(row.get('amount', 0) or 0)
        except (TypeError, ValueError):
            m[c] = 0.0
    return m


def _spending_report_months_with_data(txs: list) -> set[str]:
    """Calendar months that have at least one transaction (imported / recorded). Excludes
    'empty' previous months that would otherwise be treated as £0 in category baselines."""
    out: set[str] = set()
    for t in txs:
        rm = _report_month_for_spending_tx(t)
        if rm:
            out.add(str(rm))
    return out


def _outgoing_category_total_for_spending_month(txs: list, month_key: str, category: str) -> float:
    cat_target = str(category).strip().lower()
    month_sum = 0.0
    for tx in txs:
        if _report_month_for_spending_tx(tx) != month_key or tx.get('direction') != 'outgoing':
            continue
        if _spending_excluded_from_insight_metrics(tx):
            continue
        tx_cat = str(tx.get('category') or 'unclassified').strip().lower()
        if tx_cat not in SPENDING_CATEGORY_SET:
            tx_cat = 'unclassified'
        if tx_cat == cat_target:
            month_sum += float(tx.get('amount', 0))
    return month_sum


def _trailing_income_outgoing_averages(
    spending: dict, month_key: str, max_prior_months: int
) -> tuple[float | None, float | None, int]:
    imap = spending.get('monthly_insights') or {}
    incomes: list[float] = []
    outgoings: list[float] = []
    for i in range(1, max_prior_months + 1):
        m = _month_prev(month_key, i)
        if not m:
            break
        ins = imap.get(m)
        if not isinstance(ins, dict):
            continue
        try:
            incomes.append(float(ins.get('income_total', 0) or 0))
            outgoings.append(float(ins.get('outgoing_total', 0) or 0))
        except (TypeError, ValueError):
            continue
    n = len(outgoings)
    if n == 0:
        return None, None, 0
    return sum(incomes) / n, sum(outgoings) / n, n


def _build_budget_action_items(
    income: float,
    outgoing: float,
    net: float,
    savings_rate: float | None,
    prev_savings_rate: float | None,
    savings_rate_delta: float | None,
    outgoing_delta: float | None,
    net_delta: float | None,
    outgoing_vs_trailing_6m_pct: float | None,
    anomalies: list[dict],
    subscription_signals: list[dict],
    top_transactions_pareto_pct: float | None,
    category_trends: list[dict],
) -> list[dict]:
    items: list[dict] = []
    if net < 0:
        items.append({
            'kind': 'net_negative',
            'title': 'This month is cash-flow negative',
            'detail': f'Outgoings exceeded income by {abs(net):.2f} after non-transfer items.',
            'amount_hint': round(abs(net), 2),
        })
    if net_delta is not None and net_delta < 0 and net >= 0:
        items.append({
            'kind': 'net_tightening',
            'title': 'Net savings fell vs last month',
            'detail': f'Net is £{net_delta:.2f} lower than the previous month — check higher outgoings or lower income.',
            'amount_hint': round(abs(net_delta), 2),
        })
    if outgoing_delta is not None and outgoing_delta > 0 and len(items) < 6:
        items.append({
            'kind': 'outgoing_up',
            'title': 'Outgoing is higher than last month',
            'detail': f'Up £{outgoing_delta:.2f} vs the previous month — see category trends and largest lines.',
            'amount_hint': round(outgoing_delta, 2),
        })
    if (
        outgoing_vs_trailing_6m_pct is not None
        and outgoing_vs_trailing_6m_pct > 8.0
        and len(items) < 6
    ):
        items.append({
            'kind': 'above_trailing_baseline',
            'title': 'Well above your recent spending norm',
            'detail': f'Outgoing is about {outgoing_vs_trailing_6m_pct:.0f}% higher than your trailing 6-month average (where data exists).',
            'amount_hint': None,
        })
    for a in (anomalies or [])[:2]:
        if len(items) >= 6:
            break
        cat = a.get('category', '')
        n_base = a.get('baseline_months')
        if isinstance(n_base, int) and n_base > 0:
            base_phrase = (
                f'the average over {n_base} prior month(s) with recorded data'
                if n_base != 1
                else 'the prior month with recorded data'
            )
        else:
            base_phrase = 'the recent baseline for this category'
        items.append({
            'kind': 'category_spike',
            'title': f'"{cat}" is unusually high',
            'detail': f'Spent {a.get("delta_pct", 0):.0f}% more than {base_phrase} — worth a look.',
            'amount_hint': a.get('amount'),
        })
    if (
        savings_rate is not None
        and prev_savings_rate is not None
        and savings_rate_delta is not None
        and savings_rate_delta < -2.0
        and len(items) < 5
    ):
        items.append({
            'kind': 'savings_rate_slip',
            'title': 'Savings rate dropped vs last month',
            'detail': f'From {prev_savings_rate:.1f}% to {savings_rate:.1f}% of income kept after outgoings.',
            'amount_hint': None,
        })
    for sig in sorted(
        (subscription_signals or []),
        key=lambda s: -float(s.get('total_last_month') or s.get('last_amount') or 0),
    ):
        if len(items) >= 6:
            break
        if str(sig.get('trend') or '') != 'up':
            continue
        name = (sig.get('display_description') or sig.get('label') or 'A recurring charge')[:80]
        amt = float(sig.get('total_last_month') or sig.get('last_amount') or 0)
        items.append({
            'kind': 'subscription_up',
            'title': f'Recurring: {name}',
            'detail': 'This charge went up compared with the last time you paid it.',
            'amount_hint': round(amt, 2),
        })
    spike_cats = {str(a.get('category') or '').lower() for a in (anomalies or [])}
    for ct in (category_trends or []):
        if len(items) >= 6:
            break
        cname = str(ct.get('category') or '')[:64]
        if cname.lower() in spike_cats:
            continue
        vs3 = ct.get('vs_3m_pct')
        if vs3 is None or vs3 <= 20:
            continue
        items.append({
            'kind': 'category_vs_3m',
            'title': f'"{cname}" is up vs 3-month average',
            'detail': f'About {vs3:.0f}% more than the average in prior months (with data) for this category.',
            'amount_hint': ct.get('amount'),
        })
    if (
        top_transactions_pareto_pct is not None
        and top_transactions_pareto_pct >= 50.0
        and len(items) < 6
    ):
        items.append({
            'kind': 'pareto_concentration',
            'title': 'A few big purchases dominate spend',
            'detail': f'Your 5 largest outgoings are about {top_transactions_pareto_pct:.0f}% of total — review the largest list.',
            'amount_hint': None,
        })
    out: list[dict] = []
    seen: set[str] = set()
    for it in items:
        key = f"{it['kind']}:{it.get('title', '')[:50]}"
        if key in seen:
            continue
        seen.add(key)
        out.append(it)
    return out[:6]


def _compute_monthly_insight(spending: dict, month_key: str) -> dict:
    txs = spending.get('transactions') or []
    months_with_data = _spending_report_months_with_data(txs)
    month_rows = [t for t in txs if _report_month_for_spending_tx(t) == month_key]
    transfer_reconciliation = _spending_transfer_reconciliation_for_month(month_rows)
    # KPIs: exclude internal transfer legs and user-excluded rows
    income_rows = [
        t for t in month_rows
        if t.get('direction') == 'incoming' and not _spending_excluded_from_insight_metrics(t)
    ]
    incoming = round(sum(float(t.get('amount', 0)) for t in income_rows), 2)
    outgoing_rows = [
        t for t in month_rows
        if t.get('direction') == 'outgoing' and not _spending_excluded_from_insight_metrics(t)
    ]
    outgoing = round(sum(float(t.get('amount', 0)) for t in outgoing_rows), 2)
    net = round(incoming - outgoing, 2)
    savings_rate = round((net / incoming) * 100, 2) if incoming > 0 else None

    category_totals = defaultdict(float)
    merchant_totals = defaultdict(float)
    for tx in outgoing_rows:
        cat = str(tx.get('category') or 'unclassified').strip().lower()
        if cat not in SPENDING_CATEGORY_SET:
            cat = 'unclassified'
        category_totals[cat] += float(tx.get('amount', 0))
        merchant_key = _normalize_label(tx.get('description', ''))[:80] or 'unknown'
        merchant_totals[merchant_key] += float(tx.get('amount', 0))

    categories = [
        {'category': k, 'amount': round(v, 2)}
        for k, v in category_totals.items()
    ]
    categories.sort(key=lambda x: x['amount'], reverse=True)
    merchants = [
        {'merchant': k, 'amount': round(v, 2)}
        for k, v in merchant_totals.items()
    ]
    merchants.sort(key=lambda x: x['amount'], reverse=True)

    prev_month = _month_prev(month_key, 1)
    prev_insight = (spending.get('monthly_insights') or {}).get(prev_month) if prev_month else None
    outgoing_delta = None
    net_delta = None
    income_delta = None
    prev_savings_rate: float | None = None
    savings_rate_delta: float | None = None
    if isinstance(prev_insight, dict):
        try:
            outgoing_delta = round(outgoing - float(prev_insight.get('outgoing_total', 0)), 2)
            net_delta = round(net - float(prev_insight.get('net', 0)), 2)
            income_delta = round(incoming - float(prev_insight.get('income_total', 0)), 2)
        except (TypeError, ValueError):
            outgoing_delta = None
            net_delta = None
            income_delta = None
        try:
            psr = prev_insight.get('savings_rate')
            prev_savings_rate = float(psr) if psr is not None else None
        except (TypeError, ValueError):
            prev_savings_rate = None
        if savings_rate is not None and prev_savings_rate is not None:
            savings_rate_delta = round(savings_rate - prev_savings_rate, 2)

    inc3, out3, n3 = _trailing_income_outgoing_averages(spending, month_key, 3)
    inc6, out6, n6 = _trailing_income_outgoing_averages(spending, month_key, 6)
    outgoing_vs_trailing_6m_pct: float | None = None
    if out6 and out6 > 0.005:
        outgoing_vs_trailing_6m_pct = round((outgoing - out6) / out6 * 100.0, 1)
    income_vs_trailing_6m_pct: float | None = None
    if inc6 and inc6 > 0.005:
        income_vs_trailing_6m_pct = round((incoming - inc6) / inc6 * 100.0, 1)

    recurring = []
    count_by_desc = defaultdict(int)
    for tx in outgoing_rows:
        count_by_desc[_normalize_label(tx.get('description', ''))] += 1
    for desc, count in count_by_desc.items():
        if desc and count >= 2:
            recurring.append({'description': desc, 'occurrences': count})

    min_anomaly = max(
        SPENDING_ANOMALY_MIN_GBP,
        outgoing * SPENDING_ANOMALY_MIN_OUTGOING_PCT,
    ) if outgoing > 0 else SPENDING_ANOMALY_MIN_GBP
    prev_cat_map = _insight_category_amount_map(prev_insight if isinstance(prev_insight, dict) else None)
    anomalies = []
    for cat_row in categories:
        cat = cat_row['category']
        # Baseline: mean of *monthly* category spend (totals) for up to 3 previous calendar
        # months, not the mean of individual transaction lines. Only include months that
        # already have at least one recorded transaction — do not treat "no data yet" as £0.
        monthly_totals: list[float] = []
        for i in range(1, 4):
            m = _month_prev(month_key, i)
            if not m or m not in months_with_data:
                continue
            monthly_totals.append(_outgoing_category_total_for_spending_month(txs, m, cat))
        if monthly_totals:
            avg = sum(monthly_totals) / len(monthly_totals)
            amt = float(cat_row['amount'])
            if (
                avg > 0
                and amt > (avg * 1.5)
                and amt >= min_anomaly
            ):
                d_pct = round((amt - avg) / avg * 100.0, 1) if avg > 0 else 0.0
                anomalies.append({
                    'kind': 'category_spike',
                    'category': cat,
                    'amount': round(amt, 2),
                    'baseline_avg': round(avg, 2),
                    'delta_pct': d_pct,
                    'baseline_months': len(monthly_totals),
                })

    largest_src = sorted(outgoing_rows, key=lambda x: -float(x.get('amount', 0)))[:10]
    largest_outgoing = [
        {
            'id': t.get('id'),
            'date': t.get('date'),
            'description': str(t.get('description', ''))[:200],
            'amount': round(float(t.get('amount', 0)), 2),
            'category': str(t.get('category') or 'unclassified'),
        }
        for t in largest_src
    ]

    cat_total_sum = sum(category_totals.values())
    category_breakdown = [
        {
            'category': k,
            'amount': round(v, 2),
            'pct_of_outgoing': round(100.0 * v / cat_total_sum, 1) if cat_total_sum > 0 else 0.0,
        }
        for k, v in sorted(category_totals.items(), key=lambda x: -x[1])
    ]

    category_trends: list[dict] = []
    for cat_row in categories[:6]:
        cat = cat_row['category']
        amt = float(cat_row['amount'])
        prev_amt = float(prev_cat_map.get(str(cat).lower(), 0.0))
        mom_delta = round(amt - prev_amt, 2) if prev_month else None
        mom_delta_pct: float | None = None
        if prev_amt and prev_amt > 0.005:
            mom_delta_pct = round((amt - prev_amt) / prev_amt * 100.0, 1)
        hist_cat_amt: list[float] = []
        for i in range(1, 4):
            m = _month_prev(month_key, i)
            if not m or m not in months_with_data:
                continue
            hist_cat_amt.append(_outgoing_category_total_for_spending_month(txs, m, cat))
        vs_3m_avg: float | None
        if hist_cat_amt:
            vs_3m_avg = sum(hist_cat_amt) / len(hist_cat_amt)
        else:
            vs_3m_avg = None
        vs_3m_pct: float | None = None
        if vs_3m_avg and vs_3m_avg > 0.005:
            vs_3m_pct = round((amt - vs_3m_avg) / vs_3m_avg * 100.0, 1)
        category_trends.append({
            'category': cat,
            'amount': round(amt, 2),
            'mom_delta': mom_delta,
            'mom_delta_pct': mom_delta_pct,
            'trailing_3m_avg': round(vs_3m_avg, 2) if vs_3m_avg is not None else None,
            'vs_3m_pct': vs_3m_pct,
            'trailing_3m_months_count': len(hist_cat_amt),
        })

    sorted_out_amounts = sorted((float(t.get('amount', 0)) for t in outgoing_rows), reverse=True)
    top5_sum = sum(sorted_out_amounts[:5])
    top_transactions_pareto_pct = (
        round(100.0 * top5_sum / outgoing, 1) if outgoing > 0 else None
    )

    subscription_signals = _subscription_signals_for_month(spending, month_key)

    budget_action_items = _build_budget_action_items(
        incoming,
        outgoing,
        net,
        savings_rate,
        prev_savings_rate,
        savings_rate_delta,
        outgoing_delta,
        net_delta,
        outgoing_vs_trailing_6m_pct,
        anomalies,
        subscription_signals,
        top_transactions_pareto_pct,
        category_trends,
    )

    return {
        'month': month_key,
        'transaction_count': len(month_rows),
        'income_total': incoming,
        'outgoing_total': outgoing,
        'net': net,
        'savings_rate': savings_rate,
        'savings_rate_delta_vs_prev_month': savings_rate_delta,
        'top_categories': categories[:6],
        'top_merchants': merchants[:6],
        'outgoing_delta_vs_prev_month': outgoing_delta,
        'income_delta_vs_prev_month': income_delta,
        'net_delta_vs_prev_month': net_delta,
        'income_trailing_avg_3m': None if inc3 is None else round(inc3, 2),
        'outgoing_trailing_avg_3m': None if out3 is None else round(out3, 2),
        'trailing_3m_months_count': n3,
        'income_trailing_avg_6m': None if inc6 is None else round(inc6, 2),
        'outgoing_trailing_avg_6m': None if out6 is None else round(out6, 2),
        'trailing_6m_months_count': n6,
        'outgoing_vs_trailing_6m_pct': outgoing_vs_trailing_6m_pct,
        'income_vs_trailing_6m_pct': income_vs_trailing_6m_pct,
        'category_trends': category_trends,
        'anomalies': anomalies[:6],
        'recurring_candidates': recurring[:8],
        'largest_outgoing': largest_outgoing,
        'category_breakdown': category_breakdown,
        'top_transactions_pareto_pct': top_transactions_pareto_pct,
        'subscription_signals': subscription_signals,
        'budget_action_items': budget_action_items,
        'transfer_reconciliation': transfer_reconciliation,
        'updated_at': datetime.utcnow().isoformat() + 'Z',
    }


def _recompute_monthly_insights(spending: dict, months: set[str] | None = None) -> None:
    txs = spending.get('transactions') or []
    insight_map = spending.get('monthly_insights') or {}
    if months is None:
        months = {str(_report_month_for_spending_tx(t)) for t in txs if _report_month_for_spending_tx(t)}
        for k in list(insight_map.keys()):
            if k not in months:
                insight_map.pop(k, None)
    for m in sorted(months):
        if not m:
            continue
        has_rows = any(_report_month_for_spending_tx(t) == m for t in txs)
        if not has_rows:
            insight_map.pop(m, None)
            continue
        insight_map[m] = _compute_monthly_insight(spending, m)
    spending['monthly_insights'] = insight_map

# --- Daily Budget ----------------------------------------------------------------------------


def _daily_budget_is_discretionary_tx(t: dict) -> bool:
    """Outgoing non-bill spend that counts against the daily discretionary pool."""
    if t.get('direction') != 'outgoing':
        return False
    if _spending_excluded_from_insight_metrics(t):
        return False
    cat = str(t.get('category') or 'other').strip().lower()
    if cat in DAILY_BUDGET_BILL_CATEGORIES or cat in DAILY_BUDGET_IGNORE_CATEGORIES:
        return False
    return True


def _daily_budget_plan_figures(plan: dict) -> dict:
    try:
        income = max(0.0, float(plan.get('income_monthly') or 0))
    except (TypeError, ValueError):
        income = 0.0
    bill_items = plan.get('bill_items') if isinstance(plan.get('bill_items'), list) else []
    included_items = [b for b in bill_items if isinstance(b, dict) and b.get('included', True)]
    if included_items:
        bills = 0.0
        for b in included_items:
            try:
                bills += max(0.0, float(b.get('amount') or 0))
            except (TypeError, ValueError):
                continue
        bills = round(bills, 2)
    else:
        try:
            bills = max(0.0, float(plan.get('bills_monthly') or 0))
        except (TypeError, ValueError):
            bills = 0.0
    try:
        savings_percent = float(plan.get('savings_percent') or 0)
    except (TypeError, ValueError):
        savings_percent = 0.0
    savings_percent = max(0.0, min(100.0, savings_percent))
    savings_monthly = round(income * savings_percent / 100.0, 2)
    discretionary = round(max(0.0, income - bills - savings_monthly), 2)
    mode = str(plan.get('daily_mode') or 'envelope')
    if mode not in DAILY_BUDGET_MODES:
        mode = 'envelope'
    priority = str(plan.get('underspend_priority') or 'debt_first')
    if priority not in DAILY_BUDGET_UNDERSPEND_PRIORITIES:
        priority = 'debt_first'
    return {
        'income_monthly': round(income, 2),
        'bills_monthly': round(bills, 2),
        'savings_percent': round(savings_percent, 2),
        'savings_monthly': savings_monthly,
        'discretionary_monthly': discretionary,
        'daily_mode': mode,
        'underspend_priority': priority,
    }


def _daily_budget_spend_by_date(
    spending: dict,
    start: date,
    end: date,
) -> dict[str, float]:
    by_day: dict[str, float] = defaultdict(float)
    for t in spending.get('transactions') or []:
        if not _daily_budget_is_discretionary_tx(t):
            continue
        d = _parse_iso_date(str(t.get('date') or '')[:10])
        if d is None or d < start or d > end:
            continue
        try:
            by_day[d.isoformat()] += float(t.get('amount') or 0)
        except (TypeError, ValueError):
            continue
    return {k: round(v, 2) for k, v in by_day.items()}


def _daily_budget_parse_tracking_from(raw) -> date | None:
    if raw in (None, ''):
        return None
    return _parse_iso_date(str(raw))


def _daily_budget_parse_pay_day(raw) -> int:
    """Payday as day-of-month 1–31. Invalid / missing → 1 (calendar month)."""
    try:
        day = int(raw)
    except (TypeError, ValueError):
        return 1
    if day < 1 or day > 31:
        return 1
    return day


def _daily_budget_clamp_dom(year: int, month: int, day: int) -> date:
    last = monthrange(year, month)[1]
    return date(year, month, min(max(1, day), last))


def _daily_budget_pay_period(as_of: date, pay_day: int) -> tuple[date, date]:
    """Inclusive pay period containing ``as_of`` for a recurring monthly payday.

    Period runs from payday through the day before the next payday. Days that
    don't exist in a short month (e.g. 31 in February) clamp to the month's last
    day.
    """
    pay_day = _daily_budget_parse_pay_day(pay_day)
    start_this = _daily_budget_clamp_dom(as_of.year, as_of.month, pay_day)
    if as_of >= start_this:
        period_start = start_this
        next_month = as_of + relativedelta(months=1)
        next_start = _daily_budget_clamp_dom(next_month.year, next_month.month, pay_day)
        period_end = next_start - timedelta(days=1)
    else:
        prev_month = as_of + relativedelta(months=-1)
        period_start = _daily_budget_clamp_dom(prev_month.year, prev_month.month, pay_day)
        period_end = start_this - timedelta(days=1)
    return period_start, period_end


def _daily_budget_pacing_start(
    plan: dict,
    period_start: date,
    period_end: date,
) -> date:
    """First day that counts for daily pacing in this pay period.

    ``tracking_from`` excludes earlier days in the period so missing data is not
    treated as £0 spend. If tracking began before this period, pace from
    ``period_start``.
    """
    tf = _daily_budget_parse_tracking_from(plan.get('tracking_from') if isinstance(plan, dict) else None)
    if tf is None:
        return period_start
    if tf < period_start:
        return period_start
    if tf > period_end:
        return period_end
    return tf


def _daily_budget_goals_capacity(goals) -> float:
    total = 0.0
    if not isinstance(goals, list):
        return 0.0
    for g in goals:
        if not isinstance(g, dict):
            continue
        try:
            total += max(0.0, float(g.get('target_amount') or 0))
        except (TypeError, ValueError):
            continue
    return round(total, 2)


def _daily_budget_allocate_day_leftover(
    raw: float,
    *,
    debt_left: float,
    goals_filled: float,
    goals_capacity: float,
    priority: str,
) -> tuple[float, float, float]:
    """Split one day's leftover into (repay, roll_forward, goals_filled_delta).

    Debt repay is the only physical skim (reduces carry). ``roll_forward`` is also
    the amount that counts toward underspend/goals after skimming.
    """
    raw = max(0.0, round(float(raw or 0), 2))
    debt_left = max(0.0, round(float(debt_left or 0), 2))
    goals_filled = max(0.0, round(float(goals_filled or 0), 2))
    goals_capacity = max(0.0, round(float(goals_capacity or 0), 2))
    if priority not in DAILY_BUDGET_UNDERSPEND_PRIORITIES:
        priority = 'debt_first'

    if priority == 'goals_first' and goals_capacity > 0:
        room = max(0.0, round(goals_capacity - goals_filled, 2))
        protected = min(raw, room)
        repay = min(round(raw - protected, 2), debt_left)
        filled_delta = protected
    else:
        repay = min(raw, debt_left)
        filled_delta = round(raw - repay, 2)

    roll = round(raw - repay, 2)
    return round(repay, 2), roll, round(filled_delta, 2)


def _daily_budget_day_limits(
    figures: dict,
    spend_by_date: dict[str, float],
    period_start: date,
    period_end: date,
    pacing_start: date | None = None,
    *,
    debt_balance: float = 0.0,
    goals_capacity: float = 0.0,
    underspend_priority: str | None = None,
) -> dict[str, float]:
    """Effective daily spending allowance for each day in the pay period.

    When ``pacing_start`` is after ``period_start``, the remaining discretionary
    pool is pro-rated by days left in the period (average-pace assumption for
    unknown pre-tracking spend). Pre-start days get a 0 limit and are ignored
    by status.

    For ``carry_surplus``, overspend-debt skims reduce what rolls to tomorrow
    (debt first or after goals capacity, per ``underspend_priority``).
    """
    if period_end < period_start:
        return {}
    days_in_period = (period_end - period_start).days + 1
    discretionary = float(figures.get('discretionary_monthly') or 0)
    base = round(discretionary / days_in_period, 2) if days_in_period else 0.0
    mode = figures.get('daily_mode') or 'envelope'
    start = pacing_start or period_start
    if start < period_start:
        start = period_start
    if start > period_end:
        start = period_end
    days_in_window = (period_end - start).days + 1
    # Assume average pace before tracking started so the join-day envelope is
    # ~base, not (full pool ÷ days left).
    window_pool = (
        round(discretionary * days_in_window / days_in_period, 2) if days_in_period else 0.0
    )
    assumed_prior = round(max(0.0, discretionary - window_pool), 2)
    limits: dict[str, float] = {}
    priority = underspend_priority or figures.get('underspend_priority') or 'debt_first'
    if priority not in DAILY_BUDGET_UNDERSPEND_PRIORITIES:
        priority = 'debt_first'

    if mode == 'fixed':
        d = period_start
        while d <= period_end:
            limits[d.isoformat()] = base if d >= start else 0.0
            d += timedelta(days=1)
        return limits

    if mode == 'carry_surplus':
        leftover = 0.0
        debt_left = max(0.0, round(float(debt_balance or 0), 2))
        goals_filled = 0.0
        d = period_start
        while d <= period_end:
            if d < start:
                limits[d.isoformat()] = 0.0
                d += timedelta(days=1)
                continue
            allowance = round(base + leftover, 2)
            limits[d.isoformat()] = allowance
            spent = float(spend_by_date.get(d.isoformat(), 0) or 0)
            raw = max(0.0, round(allowance - spent, 2))
            repay, roll, filled_delta = _daily_budget_allocate_day_leftover(
                raw,
                debt_left=debt_left,
                goals_filled=goals_filled,
                goals_capacity=goals_capacity,
                priority=priority,
            )
            debt_left = round(max(0.0, debt_left - repay), 2)
            goals_filled = round(goals_filled + filled_delta, 2)
            leftover = roll
            d += timedelta(days=1)
        return limits

    # envelope: for each day, remaining pool after prior spend ÷ remaining days (incl. that day)
    spent_prior = assumed_prior
    d = period_start
    while d <= period_end:
        if d < start:
            limits[d.isoformat()] = 0.0
            d += timedelta(days=1)
            continue
        days_left = (period_end - d).days + 1
        remaining_pool = max(0.0, round(discretionary - spent_prior, 2))
        day_limit = round(remaining_pool / days_left, 2) if days_left else 0.0
        limits[d.isoformat()] = day_limit
        spent_prior += float(spend_by_date.get(d.isoformat(), 0) or 0)
        d += timedelta(days=1)
    return limits


def _daily_budget_pace_projection(
    figures: dict,
    spend_by_date: dict[str, float],
    period_start: date,
    period_end: date,
    pacing_start: date,
    today: date,
    daily_limit: float,
    *,
    debt_balance: float = 0.0,
    goals_capacity: float = 0.0,
) -> dict:
    """How MTD discretionary spend reshapes the rest-of-period daily projection."""
    days_in_period = (period_end - period_start).days + 1
    discretionary = float(figures.get('discretionary_monthly') or 0)
    base = round(discretionary / days_in_period, 2) if days_in_period else 0.0
    mode = figures.get('daily_mode') or 'envelope'
    if mode not in DAILY_BUDGET_MODES:
        mode = 'envelope'

    start = pacing_start or period_start
    if start < period_start:
        start = period_start
    if start > period_end:
        start = period_end

    days_in_window = (period_end - start).days + 1
    window_pool = (
        round(discretionary * days_in_window / days_in_period, 2) if days_in_period else 0.0
    )
    assumed_prior = round(max(0.0, discretionary - window_pool), 2)

    spent_before_today = 0.0
    d = start
    while d < today and d <= period_end:
        spent_before_today += float(spend_by_date.get(d.isoformat(), 0) or 0)
        d += timedelta(days=1)
    spent_before_today = round(spent_before_today, 2)

    spent_today = float(spend_by_date.get(today.isoformat(), 0) or 0) if start <= today <= period_end else 0.0
    spent_so_far = round(spent_before_today + spent_today, 2)

    if today < period_start or today > period_end:
        days_left = 0
        days_after_today = 0
        days_elapsed = 0
    else:
        days_left = (period_end - today).days + 1
        days_after_today = max(0, days_left - 1)
        # Elapsed days in the pacing window through today (inclusive).
        elapsed_start = start if start <= today else today
        days_elapsed = (today - elapsed_start).days + 1 if elapsed_start <= today else 0

    pool_at_start_of_today = round(
        max(0.0, discretionary - assumed_prior - spent_before_today),
        2,
    )
    remaining_after_today = round(max(0.0, window_pool - spent_so_far), 2)

    if mode == 'fixed':
        projected_daily = base
    elif days_after_today > 0:
        projected_daily = round(remaining_after_today / days_after_today, 2)
    else:
        # Last day of the period: projection is whatever remains today.
        projected_daily = round(max(0.0, float(daily_limit or 0) - spent_today), 2)

    pace_target_spend = round(base * days_elapsed, 2) if days_elapsed else 0.0
    pace_delta = round(pace_target_spend - spent_so_far, 2)
    projected_vs_base = round(projected_daily - base, 2)

    carry_from_yesterday = 0.0
    if mode == 'carry_surplus' and today > start:
        yesterday = today - timedelta(days=1)
        if yesterday >= start:
            priority = figures.get('underspend_priority') or 'debt_first'
            if priority not in DAILY_BUDGET_UNDERSPEND_PRIORITIES:
                priority = 'debt_first'
            debt_left = max(0.0, round(float(debt_balance or 0), 2))
            capacity = max(0.0, round(float(goals_capacity or 0), 2))
            leftover = 0.0
            goals_filled = 0.0
            d = start
            while d <= yesterday:
                allowance = round(base + leftover, 2)
                spent = float(spend_by_date.get(d.isoformat(), 0) or 0)
                raw = max(0.0, round(allowance - spent, 2))
                repay, roll, filled_delta = _daily_budget_allocate_day_leftover(
                    raw,
                    debt_left=debt_left,
                    goals_filled=goals_filled,
                    goals_capacity=capacity,
                    priority=priority,
                )
                debt_left = round(max(0.0, debt_left - repay), 2)
                goals_filled = round(goals_filled + filled_delta, 2)
                leftover = roll
                d += timedelta(days=1)
            carry_from_yesterday = leftover

    return {
        'mode': mode,
        'discretionary_monthly': round(discretionary, 2),
        'window_pool': window_pool,
        'assumed_prior_spend': assumed_prior,
        'spent_before_today': spent_before_today,
        'spent_so_far': spent_so_far,
        'pool_at_start_of_today': pool_at_start_of_today,
        'remaining_after_today': remaining_after_today,
        'days_in_period': days_in_period,
        'days_elapsed': days_elapsed,
        'days_left': days_left,
        'days_after_today': days_after_today,
        'base_daily': base,
        'pace_target_spend': pace_target_spend,
        'pace_delta': pace_delta,
        'projected_daily': projected_daily,
        'projected_vs_base': projected_vs_base,
        'carry_from_yesterday': round(carry_from_yesterday, 2),
        'daily_limit': round(float(daily_limit or 0), 2),
        'mid_period_start': start > period_start,
    }


# Daily budget is a UK calendar product (£ / en-GB UI). Use London civil dates so
# "today" matches the browser after midnight BST/GMT even when the host is UTC.
_DAILY_BUDGET_TZ = ZoneInfo('Europe/London')


def _daily_budget_today() -> date:
    """Calendar 'today' for daily budget (Europe/London; patchable in tests)."""
    return datetime.now(_DAILY_BUDGET_TZ).date()


def _daily_budget_period_key(period_start: date, period_end: date) -> str:
    return f'{period_start.isoformat()}_{period_end.isoformat()}'


def _daily_budget_previous_pay_period(period_start: date, pay_day: int) -> tuple[date, date]:
    """Pay period that ended the day before ``period_start``."""
    return _daily_budget_pay_period(period_start - timedelta(days=1), pay_day)


def _daily_budget_period_window_figures(
    figures: dict,
    plan: dict,
    period_start: date,
    period_end: date,
) -> tuple[date, float, int]:
    """Return (pacing_start, window_pool, days_in_period) for a pay period."""
    days_in_period = (period_end - period_start).days + 1
    pacing_start = _daily_budget_pacing_start(plan, period_start, period_end)
    discretionary = float(figures.get('discretionary_monthly') or 0)
    days_in_window = (period_end - pacing_start).days + 1
    if pacing_start == period_start or not days_in_period:
        window_pool = discretionary
    else:
        window_pool = round(discretionary * days_in_window / days_in_period, 2)
    return pacing_start, window_pool, days_in_period


def _daily_budget_period_net_overspend(
    spending: dict,
    plan: dict,
    figures: dict,
    period_start: date,
    period_end: date,
) -> dict:
    """Net overspend for a completed pay period (spent − discretionary window)."""
    pacing_start, window_pool, _days = _daily_budget_period_window_figures(
        figures, plan, period_start, period_end,
    )
    spend_by_date = _daily_budget_spend_by_date(spending, period_start, period_end)
    spent = 0.0
    d = pacing_start
    while d <= period_end:
        spent += float(spend_by_date.get(d.isoformat(), 0) or 0)
        d += timedelta(days=1)
    spent = round(spent, 2)
    net = round(max(0.0, spent - window_pool), 2)
    return {
        'period_start': period_start.isoformat(),
        'period_end': period_end.isoformat(),
        'pacing_start': pacing_start.isoformat(),
        'discretionary': round(window_pool, 2),
        'spent': spent,
        'net_overspend': net,
    }


def _daily_budget_debt_open_balance(debt) -> float:
    if not isinstance(debt, dict):
        return 0.0
    try:
        open_bal = float(debt.get('balance_at_period_open'))
    except (TypeError, ValueError):
        open_bal = None
    if open_bal is None:
        try:
            open_bal = float(debt.get('balance') or 0)
        except (TypeError, ValueError):
            open_bal = 0.0
    return max(0.0, round(open_bal, 2))


def _daily_budget_skim_totals(
    limits: dict[str, float],
    spend_by_date: dict[str, float],
    pacing_start: date,
    through: date,
    period_end: date,
    *,
    debt_open: float,
    goals_capacity: float,
    priority: str,
) -> dict:
    """Sum debt repay + post-skim underspend from daily leftovers through ``through``."""
    debt_left = max(0.0, round(float(debt_open or 0), 2))
    goals_filled = 0.0
    repaid = 0.0
    saved = 0.0
    end = min(through, period_end)
    d = pacing_start
    while d <= end:
        key = d.isoformat()
        lim = float(limits.get(key, 0) or 0)
        spent = float(spend_by_date.get(key, 0) or 0)
        raw = max(0.0, round(lim - spent, 2))
        repay, roll, filled_delta = _daily_budget_allocate_day_leftover(
            raw,
            debt_left=debt_left,
            goals_filled=goals_filled,
            goals_capacity=goals_capacity,
            priority=priority,
        )
        debt_left = round(max(0.0, debt_left - repay), 2)
        goals_filled = round(goals_filled + filled_delta, 2)
        repaid = round(repaid + repay, 2)
        saved = round(saved + roll, 2)
        d += timedelta(days=1)
    return {
        'repaid': repaid,
        'goals_pot': saved,
        'balance': debt_left,
    }


def _daily_budget_sync_overspend_state(spending: dict, as_of: date) -> bool:
    """Roll debt open-balance when the active pay period changes.

    Returns True when the bucket was mutated and should be saved.
    """
    bucket, changed = _ensure_daily_budget(spending)
    debt = bucket.get('overspend_debt')
    if not isinstance(debt, dict):
        return changed

    open_bal = _daily_budget_debt_open_balance(debt)
    if open_bal <= 0 and float(debt.get('repaid_total') or 0) <= 0:
        # Keep structure but nothing to sync.
        return changed

    plan = bucket.get('plan') or {}
    figures = _daily_budget_plan_figures(plan)
    pay_day = _daily_budget_parse_pay_day(plan.get('pay_day'))
    period_start, period_end = _daily_budget_pay_period(as_of, pay_day)
    active_raw = str(debt.get('active_period_start') or '').strip()[:10]
    active_start = _parse_iso_date(active_raw) if active_raw else None
    if active_start is None:
        debt['active_period_start'] = period_start.isoformat()
        debt['balance_at_period_open'] = open_bal
        debt['balance'] = open_bal
        return True

    if active_start == period_start:
        return changed

    # Finalize skims for each intervening period up to the current one.
    cursor = active_start
    bal = open_bal
    goals_capacity = _daily_budget_goals_capacity(bucket.get('goals'))
    priority = figures.get('underspend_priority') or 'debt_first'
    safety = 0
    while cursor < period_start and bal > 0 and safety < 36:
        safety += 1
        cur_start, cur_end = _daily_budget_pay_period(cursor, pay_day)
        pacing_start = _daily_budget_pacing_start(plan, cur_start, cur_end)
        spend_by_date = _daily_budget_spend_by_date(spending, cur_start, cur_end)
        limits = _daily_budget_day_limits(
            figures,
            spend_by_date,
            cur_start,
            cur_end,
            pacing_start=pacing_start,
            debt_balance=bal,
            goals_capacity=goals_capacity,
            underspend_priority=priority,
        )
        skim = _daily_budget_skim_totals(
            limits,
            spend_by_date,
            pacing_start,
            cur_end,
            cur_end,
            debt_open=bal,
            goals_capacity=goals_capacity,
            priority=priority,
        )
        repaid = float(skim['repaid'])
        bal = round(max(0.0, bal - repaid), 2)
        debt['repaid_total'] = round(float(debt.get('repaid_total') or 0) + repaid, 2)
        cursor = cur_end + timedelta(days=1)

    debt['balance_at_period_open'] = bal
    debt['balance'] = bal
    debt['active_period_start'] = period_start.isoformat()
    if bal <= 0:
        debt['balance_at_period_open'] = 0.0
        debt['balance'] = 0.0
    return True


def _daily_budget_overspend_prompt(
    spending: dict,
    plan: dict,
    figures: dict,
    period_start: date,
    as_of: date,
) -> dict | None:
    """Pending opt-in when the previous pay period finished net-over."""
    if as_of < period_start:
        return None
    pay_day = _daily_budget_parse_pay_day(plan.get('pay_day'))
    prev_start, prev_end = _daily_budget_previous_pay_period(period_start, pay_day)
    if as_of <= prev_end:
        return None
    # Ignore periods entirely before tracking began.
    tracking_from = _daily_budget_parse_tracking_from(plan.get('tracking_from'))
    if tracking_from is not None and prev_end < tracking_from:
        return None
    bucket = spending.get('daily_budget') or {}
    decisions = bucket.get('overspend_decisions') if isinstance(bucket.get('overspend_decisions'), dict) else {}
    key = _daily_budget_period_key(prev_start, prev_end)
    if key in decisions:
        return None
    summary = _daily_budget_period_net_overspend(
        spending, plan, figures, prev_start, prev_end,
    )
    if float(summary.get('net_overspend') or 0) <= 0:
        return None
    return {
        'period_start': summary['period_start'],
        'period_end': summary['period_end'],
        'net_overspend': summary['net_overspend'],
        'spent': summary['spent'],
        'discretionary': summary['discretionary'],
    }


def _daily_budget_serialize_debt(
    debt,
    *,
    balance: float,
    repaid_this_period: float,
) -> dict | None:
    if not isinstance(debt, dict):
        return None
    open_bal = _daily_budget_debt_open_balance(debt)
    if open_bal <= 0 and balance <= 0 and float(debt.get('repaid_total') or 0) <= 0:
        return None
    try:
        original = float(debt.get('original_amount') or open_bal or 0)
    except (TypeError, ValueError):
        original = open_bal
    try:
        repaid_total = float(debt.get('repaid_total') or 0)
    except (TypeError, ValueError):
        repaid_total = 0.0
    # Include in-period skim not yet folded into repaid_total.
    repaid_total_display = round(repaid_total + max(0.0, repaid_this_period), 2)
    return {
        'balance': round(max(0.0, balance), 2),
        'balance_at_period_open': open_bal,
        'original_amount': round(max(0.0, original), 2),
        'repaid_total': repaid_total_display,
        'repaid_this_period': round(max(0.0, repaid_this_period), 2),
        'source_period_start': debt.get('source_period_start'),
        'source_period_end': debt.get('source_period_end'),
        'created_at': debt.get('created_at'),
    }


def _daily_budget_day_insights(day_rows: list[dict]) -> dict:
    """Annotate day rows and summarise month pace for the Goals day chart."""
    if not day_rows:
        return {
            'days_elapsed': 0,
            'days_under': 0,
            'days_over': 0,
            'days_clear': 0,
            'under_streak': 0,
            'avg_spent': 0.0,
            'avg_limit': 0.0,
            'overspend_total': 0.0,
            'best_day': None,
            'worst_day': None,
        }

    days_under = 0
    days_over = 0
    days_clear = 0
    overspend_total = 0.0
    best_day = None
    worst_day = None

    for row in day_rows:
        lim = float(row.get('limit') or 0)
        spent = float(row.get('spent') or 0)
        remaining = round(lim - spent, 2)
        under = round(max(0.0, lim - spent), 2)
        over = round(max(0.0, spent - lim), 2)
        if spent <= 0.005:
            status = 'clear'
            days_clear += 1
            days_under += 1
        elif remaining < -0.005:
            status = 'over'
            days_over += 1
            overspend_total += over
            if worst_day is None or over > float(worst_day.get('overspend') or 0):
                worst_day = {
                    'date': row.get('date'),
                    'spent': spent,
                    'limit': lim,
                    'overspend': over,
                }
        elif remaining <= 0.005:
            status = 'exact'
            days_under += 1
        else:
            status = 'under'
            days_under += 1
            if under > 0 and (best_day is None or under > float(best_day.get('underspend') or 0)):
                best_day = {
                    'date': row.get('date'),
                    'spent': spent,
                    'limit': lim,
                    'underspend': under,
                }
        row['status'] = status
        row['underspend'] = under
        row['overspend'] = over
        row['remaining'] = remaining

    under_streak = 0
    for row in reversed(day_rows):
        if row.get('status') in ('under', 'clear', 'exact'):
            under_streak += 1
        else:
            break

    n = len(day_rows)
    avg_spent = round(sum(float(r.get('spent') or 0) for r in day_rows) / n, 2)
    avg_limit = round(sum(float(r.get('limit') or 0) for r in day_rows) / n, 2)
    return {
        'days_elapsed': n,
        'days_under': days_under,
        'days_over': days_over,
        'days_clear': days_clear,
        'under_streak': under_streak,
        'avg_spent': avg_spent,
        'avg_limit': avg_limit,
        'overspend_total': round(overspend_total, 2),
        'best_day': best_day,
        'worst_day': worst_day,
    }


def _daily_budget_common_titles(spending: dict, limit: int = 3) -> dict:
    """Most common payment-reference titles per category (up to ``limit`` each).

    Counts **manual** outgoing spends only (ignores statement imports), merges
    case/punctuation via ``_normalize_label``, and skips bare category labels
    used as form prefills.
    """
    # cat -> norm -> {count, displays Counter, last_date}
    tallies: dict[str, dict[str, dict]] = defaultdict(
        lambda: defaultdict(lambda: {'count': 0, 'displays': Counter(), 'last_date': ''})
    )
    entry_cats = set(DAILY_ENTRY_CATEGORIES)
    for t in spending.get('transactions') or []:
        if str(t.get('source') or '') != 'manual':
            continue
        if t.get('direction') != 'outgoing':
            continue
        if _spending_excluded_from_insight_metrics(t):
            continue
        cat = str(t.get('category') or '').strip().lower()
        if cat not in entry_cats:
            continue
        raw = str(t.get('description') or '').strip()[:200]
        if not raw:
            continue
        norm = _normalize_label(raw)
        if not norm:
            continue
        cat_norm = _normalize_label(cat.replace('_', ' '))
        if norm == cat_norm:
            continue
        bucket = tallies[cat][norm]
        bucket['count'] += 1
        bucket['displays'][raw] += 1
        d_str = str(t.get('date') or '')[:10]
        if d_str > (bucket['last_date'] or ''):
            bucket['last_date'] = d_str

    out: dict[str, list[str]] = {}
    for cat in DAILY_ENTRY_CATEGORIES:
        items = tallies.get(cat) or {}
        ranked = sorted(
            items.items(),
            key=lambda kv: (kv[1]['count'], kv[1]['last_date'] or '', kv[0]),
            reverse=True,
        )
        titles: list[str] = []
        for _norm, info in ranked[: max(0, int(limit))]:
            displays = info['displays']
            if not displays:
                continue
            titles.append(displays.most_common(1)[0][0])
        out[cat] = titles
    return out


def _daily_budget_status(spending: dict, as_of: date | None = None) -> dict:
    bucket, _ = _ensure_daily_budget(spending)
    plan = bucket.get('plan') or {}
    figures = _daily_budget_plan_figures(plan)
    today = as_of or _daily_budget_today()
    pay_day = _daily_budget_parse_pay_day(plan.get('pay_day'))
    period_start, period_end = _daily_budget_pay_period(today, pay_day)
    days_in_period = (period_end - period_start).days + 1
    pacing_start = _daily_budget_pacing_start(plan, period_start, period_end)
    spend_by_date = _daily_budget_spend_by_date(spending, period_start, period_end)
    goals = bucket.get('goals') or []
    goals_capacity = _daily_budget_goals_capacity(goals)
    priority = figures.get('underspend_priority') or 'debt_first'
    debt = bucket.get('overspend_debt') if isinstance(bucket.get('overspend_debt'), dict) else None
    debt_open = _daily_budget_debt_open_balance(debt)
    limits = _daily_budget_day_limits(
        figures,
        spend_by_date,
        period_start,
        period_end,
        pacing_start=pacing_start,
        debt_balance=debt_open,
        goals_capacity=goals_capacity,
        underspend_priority=priority,
    )
    today_key = today.isoformat()
    spent_today = float(spend_by_date.get(today_key, 0) or 0)
    daily_limit = float(limits.get(today_key, 0) or 0)
    remaining_today = round(daily_limit - spent_today, 2)
    discretionary = float(figures['discretionary_monthly'])
    pace_projection = _daily_budget_pace_projection(
        figures,
        spend_by_date,
        period_start,
        period_end,
        pacing_start,
        today,
        daily_limit,
        debt_balance=debt_open,
        goals_capacity=goals_capacity,
    )

    # Only count spend inside the pacing window (pre-tracking days are unknown /
    # intentionally excluded when tracking_from is mid-period).
    spent_mtd = 0.0
    d = pacing_start
    while d <= period_end:
        spent_mtd += float(spend_by_date.get(d.isoformat(), 0) or 0)
        d += timedelta(days=1)
    spent_mtd = round(spent_mtd, 2)

    days_in_window = (period_end - pacing_start).days + 1
    if pacing_start == period_start or not days_in_period:
        window_pool = discretionary
    else:
        window_pool = round(discretionary * days_in_window / days_in_period, 2)

    skim_through = min(today, period_end)
    skim = _daily_budget_skim_totals(
        limits,
        spend_by_date,
        pacing_start,
        skim_through,
        period_end,
        debt_open=debt_open,
        goals_capacity=goals_capacity,
        priority=priority,
    )
    underspend_total = float(skim['goals_pot'])
    repaid_this_period = float(skim['repaid'])
    debt_balance = float(skim['balance'])

    day_rows = []
    d = pacing_start
    while d <= min(today, period_end):
        key = d.isoformat()
        lim = float(limits.get(key, 0) or 0)
        spent = float(spend_by_date.get(key, 0) or 0)
        under = round(max(0.0, lim - spent), 2)
        day_rows.append({
            'date': key,
            'limit': lim,
            'spent': spent,
            'remaining': round(lim - spent, 2),
            'underspend': under,
            'weekday': d.weekday(),
            'is_today': key == today_key,
        })
        d += timedelta(days=1)
    day_insights = _daily_budget_day_insights(day_rows)

    txs_today = []
    for t in spending.get('transactions') or []:
        if str(t.get('date') or '')[:10] != today_key:
            continue
        if t.get('direction') != 'outgoing':
            continue
        if _spending_excluded_from_insight_metrics(t):
            continue
        txs_today.append(t)
    txs_today.sort(key=lambda r: (r.get('created_at') or '', r.get('description') or ''), reverse=True)

    tracking_from = _daily_budget_parse_tracking_from(plan.get('tracking_from'))
    overspend_prompt = _daily_budget_overspend_prompt(
        spending, plan, figures, period_start, today,
    )
    overspend_debt = _daily_budget_serialize_debt(
        debt,
        balance=debt_balance,
        repaid_this_period=repaid_this_period,
    )
    return {
        'as_of': today_key,
        'month': today.strftime('%Y-%m'),
        'pay_day': pay_day,
        'period_start': period_start.isoformat(),
        'period_end': period_end.isoformat(),
        'days_in_period': days_in_period,
        # Alias kept for older UI that divides discretionary by days_in_month.
        'days_in_month': days_in_period,
        'pacing_start': pacing_start.isoformat(),
        'tracking_from': tracking_from.isoformat() if tracking_from else None,
        'plan': figures,
        'source_month': plan.get('source_month'),
        'bill_items': plan.get('bill_items') or [],
        'daily_limit': round(daily_limit, 2),
        'spent_today': round(spent_today, 2),
        'remaining_today': remaining_today,
        'spent_mtd': spent_mtd,
        'discretionary_remaining_month': round(max(0.0, window_pool - spent_mtd), 2),
        # Unclamped period result for Goals hero (negative when net over discretionary).
        'period_net_saved': round(window_pool - spent_mtd, 2),
        'pace_projection': pace_projection,
        'underspend_saved': underspend_total,
        'underspend_priority': priority,
        'overspend_debt': overspend_debt,
        'overspend_prompt': overspend_prompt,
        'days': day_rows,
        'day_insights': day_insights,
        'transactions_today': txs_today,
        'goals': goals,
        'common_titles_by_category': _daily_budget_common_titles(spending, limit=3),
    }


def _manual_match_amounts_close(a: float, b: float) -> bool:
    """Slight amount differences allowed when reconciling manual vs statement."""
    try:
        aa = float(a)
        bb = float(b)
    except (TypeError, ValueError):
        return False
    if round(abs(aa - bb), 2) <= DAILY_BUDGET_MANUAL_MATCH_AMOUNT_TOL:
        return True
    # Penny-level rounding only — no broad percentage slack for manual reconciliation.
    if round(abs(aa - bb), 2) <= 0.02:
        return True
    m = max(abs(aa), abs(bb), 1e-9)
    return abs(aa - bb) / m <= 0.01


def _manual_match_dates_close(manual_date: str, statement_date: str) -> bool:
    """True when the statement posts on or up to N days after the manual entry date."""
    dm = _parse_iso_date(str(manual_date or '')[:10])
    ds = _parse_iso_date(str(statement_date or '')[:10])
    if dm is None or ds is None:
        return str(manual_date or '')[:10] == str(statement_date or '')[:10]
    delta = (ds - dm).days
    return 0 <= delta <= DAILY_BUDGET_MANUAL_MATCH_DATE_SLACK_DAYS


def _manual_description_match_ratio(man_n: str, desc_n: str) -> float:
    if not man_n and not desc_n:
        return 1.0
    if not man_n or not desc_n:
        return 0.55
    ratio = SequenceMatcher(None, man_n, desc_n).ratio()
    # Token overlap: manual "costa" inside bank "costa coffee cambridge ref 99"
    if man_n in desc_n or desc_n in man_n:
        ratio = max(ratio, 0.85)
    man_toks = set(man_n.split())
    desc_toks = set(desc_n.split())
    if man_toks and man_toks <= desc_toks:
        ratio = max(ratio, 0.88)
    return ratio


def _spending_included_bill_items(spending: dict) -> list:
    db = spending.get('daily_budget') if isinstance(spending.get('daily_budget'), dict) else {}
    plan = db.get('plan') if isinstance(db.get('plan'), dict) else {}
    raw = plan.get('bill_items') if isinstance(plan.get('bill_items'), list) else []
    out = []
    for b in raw:
        if not isinstance(b, dict):
            continue
        if b.get('included', True) is False:
            continue
        out.append(b)
    return out


def _match_expected_bill_item(
    description: str,
    amount: float,
    bill_items: list,
    *,
    used_idxs: set[int] | None = None,
) -> int | None:
    """Return index into bill_items when statement line looks like a monthly expected bill."""
    used = used_idxs if used_idxs is not None else set()
    desc_n = _normalize_label(description)
    try:
        amt = float(amount)
    except (TypeError, ValueError):
        return None
    best_j = None
    best_sim = -1.0
    for j, b in enumerate(bill_items):
        if j in used:
            continue
        label = str(b.get('label') or b.get('description') or '')
        try:
            b_amt = float(b.get('amount') or 0)
        except (TypeError, ValueError):
            continue
        if not _amounts_close_for_compare(amt, b_amt):
            continue
        sim = _manual_description_match_ratio(_normalize_label(label), desc_n)
        if sim > best_sim:
            best_sim = sim
            best_j = j
    if best_j is None or best_sim < SPENDING_EXPECTED_BILL_SIM_THRESHOLD:
        return None
    return best_j


def _daily_budget_fuzzy_match_manual(
    spending: dict,
    *,
    date_str: str,
    amount: float,
    description: str,
    direction: str,
    exclude_ids: set[str] | None = None,
) -> dict | None:
    """Find a manual ledger row that likely duplicates an incoming statement line.

    Matching prioritises date (statement on or up to a few days after manual) and
    amount (tight tolerance). When that pair uniquely identifies one unmatched manual
    entry, different payment references / titles are ignored. Description similarity
    is used to break ties when several manuals share a near date and amount.
    """
    amount = round(float(amount), 2)
    desc_n = _normalize_label(description)
    exclude = exclude_ids if exclude_ids is not None else set()
    candidates = []
    for t in spending.get('transactions') or []:
        tid = str(t.get('id') or '')
        if tid and tid in exclude:
            continue
        if str(t.get('source') or '') != 'manual':
            continue
        if t.get('bank_matched'):
            continue
        if str(t.get('direction') or '') != direction:
            continue
        if not _manual_match_dates_close(str(t.get('date') or '')[:10], date_str):
            continue
        try:
            man_amt = round(float(t.get('amount') or 0), 2)
        except (TypeError, ValueError):
            continue
        if not _manual_match_amounts_close(man_amt, amount):
            continue
        man_n = _normalize_label(str(t.get('description') or ''))
        ratio = _manual_description_match_ratio(man_n, desc_n)
        amount_delta = abs(man_amt - amount)
        candidates.append((ratio, amount_delta, t))
    if not candidates:
        return None
    # Unique on date+amount window → accept even when payment refs differ.
    if len(candidates) == 1:
        return candidates[0][2]
    candidates.sort(key=lambda x: (-x[0], x[1]))
    # Prefer description when several manuals collide on date/amount.
    labeled = [
        c for c in candidates
        if c[0] >= DAILY_BUDGET_MANUAL_MATCH_RATIO
        or (c[0] >= 0.55 and not _normalize_label(str(c[2].get('description') or '')))
    ]
    pool = labeled if labeled else []
    if not pool:
        # No usable labels: only auto-match if one candidate is clearly closer in amount.
        candidates.sort(key=lambda x: (x[1], -x[0]))
        if len(candidates) >= 2 and abs(candidates[0][1] - candidates[1][1]) < 0.01:
            return None
        return candidates[0][2]
    if len(pool) > 1 and abs(pool[0][0] - pool[1][0]) < 0.05:
        if pool[0][0] < 0.9:
            return None
    return pool[0][2]


def _daily_budget_claim_manual_match(manual_tx: dict, row: dict, statement_id: str, fingerprint: str) -> None:
    manual_tx['bank_matched'] = True
    manual_tx['source_statement_id'] = statement_id
    manual_tx['fingerprint'] = fingerprint
    bank_desc = str(row.get('description') or '').strip()
    if bank_desc:
        manual_tx['bank_description'] = bank_desc[:500]
        # Prefer richer bank label when manual title is empty/generic
        man = str(manual_tx.get('description') or '').strip()
        if not man or man.lower() in ('spend', 'expense', 'purchase'):
            manual_tx['description'] = bank_desc[:500]
    # Bank amount is ground truth when reconciling slight differences.
    try:
        bank_amt = round(float(row.get('amount') or 0), 2)
        man_amt = round(float(manual_tx.get('amount') or 0), 2)
        if bank_amt > 0 and abs(bank_amt - man_amt) > 0.001:
            manual_tx['manual_amount'] = man_amt
            manual_tx['amount'] = bank_amt
    except (TypeError, ValueError):
        pass
    bank_date = str(row.get('date') or '')[:10]
    man_date = str(manual_tx.get('date') or '')[:10]
    if bank_date and man_date and bank_date != man_date:
        manual_tx['manual_date'] = man_date
        manual_tx['date'] = bank_date
        manual_tx['month'] = bank_date[:7]
    if manual_tx.get('category') in (None, '', 'unclassified') and row.get('category'):
        cat = str(row.get('category')).strip().lower()
        if cat in SPENDING_CATEGORY_SET:
            manual_tx['category'] = cat


def _daily_budget_suggest_manual_matches(
    spending: dict,
    *,
    date_str: str,
    amount: float,
    description: str,
    direction: str,
    exclude_ids: set[str] | None = None,
    limit: int | None = None,
) -> list[dict]:
    """Near-miss manuals for preview UI — user can dismiss a statement row as already covered.

    Looser than auto-match: wider date/amount windows. Does not claim or mutate ledger rows.
    """
    amount = round(float(amount), 2)
    desc_n = _normalize_label(description)
    exclude = exclude_ids if exclude_ids is not None else set()
    max_n = DAILY_BUDGET_MANUAL_SUGGEST_LIMIT if limit is None else max(0, int(limit))
    stmt_d = _parse_iso_date(str(date_str or '')[:10])
    scored: list[tuple[float, float, int, dict]] = []
    for t in spending.get('transactions') or []:
        tid = str(t.get('id') or '')
        if tid and tid in exclude:
            continue
        if str(t.get('source') or '') != 'manual':
            continue
        if t.get('bank_matched'):
            continue
        if str(t.get('direction') or '') != direction:
            continue
        man_date = str(t.get('date') or '')[:10]
        man_d = _parse_iso_date(man_date)
        if stmt_d is None or man_d is None:
            if man_date != str(date_str or '')[:10]:
                continue
            date_delta = 0
        else:
            date_delta = (stmt_d - man_d).days
            if date_delta < 0 or date_delta > DAILY_BUDGET_MANUAL_SUGGEST_DATE_SLACK_DAYS:
                continue
        try:
            man_amt = round(float(t.get('amount') or 0), 2)
        except (TypeError, ValueError):
            continue
        amount_delta = round(abs(man_amt - amount), 2)
        if amount_delta > DAILY_BUDGET_MANUAL_SUGGEST_AMOUNT_TOL:
            m = max(abs(man_amt), abs(amount), 1e-9)
            if amount_delta / m > 0.15:
                continue
        man_n = _normalize_label(str(t.get('description') or ''))
        ratio = _manual_description_match_ratio(man_n, desc_n)
        # Prefer closer amounts, nearer dates, better labels.
        score = (amount_delta * 2.0) + (date_delta * 0.35) - (ratio * 3.0)
        scored.append((score, amount_delta, date_delta, {
            'id': tid,
            'date': man_date,
            'amount': man_amt,
            'description': str(t.get('description') or '')[:200],
            'direction': str(t.get('direction') or 'outgoing'),
            'category': str(t.get('category') or '') or None,
            'amount_delta': amount_delta,
            'date_delta_days': date_delta,
            'label_ratio': round(ratio, 3),
        }))
    scored.sort(key=lambda x: (x[0], x[1], x[2]))
    return [row for _, _, _, row in scored[:max_n]]


def _normalize_daily_bill_items(raw_items) -> list:
    out = []
    if not isinstance(raw_items, list):
        return out
    for row in raw_items:
        if not isinstance(row, dict):
            continue
        try:
            amount = round(max(0.0, float(row.get('amount') or 0)), 2)
        except (TypeError, ValueError):
            continue
        if amount <= 0:
            continue
        label = str(row.get('label') or row.get('description') or 'Bill').strip()[:120]
        cat = str(row.get('category') or 'other').strip().lower()
        if cat not in SPENDING_CATEGORY_SET:
            cat = 'other'
        out.append({
            'id': str(row.get('id') or uuid.uuid4()),
            'label': label,
            'amount': amount,
            'category': cat,
            'included': bool(row.get('included', True)),
            'source': str(row.get('source') or 'manual')[:40],
        })
    return out


def _build_hybrid_bill_estimate(spending: dict, month_key: str, *, use_llm: bool = True) -> dict:
    """Hybrid bills: category rollup + subscription signals + optional LLM regular-bill pass."""
    mk = _normalize_spending_month_key(month_key)
    if not mk:
        return {'error': 'Invalid month', 'bill_items': [], 'income_monthly': 0.0}

    insight = (spending.get('monthly_insights') or {}).get(mk) or {}
    try:
        income = float(insight.get('income_total') or 0)
    except (TypeError, ValueError):
        income = 0.0
    if income <= 0:
        # Fall back to summing incoming txs
        for t in spending.get('transactions') or []:
            if _report_month_for_spending_tx(t) != mk:
                continue
            if t.get('direction') != 'incoming':
                continue
            if _spending_excluded_from_insight_metrics(t):
                continue
            try:
                income += float(t.get('amount') or 0)
            except (TypeError, ValueError):
                continue
    income = round(income, 2)

    category_totals: dict[str, float] = defaultdict(float)
    merchant_totals: dict[str, dict] = {}
    for t in spending.get('transactions') or []:
        if _report_month_for_spending_tx(t) != mk:
            continue
        if t.get('direction') != 'outgoing':
            continue
        if _spending_excluded_from_insight_metrics(t):
            continue
        try:
            amt = float(t.get('amount') or 0)
        except (TypeError, ValueError):
            continue
        if amt <= 0.005:
            continue
        cat = str(t.get('category') or 'other').strip().lower()
        if cat not in SPENDING_CATEGORY_SET:
            cat = 'other'
        if cat in DAILY_BUDGET_BILL_CATEGORIES:
            category_totals[cat] += amt
        label = _normalize_label(str(t.get('description') or ''))
        if not label:
            continue
        slot = merchant_totals.setdefault(label, {
            'label': str(t.get('description') or '')[:120],
            'amount': 0.0,
            'charges': [],
            'category': cat,
            'norm': label,
        })
        slot['amount'] += amt
        parsed = _spending_tx_parsed_date(t)
        slot['charges'].append({
            'date': parsed,
            'amount': amt,
            'report_month': mk,
        })
        # Prefer bill-like category if any tx in the group has one
        if cat in DAILY_BUDGET_BILL_CATEGORIES:
            slot['category'] = cat

    bill_items = []
    seen_norms: set[str] = set()
    # Prefer merchant-level lines for bill categories (avoids double-counting with signals)
    for slot in sorted(merchant_totals.values(), key=lambda s: -float(s.get('amount') or 0)):
        cat = slot['category']
        if cat not in DAILY_BUDGET_BILL_CATEGORIES:
            continue
        # Avoid combining a delayed prior cycle with this month's charge into one total.
        amt = _representative_monthly_bill_amount(slot.get('charges') or [])
        if amt < 0.01:
            amt = round(float(slot['amount'] or 0), 2)
        if amt < 0.01:
            continue
        bill_items.append({
            'id': str(uuid.uuid4()),
            'label': slot['label'],
            'amount': amt,
            'category': cat,
            'included': True,
            'source': 'category',
        })
        seen_norms.add(slot['norm'])

    # If a bill category had spend but no usable merchant labels, keep a category bucket
    covered_cats = {b['category'] for b in bill_items}
    for cat, amt in sorted(category_totals.items(), key=lambda x: -x[1]):
        if cat in covered_cats or amt < 0.01:
            continue
        bill_items.append({
            'id': str(uuid.uuid4()),
            'label': f'{cat.replace("_", " ").title()} (category total)',
            'amount': round(amt, 2),
            'category': cat,
            'included': True,
            'source': 'category',
        })
    for sig in _subscription_signals_for_month(spending, mk):
        label = str(sig.get('display_description') or sig.get('label') or 'Subscription')[:120]
        try:
            amt = float(sig.get('total_last_month') or sig.get('last_amount') or 0)
        except (TypeError, ValueError):
            amt = 0.0
        if amt < 0.01:
            continue
        norm = _normalize_label(label)
        # Skip if already covered loosely by category rollup merchants later
        if norm in seen_norms:
            continue
        bill_items.append({
            'id': str(uuid.uuid4()),
            'label': label,
            'amount': round(amt, 2),
            'category': 'subscriptions',
            'included': True,
            'source': 'subscription_signal',
        })
        seen_norms.add(norm)

    llm_candidates = []
    for slot in merchant_totals.values():
        cat = slot['category']
        if cat in DAILY_BUDGET_BILL_CATEGORIES:
            continue
        if slot['amount'] < 15:
            continue
        llm_candidates.append(slot)

    llm_flagged = []
    if use_llm and llm_candidates:
        llm_flagged = _llm_flag_regular_bills(llm_candidates[:40])
        for item in llm_flagged:
            norm = _normalize_label(item.get('label', ''))
            if not norm or norm in seen_norms:
                continue
            bill_items.append({
                'id': str(uuid.uuid4()),
                'label': str(item.get('label') or '')[:120],
                'amount': round(float(item.get('amount') or 0), 2),
                'category': str(item.get('category') or 'other'),
                'included': True,
                'source': 'llm_bill',
                'rationale': str(item.get('rationale') or '')[:240],
            })
            seen_norms.add(norm)

    bills_total = round(sum(float(b['amount']) for b in bill_items if b.get('included', True)), 2)
    return {
        'month': mk,
        'income_monthly': income,
        'bill_items': bill_items,
        'bills_monthly': bills_total,
        'llm_used': bool(llm_flagged),
        'llm_flagged_count': len(llm_flagged),
    }


def _llm_flag_regular_bills(candidates: list[dict]) -> list[dict]:
    """Ask the model which merchant lines look like regular monthly bills (e.g. council tax)."""
    client = _get_openai_client()
    if not client or not candidates:
        return []
    model = os.getenv('OPENAI_MODEL', 'gpt-4o-mini').strip() or 'gpt-4o-mini'
    lines = []
    for c in candidates:
        lines.append(
            f'- {c.get("label")}|amount={round(float(c.get("amount") or 0), 2)}|category={c.get("category")}'
        )
    system = (
        'You identify regular monthly household bills from bank spending lines. '
        'Regular bills include rent/mortgage, council tax, insurance, utilities, phone/broadband, '
        'TV licence, childcare, loan repayments, and similar recurring obligations. '
        'Do NOT flag groceries, dining, shopping, one-off purchases, or variable everyday spend. '
        'Return JSON only: {"items":[{"description":"exact label from input","is_regular_bill":true,'
        '"category":"one of housing|utilities|subscriptions|debt|other","reason":"short"}]}'
    )
    try:
        completion = client.chat.completions.create(
            model=model,
            messages=[
                {'role': 'system', 'content': system},
                {'role': 'user', 'content': 'Lines:\n' + '\n'.join(lines)},
            ],
            response_format={'type': 'json_object'},
            temperature=0.1,
        )
        raw = completion.choices[0].message.content or '{}'
        parsed, jerr = _parse_llm_json_object(raw, context='daily_budget_bills')
        if jerr or not isinstance(parsed, dict):
            return []
    except Exception as e:
        logger.warning('LLM regular-bill detection failed: %s', e)
        return []

    by_norm = {_normalize_label(c.get('label', '')): c for c in candidates}
    out = []
    for item in parsed.get('items') or []:
        if not isinstance(item, dict) or not item.get('is_regular_bill'):
            continue
        desc = str(item.get('description') or '').strip()
        norm = _normalize_label(desc)
        src = by_norm.get(norm)
        if not src:
            # fuzzy: match substring
            for n, c in by_norm.items():
                if norm and (norm in n or n in norm):
                    src = c
                    break
        if not src:
            continue
        cat = str(item.get('category') or 'other').strip().lower()
        if cat not in SPENDING_CATEGORY_SET:
            cat = 'other'
        if cat not in DAILY_BUDGET_BILL_CATEGORIES and cat != 'other':
            cat = 'other'
        out.append({
            'label': src['label'],
            'amount': round(float(src.get('amount') or 0), 2),
            'category': cat if cat in DAILY_BUDGET_BILL_CATEGORIES else 'other',
            'rationale': str(item.get('reason') or '').strip()[:240],
        })
    return out


def _serialize_daily_budget_plan(plan: dict) -> dict:
    figures = _daily_budget_plan_figures(plan)
    tf = _daily_budget_parse_tracking_from(plan.get('tracking_from'))
    return {
        **figures,
        'pay_day': _daily_budget_parse_pay_day(plan.get('pay_day')),
        'tracking_from': tf.isoformat() if tf else None,
        'source_month': plan.get('source_month'),
        'bill_items': plan.get('bill_items') or [],
        'updated_at': plan.get('updated_at'),
    }


def _daily_budget_accept_overspend(
    bucket: dict,
    *,
    period_start: date,
    period_end: date,
    net_overspend: float,
    current_period_start: date,
    current_display_balance: float,
) -> dict:
    """Record accept decision and add net overspend to the debt balance."""
    net = round(max(0.0, float(net_overspend)), 2)
    key = _daily_budget_period_key(period_start, period_end)
    decisions = bucket.setdefault('overspend_decisions', {})
    if not isinstance(decisions, dict):
        decisions = {}
        bucket['overspend_decisions'] = decisions
    now_iso = datetime.utcnow().isoformat() + 'Z'
    decisions[key] = {
        'decision': 'accepted',
        'net_overspend': net,
        'decided_at': now_iso,
    }
    debt = bucket.get('overspend_debt') if isinstance(bucket.get('overspend_debt'), dict) else None
    base_bal = max(0.0, round(float(current_display_balance or 0), 2))
    new_open = round(base_bal + net, 2)
    if debt is None or (
        _daily_budget_debt_open_balance(debt) <= 0
        and base_bal <= 0
        and float(debt.get('repaid_total') or 0) <= 0
    ):
        debt = {
            'balance_at_period_open': new_open,
            'balance': new_open,
            'original_amount': net,
            'repaid_total': 0.0,
            'source_period_start': period_start.isoformat(),
            'source_period_end': period_end.isoformat(),
            'created_at': now_iso,
            'active_period_start': current_period_start.isoformat(),
        }
    else:
        debt['balance_at_period_open'] = new_open
        debt['balance'] = new_open
        debt['original_amount'] = round(float(debt.get('original_amount') or 0) + net, 2)
        debt['last_source_period_start'] = period_start.isoformat()
        debt['last_source_period_end'] = period_end.isoformat()
        debt['active_period_start'] = current_period_start.isoformat()
    bucket['overspend_debt'] = debt
    return debt


def _daily_budget_decline_overspend(
    bucket: dict,
    *,
    period_start: date,
    period_end: date,
    net_overspend: float,
) -> None:
    key = _daily_budget_period_key(period_start, period_end)
    decisions = bucket.setdefault('overspend_decisions', {})
    if not isinstance(decisions, dict):
        decisions = {}
        bucket['overspend_decisions'] = decisions
    decisions[key] = {
        'decision': 'declined',
        'net_overspend': round(max(0.0, float(net_overspend)), 2),
        'decided_at': datetime.utcnow().isoformat() + 'Z',
    }


def _daily_budget_write_off_debt(bucket: dict) -> dict | None:
    debt = bucket.get('overspend_debt') if isinstance(bucket.get('overspend_debt'), dict) else None
    if debt is None:
        return None
    now_iso = datetime.utcnow().isoformat() + 'Z'
    try:
        wiped = max(0.0, float(debt.get('balance') or debt.get('balance_at_period_open') or 0))
    except (TypeError, ValueError):
        wiped = 0.0
    debt['balance_at_period_open'] = 0.0
    debt['balance'] = 0.0
    debt['written_off_at'] = now_iso
    debt['written_off_amount'] = wiped
    # Keep a trail but clear active debt for status serialization.
    bucket['overspend_debt'] = None
    history = bucket.get('overspend_write_offs')
    if not isinstance(history, list):
        history = []
        bucket['overspend_write_offs'] = history
    history.append({
        'written_off_at': now_iso,
        'amount': wiped,
        'original_amount': debt.get('original_amount'),
        'source_period_start': debt.get('source_period_start'),
        'source_period_end': debt.get('source_period_end'),
    })
    return debt


# --- LLM "savings coach" --------------------------------------------------------------------
SAVINGS_ADVICE_TREND_MAX_PRIOR = 12
SAVINGS_ADVICE_LLM_MAX_CONTEXT_CHARS = 200_000
# Line-level context (capped: third-party still receives real descriptions/amounts).
SAVINGS_ADVICE_FOCAL_TX_MAX = 300
SAVINGS_ADVICE_DESC_MAX = 200
SAVINGS_ADVICE_PRIOR_MONTHS_TX_DETAIL = 2
SAVINGS_ADVICE_PRIOR_TX_PER_MONTH = 60


def _savings_trend_row_compact(ins: dict, month_key: str) -> dict:
    top_raw = ins.get('top_categories') or []
    top3 = []
    for t in top_raw[:3]:
        if not isinstance(t, dict):
            continue
        try:
            amt = round(float(t.get('amount', 0) or 0), 2)
        except (TypeError, ValueError):
            amt = 0.0
        c = str(t.get('category', '') or '').strip()
        if c:
            top3.append({'category': c, 'amount': amt})
    try:
        inc = float(ins.get('income_total', 0) or 0)
    except (TypeError, ValueError):
        inc = 0.0
    try:
        outg = float(ins.get('outgoing_total', 0) or 0)
    except (TypeError, ValueError):
        outg = 0.0
    try:
        net = float(ins.get('net', 0) or 0)
    except (TypeError, ValueError):
        net = 0.0
    sr = ins.get('savings_rate')
    if sr is not None:
        try:
            sr = round(float(sr), 2)
        except (TypeError, ValueError):
            sr = None
    return {
        'month': month_key,
        'income_total': round(inc, 2),
        'outgoing_total': round(outg, 2),
        'net': round(net, 2),
        'savings_rate': sr,
        'top_categories': top3,
    }


def _savings_advice_serialize_tx(t: dict) -> dict:
    desc = str(t.get('description') or '').strip()
    if len(desc) > SAVINGS_ADVICE_DESC_MAX:
        desc = desc[:SAVINGS_ADVICE_DESC_MAX - 1] + '…'
    try:
        amt = round(float(t.get('amount', 0) or 0), 2)
    except (TypeError, ValueError):
        amt = 0.0
    in_kpi = not _spending_excluded_from_insight_metrics(t)
    return {
        'date': str(t.get('date') or '')[:10],
        'description': desc,
        'direction': str(t.get('direction') or ''),
        'amount': amt,
        'category': (str(t.get('category') or '').strip() or None),
        'counts_in_month_totals': in_kpi,
        'internal_transfer_leg': _spending_is_paired_leg(t),
    }


def _collect_savings_advice_tx_for_month(
    spending: dict, month_key: str, max_count: int
) -> tuple[list[dict], dict]:
    """
    Serialize transactions for one report month, newest-first by amount when over max_count
    (keeps the largest lines so limits still allow deeper insights). Sorted by date, then
    description in the final list.
    """
    txs = [t for t in (spending.get('transactions') or []) if _report_month_for_spending_tx(t) == month_key]
    n = len(txs)
    meta: dict = {'count_in_month': n, 'count_sent': 0, 'truncated': False}
    if n == 0:
        return [], meta
    if n <= max_count:
        rows = sorted(txs, key=lambda x: (str(x.get('date') or ''), str(x.get('description') or '')))
        ser = [_savings_advice_serialize_tx(t) for t in rows]
        meta['count_sent'] = n
        return ser, meta
    def _abs_amt(t: dict) -> float:
        try:
            return abs(float(t.get('amount', 0) or 0))
        except (TypeError, ValueError):
            return 0.0
    ranked = sorted(txs, key=_abs_amt, reverse=True)
    kept = ranked[:max_count]
    kept.sort(key=lambda x: (str(x.get('date') or ''), str(x.get('description') or '')))
    ser = [_savings_advice_serialize_tx(t) for t in kept]
    meta['count_sent'] = max_count
    meta['truncated'] = True
    return ser, meta


def _build_savings_advice_context(
    spending: dict, focal_month: str
) -> tuple[dict | None, int, str | None, dict | None]:
    """
    Focal = full stored insight + per-line transactions (capped). Prior months = trend_series
    and optional per-line sample for the two most recent months before focal. Transaction lists
    are sorted by date for the model unless truncation applied (largest |amount| lines kept).
    """
    insight_map = spending.get('monthly_insights') or {}
    focal = insight_map.get(focal_month)
    if not isinstance(focal, dict) or not focal:
        return None, 0, 'no_insight', None
    all_keys = [str(k) for k in insight_map.keys() if k]
    all_sorted_desc = sorted(all_keys, reverse=True)
    priors_newest_first = [m for m in all_sorted_desc if m < focal_month][:SAVINGS_ADVICE_TREND_MAX_PRIOR]
    trend_months_oldest_first = list(reversed(priors_newest_first))
    trend_series: list[dict] = []
    for m in trend_months_oldest_first:
        ins = insight_map.get(m)
        if not isinstance(ins, dict):
            continue
        trend_series.append(_savings_trend_row_compact(ins, m))
    focal_tx, focal_meta = _collect_savings_advice_tx_for_month(
        spending, focal_month, SAVINGS_ADVICE_FOCAL_TX_MAX
    )
    prior_tx_blocks: list[dict] = []
    for pm in priors_newest_first[:SAVINGS_ADVICE_PRIOR_MONTHS_TX_DETAIL]:
        p_rows, p_meta = _collect_savings_advice_tx_for_month(
            spending, pm, SAVINGS_ADVICE_PRIOR_TX_PER_MONTH
        )
        if not p_rows and p_meta.get('count_in_month', 0) == 0:
            continue
        prior_tx_blocks.append({
            'month': pm,
            'meta': p_meta,
            'transactions': p_rows,
        })
    meta_out = {
        'focal_transactions': dict(focal_meta),
        'prior_transaction_months_included': len(prior_tx_blocks),
    }
    context = {
        'focal_month': focal_month,
        'insight_focal': copy.deepcopy(focal),
        'focal_month_transactions': focal_tx,
        'focal_transaction_meta': focal_meta,
        'trend_series': trend_series,
        'trend_months_included': len(trend_series),
        'prior_month_transaction_samples': prior_tx_blocks,
    }
    return context, len(trend_series), None, meta_out


def _validate_savings_advice_payload(d: dict) -> dict | None:
    if not isinstance(d, dict):
        return None
    summary = d.get('summary')
    if not isinstance(summary, str) or not summary.strip():
        return None
    recs = d.get('recommendations')
    if not isinstance(recs, list):
        return None
    out: list[dict] = []
    for r in recs[:8]:
        if not isinstance(r, dict):
            continue
        title = str(r.get('title') or '').strip()
        detail = str(r.get('detail') or '').strip()
        ev = str(r.get('evidence') or '').strip()
        if not title:
            continue
        try:
            pr = int(r.get('priority', 2))
        except (TypeError, ValueError):
            pr = 2
        pr = max(1, min(3, pr))
        out.append({
            'title': title[:300],
            'detail': detail[:2000],
            'priority': pr,
            'evidence': ev[:500],
        })
    if not out:
        return None
    return {
        'summary': summary.strip()[:2000],
        'recommendations': out,
    }


def _generate_savings_advice_with_llm(context: dict) -> tuple[dict | None, str | None, str | None]:
    client = _get_openai_client()
    if not client:
        return None, 'Savings ideas require OPENAI_API_KEY to be set on the server.', None
    model = os.getenv('OPENAI_MODEL', 'gpt-4o-mini').strip() or 'gpt-4o-mini'
    system = (
        'You are a savings coach. The user message is JSON with:\n'
        '- "insight_focal": full monthly spending insight for the selected month.\n'
        '- "focal_month_transactions": individual lines for the focal month (date, description, amount, direction, category; '
        'flags for internal-transfer legs and whether the line counts in month KPI totals). Use these for concrete, line-level insights. '
        'If focal_transaction_meta.truncated is true, the list is not exhaustive—larger lines are prioritised. '
        '- "prior_month_transaction_samples": optional per-line samples for the most recent prior month(s) (capped), same shape.\n'
        '- "trend_series": up to 12 months before the focal month, each with income/outgoing/net, savings_rate, and top 3 category amounts. Oldest to newest.\n'
        'Rules: Only use numbers and facts that appear in the JSON. Do not invent amounts or payees. '
        'For deeper ideas, connect insight_focal and focal_month_transactions (e.g. specific merchants, categories, or large lines). '
        'Prefer also using trend_series and prior month samples when non-empty. '
        'If history is thin, say so briefly—do not fabricate long trends. '
        'Explain internal_transfer_leg and counts_in_month_totals when relevant. '
        'Phrasing: general information only, not regulated financial, tax, or legal advice. No guaranteed returns.\n'
        'Return only valid JSON with: '
        '{"summary": string (1-2 sentences), "recommendations": ['
        '{"title": string, "detail": string, "priority": integer 1-3, "evidence": string} '
        ']} . priority 1 = highest. evidence should cite the data: category, budget_action_items, line descriptions/dates, or trend_series. '
        'Between 1 and 8 recommendations. British English; use £ where relevant.'
    )
    user_body = json.dumps(context, ensure_ascii=False, indent=2)
    if len(user_body) > SAVINGS_ADVICE_LLM_MAX_CONTEXT_CHARS:
        user_body = user_body[:SAVINGS_ADVICE_LLM_MAX_CONTEXT_CHARS] + '\n... [truncated for length]'
    err_note: str | None = None
    for attempt in range(2):
        try:
            extra = ' Reply with the JSON object only, no other text.' if attempt else ''
            completion = client.chat.completions.create(
                model=model,
                messages=[
                    {'role': 'system', 'content': system},
                    {'role': 'user', 'content': user_body + extra},
                ],
                response_format={'type': 'json_object'},
                temperature=0.15 if attempt == 0 else 0.0,
            )
            raw = completion.choices[0].message.content or '{}'
            parsed, jerr = _parse_llm_json_object(raw, context='savings_advice')
            if jerr:
                err_note = 'Could not parse model response.'
                continue
            val = _validate_savings_advice_payload(parsed)
            if val:
                return val, None, model
            err_note = 'Model response was incomplete.'
        except Exception as e:
            logger.exception('savings advice LLM call failed: %s', e)
            err_note = str(e) or 'LLM call failed'
    return None, err_note or 'Could not generate savings ideas. Try again.', model


def should_apply_interest(target_day):
    today = date.today()
    if today.day == target_day:
        return True
    last_day_of_month = monthrange(today.year, today.month)[1]
    return today.day == last_day_of_month


def resolve_interest_as_of_date(year, month, interest_day):
    """Return the interest date for a month, clamping interest_day to the month length."""
    year = int(year)
    month = int(month)
    interest_day = int(interest_day)
    if not (1 <= month <= 12):
        raise ValueError(f'Invalid month: {month}')
    if interest_day < 1:
        raise ValueError(f'Invalid interest day: {interest_day}')
    last_day = monthrange(year, month)[1]
    return date(year, month, min(interest_day, last_day))


def loan_balance_as_of_end_of_day(transactions, as_of_date):
    """Sum transaction amounts with date <= as_of (end-of-day balance before a new interest row)."""
    if isinstance(as_of_date, date):
        as_of_str = as_of_date.strftime('%Y-%m-%d')
    else:
        as_of_str = str(as_of_date)
    total = 0.0
    for tx in transactions or []:
        tx_date = str(tx.get('date') or '')
        if tx_date and tx_date <= as_of_str:
            total += float(tx.get('amount') or 0)
    return total


def interest_applied_in_month(transactions, year, month):
    """True if any interest transaction falls in the given calendar year/month."""
    prefix = f'{int(year):04d}-{int(month):02d}'
    for tx in transactions or []:
        if tx.get('type') == 'interest' and str(tx.get('date') or '').startswith(prefix):
            return True
    return False


def compute_monthly_interest_amount(balance, rate):
    """Simple APR/12 monthly interest, rounded to 2 decimal places."""
    return round((float(balance) * (float(rate) / 100)) / 12, 2)


def apply_monthly_interest(loan_id):
    data = load_data()
    if loan_id not in data['loans']:
        return
    
    loan = data['loans'][loan_id]
    target_day = loan['interest_day']
    
    if not should_apply_interest(target_day):
        return  # Skip if not the correct day

    today = date.today()
    if interest_applied_in_month(loan.get('transactions'), today.year, today.month):
        return  # Already applied for this calendar month
    
    current_amount = loan['loan_amount']
    rate = loan['interest_rate']
    
    monthly_interest = compute_monthly_interest_amount(current_amount, rate)
    
    new_transaction = {
        'date': today.strftime('%Y-%m-%d'),
        'type': 'interest',
        'amount': monthly_interest,
        'description': f'Monthly interest at {rate}% APR (automated)',
        'user': 'system'
    }
    
    loan['transactions'].append(new_transaction)
    loan['loan_amount'] += monthly_interest
    save_data(data)

def schedule_interest_task(loan_id, day_of_month):
    # Remove existing interest jobs for this loan
    job_id = f'monthly_interest_{loan_id}'
    for job in scheduler.get_jobs():
        if job.id == job_id:
            scheduler.remove_job(job_id)
    
    # Schedule new interest job
    if 1 <= day_of_month <= 31:
        if day_of_month >= 28:
            scheduler.add_job(
                lambda: apply_monthly_interest(loan_id),
                CronTrigger(day="28-31"),
                id=job_id,
                replace_existing=True
            )
        else:
            scheduler.add_job(
                lambda: apply_monthly_interest(loan_id),
                CronTrigger(day=str(day_of_month)),
                id=job_id,
                replace_existing=True
            )
    else:
        raise ValueError(f"Invalid day of month: {day_of_month}")

def calculate_loan_stats(transactions):
    total_loan_quantity = 0
    total_paid = 0
    
    for transaction in transactions:
        if transaction['type'] in ['initial', 'addition']:
            total_loan_quantity += transaction['amount']
        elif transaction['type'] == 'repayment':
            total_paid += abs(transaction['amount'])
    
    return total_loan_quantity, total_paid


def _openai_env_timeout(name: str, default: float) -> float:
    raw = os.getenv(name, '').strip()
    try:
        val = float(raw) if raw else default
    except ValueError:
        val = default
    return val if val > 0 else default


def _openai_timeout_seconds() -> float:
    """Default per-request OpenAI timeout for the lighter calls (classification,
    layout hints, advice). A hung upstream connection surfaces as a handled
    error instead of the worker being force-killed (WORKER TIMEOUT / SystemExit).
    """
    return _openai_env_timeout('OPENAI_TIMEOUT', 60.0)


def _openai_extraction_timeout_seconds() -> float:
    """Per-chunk timeout for statement extraction. Keep below gunicorn --timeout
    (production uses 120s) so a hung chunk fails cleanly instead of killing the worker.
    Long statements are split into chunks; streaming preview yields between chunks.
    """
    return _openai_env_timeout('OPENAI_EXTRACTION_TIMEOUT', 90.0)


def _openai_max_retries() -> int:
    raw = os.getenv('OPENAI_MAX_RETRIES', '').strip()
    try:
        val = int(raw) if raw else 2
    except ValueError:
        val = 2
    return max(0, val)


def _get_openai_client():
    if OpenAI is None:
        return None
    api_key = os.getenv('OPENAI_API_KEY', '').strip()
    if not api_key:
        return None
    base_url = os.getenv('OPENAI_BASE_URL', '').strip() or None
    kwargs = {
        'api_key': api_key,
        'timeout': _openai_timeout_seconds(),
        'max_retries': _openai_max_retries(),
    }
    if base_url:
        kwargs['base_url'] = base_url
    return OpenAI(**kwargs)


def _is_spreadsheet_filename(name: str) -> bool:
    """True for Excel workbook extensions we can convert (or reject with a clear error)."""
    n = (name or '').lower()
    return n.endswith('.xlsx') or n.endswith('.xlsm') or n.endswith('.xls')


def _csv_escape_cell(value: str) -> str:
    if any(ch in value for ch in ',"\n\r'):
        return '"' + value.replace('"', '""') + '"'
    return value


def _spreadsheet_cell_to_str(value) -> str:
    if value is None:
        return ''
    if isinstance(value, datetime):
        if value.hour or value.minute or value.second or value.microsecond:
            return value.strftime('%Y-%m-%d %H:%M:%S')
        return value.strftime('%Y-%m-%d')
    if isinstance(value, date):
        return value.isoformat()
    if isinstance(value, bool):
        return 'TRUE' if value else 'FALSE'
    if isinstance(value, float):
        # Avoid 4.800000000000001 noise for money-like values.
        if value == int(value) and abs(value) < 1e15:
            return str(int(value))
        return f'{value:.10g}'
    return str(value)


def _extract_spreadsheet_text(raw: bytes, name: str = '') -> str:
    """
    Convert Excel (.xlsx / .xlsm) workbook bytes to CSV-like text for the statement LLM pipeline.
    Old binary .xls is rejected with a clear message (export as .xlsx or CSV instead).
    """
    n = (name or '').lower()
    if n.endswith('.xls') and not (n.endswith('.xlsx') or n.endswith('.xlsm')):
        raise ValueError(
            'Old .xls Excel format is not supported. Re-export as .xlsx or CSV from your bank (e.g. Revolut).'
        )
    try:
        from openpyxl import load_workbook
    except ImportError as e:  # pragma: no cover
        raise RuntimeError('Excel (.xlsx) support requires the openpyxl package.') from e

    wb = load_workbook(io.BytesIO(raw), read_only=True, data_only=True)
    try:
        sheet_blocks: list[str] = []
        for sheet in wb.worksheets:
            rows_out: list[str] = []
            for row in sheet.iter_rows(values_only=True):
                if row is None:
                    continue
                cells = [_spreadsheet_cell_to_str(c) for c in row]
                if not any(c.strip() for c in cells):
                    continue
                # Drop trailing empty cells so sparse Excel rows stay compact.
                while cells and not cells[-1].strip():
                    cells.pop()
                rows_out.append(','.join(_csv_escape_cell(c) for c in cells))
            if not rows_out:
                continue
            block = '\n'.join(rows_out)
            if len(wb.worksheets) > 1:
                sheet_blocks.append(f'--- Sheet: {sheet.title} ---\n{block}')
            else:
                sheet_blocks.append(block)
        text = '\n\n'.join(sheet_blocks).strip()
        if not text:
            raise ValueError('Excel file has no readable rows.')
        return text
    finally:
        wb.close()


def _extract_pdf_text(raw: bytes, *, meta_out: dict | None = None) -> str:
    """
    Bank PDFs are often table-heavy; pypdf alone loses column order.
    Prefer pdfplumber (layout-preserving) when available, then fall back to pypdf.
    Use whichever yields more usable characters.
    If meta_out is a dict, it is filled with keys: engine (str), engines (list of {name, non_ws_chars}).
    """
    def _non_ws_len(t: str) -> int:
        return len(re.sub(r'\s+', '', t or ''))

    candidates: list[tuple[str, str]] = []
    try:
        import pdfplumber  # type: ignore[import-untyped]

        parts = []
        with pdfplumber.open(io.BytesIO(raw)) as pdf:
            for page in pdf.pages:
                t = page.extract_text(layout=True, x_tolerance=2, y_tolerance=2)
                if t and t.strip():
                    parts.append(t)
        joined = '\n\n'.join(parts)
        if _non_ws_len(joined) >= 20:
            candidates.append(('pdfplumber', joined))
    except Exception as e:
        logger.debug('pdfplumber extraction skipped: %s', e)

    try:
        reader = PdfReader(io.BytesIO(raw))
        parts = []
        for page in reader.pages:
            t = page.extract_text()
            if t:
                parts.append(t)
        joined = '\n'.join(parts)
        if _non_ws_len(joined) >= 20:
            candidates.append(('pypdf', joined))
    except Exception as e:
        logger.warning('pypdf extraction failed: %s', e)

    if not candidates:
        raise RuntimeError('Could not extract readable text from this PDF. Try a CSV export from your bank.')
    candidates.sort(key=lambda x: _non_ws_len(x[1]), reverse=True)
    best = candidates[0][1]
    if meta_out is not None:
        meta_out['engine'] = candidates[0][0]
        meta_out['engines'] = [{'name': c[0], 'non_ws_chars': _non_ws_len(c[1])} for c in candidates]
    if len(candidates) > 1:
        logger.info(
            'PDF text: using %s (%s chars non-ws); alt had %s',
            candidates[0][0],
            _non_ws_len(best),
            [(c[0], _non_ws_len(c[1])) for c in candidates[1:]],
        )
    return best


def _group_pdf_words_into_lines(words: list, y_tol: float = 3.0) -> list[list]:
    if not words:
        return []
    sorted_w = sorted(words, key=lambda w: (w.get('top', 0), w.get('x0', 0)))
    lines: list[list] = []
    for w in sorted_w:
        t = w.get('top', 0)
        placed = False
        for line in lines:
            if abs(line[0].get('top', 0) - t) <= y_tol:
                line.append(w)
                placed = True
                break
        if not placed:
            lines.append([w])
    for line in lines:
        line.sort(key=lambda w: w.get('x0', 0))
    lines.sort(key=lambda line: line[0].get('top', 0))
    return lines


def _word_as_money_amount(text: str) -> float | None:
    t = (text or '').strip()
    if not t:
        return None
    t = re.sub(r'^[$£€]\s*', '', t)
    t = t.replace(',', '')
    if not re.match(r'^\d+\.\d{2}$', t):
        return None
    try:
        v = float(t)
    except ValueError:
        return None
    return v if v > 0 else None


def _word_center_x(w: dict) -> float:
    return (w.get('x0', 0) + w.get('x1', 0)) / 2


def _nearest_column(x: float, inc_x: float | None, out_x: float | None) -> str | None:
    if inc_x is None or out_x is None:
        return None
    mid = (inc_x + out_x) / 2
    d_in = abs(x - inc_x)
    d_out = abs(x - out_x)
    if d_in < d_out:
        return 'incoming'
    if d_out < d_in:
        return 'outgoing'
    return 'incoming' if x < mid else 'outgoing'


_SPENDING_MONEY_RE = re.compile(r'(?<![\d,.])(\d{1,3}(?:,\d{3})+|\d+)\.(\d{2})(?![\d])')


def _spending_direction_hints_from_layout_text(raw: bytes) -> list[dict]:
    """
    Robust fallback for PDFs where pdfplumber.extract_words splits each character
    into its own token (common with custom-spacing fonts in bank exports).

    We read page.extract_text(layout=True) and use character-column positions
    as a proxy for horizontal x coordinates. Any page whose header row exposes
    an incoming AND outgoing money column contributes hints.
    """
    try:
        import pdfplumber  # type: ignore[import-untyped]
    except Exception:
        return []

    in_phrases = ('paid in', 'money in', 'amount in', 'amt in', 'cash in')
    out_phrases = ('paid out', 'money out', 'amount out', 'amt out', 'cash out')
    in_singles = ('credits', 'deposits', 'credit', 'deposit')
    out_singles = ('debits', 'withdrawals', 'debit', 'withdrawal')

    def find_header_cols(line: str) -> tuple[int | None, int | None, int | None]:
        low = line.lower()
        inc_c = out_c = bal_c = None
        for phrase in in_phrases:
            i = low.find(phrase)
            if i != -1:
                inc_c = i + len(phrase) // 2
                break
        for phrase in out_phrases:
            i = low.find(phrase)
            if i != -1:
                out_c = i + len(phrase) // 2
                break
        if inc_c is None:
            for phrase in in_singles:
                i = low.find(phrase)
                if i != -1:
                    inc_c = i + len(phrase) // 2
                    break
        if out_c is None:
            for phrase in out_singles:
                i = low.find(phrase)
                if i != -1:
                    out_c = i + len(phrase) // 2
                    break
        i = low.find('balance')
        if i != -1:
            bal_c = i + len('balance') // 2
        return inc_c, out_c, bal_c

    hints: list[dict] = []
    try:
        with pdfplumber.open(io.BytesIO(raw)) as pdf:
            for page in pdf.pages:
                text = page.extract_text(layout=True, x_tolerance=2, y_tolerance=2) or ''
                if not text.strip():
                    continue
                lines = text.splitlines()
                inc_c = out_c = bal_c = None
                header_idx = -1
                for idx, line in enumerate(lines[: min(40, len(lines))]):
                    ic, oc, bc = find_header_cols(line)
                    if ic is not None and oc is not None:
                        inc_c, out_c, bal_c = ic, oc, bc
                        header_idx = idx
                        break
                if inc_c is None or out_c is None or inc_c == out_c:
                    continue

                for line in lines[header_idx + 1:]:
                    if not line.strip():
                        continue
                    low = line.lower()
                    # skip repeated header rows
                    if any(p in low for p in in_phrases) and any(p in low for p in out_phrases):
                        continue

                    matches = list(_SPENDING_MONEY_RE.finditer(line))
                    if not matches:
                        continue
                    clean_line = line.strip()[:200]
                    tokens = _hint_tokens_from_line(clean_line)
                    for m in matches:
                        center = (m.start() + m.end()) / 2
                        if bal_c is not None and abs(center - bal_c) <= 3:
                            continue
                        try:
                            amt = float(m.group(0).replace(',', ''))
                        except ValueError:
                            continue
                        if amt <= 0:
                            continue
                        d_inc = abs(center - inc_c)
                        d_out = abs(center - out_c)
                        if d_inc == d_out:
                            continue
                        direction = 'incoming' if d_inc < d_out else 'outgoing'
                        hints.append({
                            'amount': round(amt, 2),
                            'direction': direction,
                            'tokens': tokens,
                            'line_text': clean_line,
                        })
    except Exception as e:
        logger.debug('layout-text hints failed: %s', e)
        return []

    return hints


def _merge_spending_hints(*groups: list[dict]) -> list[dict]:
    seen: set[tuple[float, str, str]] = set()
    merged: list[dict] = []
    for group in groups:
        for h in group or []:
            key = (
                round(float(h.get('amount', 0)), 2),
                str(h.get('direction') or ''),
                str(h.get('line_text') or ''),
            )
            if key in seen:
                continue
            seen.add(key)
            merged.append(h)
    return merged


def _format_spending_hints_as_line_hints(hints: list[dict]) -> str:
    """
    Group per-line hints into one LINE_HINT row per source line so the model
    sees both money_in and money_out when a single line has both.
    """
    if not hints:
        return ''
    by_line: dict[str, dict] = {}
    order: list[str] = []
    for h in hints:
        line = str(h.get('line_text') or '').strip()
        if not line:
            continue
        if line not in by_line:
            by_line[line] = {'in': None, 'out': None}
            order.append(line)
        slot = 'in' if h.get('direction') == 'incoming' else 'out'
        prev = by_line[line][slot]
        amt = float(h.get('amount', 0))
        # keep the larger amount if duplicate hints for same column
        if prev is None or amt > prev:
            by_line[line][slot] = amt
    out_lines: list[str] = []
    for line in order[:450]:
        cols = by_line[line]
        in_s = f'{cols["in"]:.2f}' if cols['in'] is not None else '0'
        out_s = f'{cols["out"]:.2f}' if cols['out'] is not None else '0'
        out_lines.append(f'LINE_HINT money_in={in_s} money_out={out_s} | {line}')
    return '\n'.join(out_lines)


def _pdfplumber_spending_column_hints(raw: bytes) -> str:
    """
    Infer Paid in / Paid out column positions from header text and emit LINE_HINT rows
    so the model can set direction from geometry, not merchant names.
    """
    try:
        import pdfplumber  # type: ignore[import-untyped]
    except Exception:
        return ''

    out_lines: list[str] = []
    max_lines = 450

    def scan_headers(lines: list) -> tuple[float | None, float | None, float | None]:
        inc_x = None
        out_x = None
        bal_x = None
        in_prefixes = ('paid', 'money', 'amount', 'amt', 'cash')
        out_prefixes = ('paid', 'money', 'amount', 'amt', 'cash')
        header_scan = lines[: min(45, len(lines))]
        for line in header_scan[:25]:
            words = sorted(line, key=lambda w: w.get('x0', 0))
            lowers = [w.get('text', '').lower().rstrip(':') for w in words]
            for i in range(len(words) - 1):
                a, b = lowers[i], lowers[i + 1]
                if a in in_prefixes and b == 'in':
                    inc_x = (words[i]['x0'] + words[i + 1]['x1']) / 2
                if a in out_prefixes and b == 'out':
                    out_x = (words[i]['x0'] + words[i + 1]['x1']) / 2
            for w in words:
                tl = w.get('text', '').lower().rstrip(':')
                if tl == 'balance':
                    bal_x = _word_center_x(w)
                if inc_x is None and tl in ('credits', 'credit', 'deposits', 'deposit'):
                    inc_x = _word_center_x(w)
                if out_x is None and tl in ('debits', 'debit', 'withdrawals', 'withdrawal'):
                    out_x = _word_center_x(w)
        return inc_x, out_x, bal_x

    try:
        with pdfplumber.open(io.BytesIO(raw)) as pdf:
            for page in pdf.pages:
                words = page.extract_words(
                    extra_attrs=['x0', 'x1', 'top', 'bottom'],
                )
                if not words:
                    continue
                lines = _group_pdf_words_into_lines(words)
                inc_x, out_x, bal_x = scan_headers(lines)
                if inc_x is None or out_x is None:
                    # Fallback: split amount columns by largest horizontal gap on this page
                    xs: list[float] = []
                    for line in lines:
                        for w in line:
                            if _word_as_money_amount(w.get('text', '')) is not None:
                                xs.append(_word_center_x(w))
                    xs.sort()
                    if len(xs) >= 6:
                        gaps = [(xs[i + 1] - xs[i], i) for i in range(len(xs) - 1)]
                        biggest, idx = max(gaps, key=lambda g: g[0])
                        if biggest >= 35:
                            left = xs[: idx + 1]
                            right = xs[idx + 1 :]
                            inc_x = sum(left) / len(left)
                            out_x = sum(right) / len(right)

                if inc_x is None or out_x is None:
                    continue

                for line in lines:
                    if len(out_lines) >= max_lines:
                        break
                    money_words = []
                    for w in line:
                        amt = _word_as_money_amount(w.get('text', ''))
                        if amt is None:
                            continue
                        xc = _word_center_x(w)
                        if bal_x is not None and abs(xc - bal_x) < 12:
                            continue
                        money_words.append((amt, xc))
                    if not money_words:
                        continue
                    line_text = ' '.join(w.get('text', '') for w in line).strip()
                    if len(line_text) < 4:
                        continue
                    inc_cands = [(a, xc) for a, xc in money_words if _nearest_column(xc, inc_x, out_x) == 'incoming']
                    out_cands = [(a, xc) for a, xc in money_words if _nearest_column(xc, inc_x, out_x) == 'outgoing']
                    inc_amt = min(inc_cands, key=lambda t: abs(t[1] - inc_x))[0] if inc_cands else None
                    out_amt = min(out_cands, key=lambda t: abs(t[1] - out_x))[0] if out_cands else None
                    if inc_amt is None and out_amt is None:
                        continue
                    in_s = f'{inc_amt:.2f}' if inc_amt is not None else ''
                    out_s = f'{out_amt:.2f}' if out_amt is not None else ''
                    out_lines.append(f'LINE_HINT money_in={in_s or "0"} money_out={out_s or "0"} | {line_text}')
                if len(out_lines) >= max_lines:
                    break
    except Exception as e:
        logger.debug('pdfplumber column hints failed: %s', e)
        out_lines = []

    layout_hints = _spending_direction_hints_from_layout_text(raw)
    layout_text = _format_spending_hints_as_line_hints(layout_hints)
    if out_lines and layout_text:
        return '\n'.join(out_lines) + '\n' + layout_text
    if layout_text:
        return layout_text
    return '\n'.join(out_lines) if out_lines else ''


def _pipeline_text_preview(text: str, max_chars: int) -> dict:
    """API-safe preview of a long string (full lengths + capped content for the UI)."""
    if not text:
        return {'content': '', 'preview_truncated': False, 'total_length': 0}
    total = len(text)
    if total <= max_chars:
        return {'content': text, 'preview_truncated': False, 'total_length': total}
    return {'content': text[:max_chars], 'preview_truncated': True, 'total_length': total}


def _spending_direction_hints_json_sample(hints: list, limit: int) -> list[dict]:
    out: list[dict] = []
    for h in (hints or [])[:limit]:
        toks = h.get('tokens')
        if isinstance(toks, set):
            toks_list = sorted(toks)[:50]
        elif toks:
            toks_list = list(toks)[:50]
        else:
            toks_list = []
        out.append({
            'amount': h.get('amount'),
            'direction': h.get('direction'),
            'line_text': str(h.get('line_text') or '')[:260],
            'match_tokens_sample': toks_list,
        })
    return out


def _build_spending_pdf_statement_text(raw: bytes) -> tuple[str, dict]:
    """
    PDF → plain text plus optional layout-derived LINE_HINT block (same string the model sees,
    before the global char cap). Second return value has pieces for API transparency.
    """
    pdf_meta: dict = {}
    base = _extract_pdf_text(raw, meta_out=pdf_meta)
    hints_str = _pdfplumber_spending_column_hints(raw)
    if hints_str:
        full = base + SPENDING_PDF_COLUMN_HINTS_BANNER + hints_str
    else:
        full = base
    return full, {
        'extracted_text': base,
        'column_hints_block': hints_str,
        'pdf_text_engine': pdf_meta.get('engine'),
        'pdf_engines_considered': pdf_meta.get('engines') or [],
    }


def _extract_pdf_text_for_spending(raw: bytes) -> str:
    text, _ = _build_spending_pdf_statement_text(raw)
    return text


def _hint_tokens_from_line(line_text: str) -> set[str]:
    toks = set(_normalize_label(line_text).split())
    return {t for t in toks if len(t) >= 3 and not t.isdigit()}


def _build_spending_direction_hints(raw: bytes) -> list[dict]:
    """
    Deterministic per-amount direction hints from PDF geometry.
    Returns a list of {'amount': float, 'direction': 'incoming'|'outgoing',
    'tokens': set[str], 'line_text': str} derived from the same column detection
    used for LINE_HINT text. These are used to override LLM direction after extraction.
    """
    try:
        import pdfplumber  # type: ignore[import-untyped]
    except Exception:
        return []

    hints: list[dict] = []

    def scan_headers(lines: list) -> tuple[float | None, float | None, float | None]:
        inc_x = None
        out_x = None
        bal_x = None
        in_prefixes = ('paid', 'money', 'amount', 'amt', 'cash')
        out_prefixes = ('paid', 'money', 'amount', 'amt', 'cash')
        for line in lines[:25]:
            words = sorted(line, key=lambda w: w.get('x0', 0))
            lowers = [w.get('text', '').lower().rstrip(':') for w in words]
            for i in range(len(words) - 1):
                a, b = lowers[i], lowers[i + 1]
                if a in in_prefixes and b == 'in':
                    inc_x = (words[i]['x0'] + words[i + 1]['x1']) / 2
                if a in out_prefixes and b == 'out':
                    out_x = (words[i]['x0'] + words[i + 1]['x1']) / 2
            for w in words:
                tl = w.get('text', '').lower().rstrip(':')
                if tl == 'balance':
                    bal_x = _word_center_x(w)
                if inc_x is None and tl in ('credits', 'credit', 'deposits', 'deposit'):
                    inc_x = _word_center_x(w)
                if out_x is None and tl in ('debits', 'debit', 'withdrawals', 'withdrawal'):
                    out_x = _word_center_x(w)
        return inc_x, out_x, bal_x

    try:
        with pdfplumber.open(io.BytesIO(raw)) as pdf:
            for page in pdf.pages:
                words = page.extract_words(extra_attrs=['x0', 'x1', 'top', 'bottom'])
                if not words:
                    continue
                lines = _group_pdf_words_into_lines(words)
                inc_x, out_x, bal_x = scan_headers(lines)
                if inc_x is None or out_x is None:
                    xs: list[float] = []
                    for line in lines:
                        for w in line:
                            if _word_as_money_amount(w.get('text', '')) is not None:
                                xs.append(_word_center_x(w))
                    xs.sort()
                    if len(xs) >= 6:
                        gaps = [(xs[i + 1] - xs[i], i) for i in range(len(xs) - 1)]
                        biggest, idx = max(gaps, key=lambda g: g[0])
                        if biggest >= 35:
                            left = xs[: idx + 1]
                            right = xs[idx + 1 :]
                            inc_x = sum(left) / len(left)
                            out_x = sum(right) / len(right)
                if inc_x is None or out_x is None:
                    continue
                for line in lines:
                    money_words = []
                    for w in line:
                        amt = _word_as_money_amount(w.get('text', ''))
                        if amt is None:
                            continue
                        xc = _word_center_x(w)
                        if bal_x is not None and abs(xc - bal_x) < 12:
                            continue
                        money_words.append((amt, xc))
                    if not money_words:
                        continue
                    line_text = ' '.join(w.get('text', '') for w in line).strip()
                    if len(line_text) < 4:
                        continue
                    inc_cands = [(a, xc) for a, xc in money_words if _nearest_column(xc, inc_x, out_x) == 'incoming']
                    out_cands = [(a, xc) for a, xc in money_words if _nearest_column(xc, inc_x, out_x) == 'outgoing']
                    inc_amt = min(inc_cands, key=lambda t: abs(t[1] - inc_x))[0] if inc_cands else None
                    out_amt = min(out_cands, key=lambda t: abs(t[1] - out_x))[0] if out_cands else None
                    tokens = _hint_tokens_from_line(line_text)
                    if inc_amt is not None:
                        hints.append({
                            'amount': round(inc_amt, 2),
                            'direction': 'incoming',
                            'tokens': tokens,
                            'line_text': line_text,
                        })
                    if out_amt is not None:
                        hints.append({
                            'amount': round(out_amt, 2),
                            'direction': 'outgoing',
                            'tokens': tokens,
                            'line_text': line_text,
                        })
    except Exception as e:
        logger.debug('direction hints build failed: %s', e)
        hints = []

    layout_hints = _spending_direction_hints_from_layout_text(raw)
    return _merge_spending_hints(hints, layout_hints)


def _reconcile_spending_directions(rows: list, hints: list) -> dict:
    """
    Override each row's direction when PDF geometry indicates otherwise.
    Matching is by amount (to 2dp) plus token overlap with the hint's line text.
    Returns counters {'overridden', 'confirmed', 'unmatched_rows', 'hints_seen'}.
    """
    stats = {
        'overridden': 0,
        'confirmed': 0,
        'unmatched_rows': 0,
        'hints_seen': len(hints) if hints else 0,
    }
    if not rows or not hints:
        return stats

    by_amount: dict[float, list] = {}
    for h in hints:
        by_amount.setdefault(round(float(h['amount']), 2), []).append(h)

    for row in rows:
        try:
            amt = round(float(row.get('amount', 0)), 2)
        except (TypeError, ValueError):
            stats['unmatched_rows'] += 1
            continue
        cands = by_amount.get(amt) or []
        if not cands:
            for delta in (0.01, -0.01):
                cands = by_amount.get(round(amt + delta, 2)) or []
                if cands:
                    break
        if not cands:
            stats['unmatched_rows'] += 1
            continue

        directions = {h['direction'] for h in cands}
        chosen = None
        if len(directions) == 1:
            chosen = next(iter(directions))
        else:
            row_tokens = _hint_tokens_from_line(row.get('description', ''))
            scored = sorted(
                ((len(row_tokens & h['tokens']) if row_tokens else 0, h) for h in cands),
                key=lambda t: t[0],
                reverse=True,
            )
            top_score, top_hint = scored[0]
            second_score = scored[1][0] if len(scored) > 1 else -1
            if top_score > 0 and top_score > second_score:
                chosen = top_hint['direction']

        if not chosen:
            continue
        if row.get('direction') != chosen:
            row['direction'] = chosen
            stats['overridden'] += 1
        else:
            stats['confirmed'] += 1

    return stats


def _parse_iso_date(s: str):
    if not s or not isinstance(s, str):
        return None
    s = s.strip()[:32]
    try:
        return datetime.strptime(s[:10], '%Y-%m-%d').date()
    except ValueError:
        return None


def _parse_bank_statement_date(s: str) -> date | None:
    """
    Parse dates from LLM output or messy statement text.
    Prefer ISO; then UK-style DD/MM/YYYY; then dateutil as last resort.
    """
    if not s or not isinstance(s, str):
        return None
    raw = s.strip()[:64]
    d = _parse_iso_date(raw)
    if d is not None:
        return d
    m = re.match(r'^(\d{1,2})[/.-](\d{1,2})[/.-](\d{2,4})$', raw)
    if m:
        day, month, year = int(m.group(1)), int(m.group(2)), int(m.group(3))
        if year < 100:
            year += 2000
        try:
            return date(year, month, day)
        except ValueError:
            return None
    try:
        return dateutil_parser.parse(raw, dayfirst=True).date()
    except (ValueError, TypeError, OverflowError):
        return None


def _description_similarity(a: str, b: str) -> float:
    a = (a or '').lower().strip()
    b = (b or '').lower().strip()
    if not a or not b:
        return 0.0
    return SequenceMatcher(None, a, b).ratio()


def _extract_transactions_llm(statement_text: str) -> list:
    """
    Sends statement text to OpenAI; returns a list of dicts with statement_date, amount, description, category.
    Third-party: data leaves your server. Set OPENAI_API_KEY (and optionally OPENAI_MODEL, OPENAI_BASE_URL).
    """
    client = _get_openai_client()
    if not client:
        raise RuntimeError('OPENAI_API_KEY is not set')

    model = os.getenv('OPENAI_MODEL', 'gpt-4o-mini').strip() or 'gpt-4o-mini'
    truncated = False
    text = statement_text
    if len(text) > STATEMENT_MAX_CHARS_FOR_LLM:
        text = text[:STATEMENT_MAX_CHARS_FOR_LLM]
        truncated = True

    system = (
        'You extract ONLY outgoing debits from UK bank statements that are shared household bills: '
        'council tax, water, sewerage, electricity, gas, broadband, internet, landline, TV licence, '
        'TV/broadband bundles, home insurance, service charge, or similar utilities. '
        'EXCLUDE: groceries, supermarkets, restaurants, retail, ATM cash, transfers to people, '
        'credit card payments, investments, salary, benefits, general shopping, petrol unless clearly a utility DD. '
        'Return ONLY valid JSON: {"transactions":[{"statement_date":"YYYY-MM-DD","amount":number,"description":string,"category":string|null}]} '
        'where amount is the positive debit amount as on the statement. If nothing matches, return {"transactions":[]}.'
    )
    user_msg = 'Bank statement text:\n\n' + text
    if truncated:
        user_msg += '\n\n(Note: text was truncated; extract from the visible portion only.)'

    completion = client.chat.completions.create(
        model=model,
        messages=[
            {'role': 'system', 'content': system},
            {'role': 'user', 'content': user_msg},
        ],
        response_format={'type': 'json_object'},
        temperature=0.2,
    )
    raw = completion.choices[0].message.content or '{}'
    data, jerr = _parse_llm_json_object(raw, context='loan_statement_extract')
    if jerr:
        raise RuntimeError(
            'Could not parse the model response as JSON. Try again or use a shorter statement excerpt. '
            f'Detail: {jerr}'
        )
    txs = data.get('transactions')
    if txs is None:
        txs = data.get('items') or []
    if not isinstance(txs, list):
        return []
    return txs


def _normalize_candidates(raw_list: list) -> list:
    out = []
    idx = 0
    for row in raw_list:
        if not isinstance(row, dict):
            continue
        d = _parse_iso_date(str(row.get('statement_date') or row.get('date') or ''))
        if d is None:
            continue
        try:
            amount_bank = float(row.get('amount'))
        except (TypeError, ValueError):
            continue
        if amount_bank <= 0:
            continue
        desc = str(row.get('description') or '').strip()[:500]
        if not desc:
            desc = 'Imported repayment'
        cat = row.get('category')
        if cat is not None:
            cat = str(cat)[:120]
        amount_default = round(amount_bank / 2.0, 2)
        out.append({
            'id': idx,
            'statement_date': d.strftime('%Y-%m-%d'),
            'amount_bank': round(amount_bank, 2),
            'amount_default': amount_default,
            'description': desc,
            'category': cat,
        })
        idx += 1
    return out


def _mark_duplicates(candidates: list, loan: dict) -> None:
    """Set possible_duplicate when a candidate matches any existing loan transaction on date and amount."""
    txs = loan.get('transactions', []) or []
    for c in candidates:
        c['possible_duplicate'] = False
        try:
            bank = float(c.get('amount_bank', 0))
            share = float(c.get('amount_default', 0))
        except (TypeError, ValueError):
            continue
        for tx in txs:
            if tx.get('date') != c['statement_date']:
                continue
            try:
                stored = abs(float(tx['amount']))
            except (TypeError, ValueError):
                continue
            if (
                abs(stored - bank) <= STATEMENT_AUDIT_MATCH_TOLERANCE
                or abs(stored - share) <= STATEMENT_AUDIT_MATCH_TOLERANCE
            ):
                c['possible_duplicate'] = True
                break


def _spending_pipeline_dict_pdf(pdf_parts: dict, hints: list, raw_len: int, combined_text: str) -> dict:
    ext = pdf_parts['extracted_text']
    col = pdf_parts['column_hints_block'] or ''
    return {
        'source_format': 'pdf',
        'lengths': {
            'raw_bytes': raw_len,
            'extracted_text': len(ext),
            'column_hints_block': len(col),
            'combined_before_truncation': len(combined_text),
        },
        'pdf': {
            'text_engine': pdf_parts['pdf_text_engine'],
            'engines_considered': pdf_parts['pdf_engines_considered'],
        },
        'previews': {
            'extracted_text': _pipeline_text_preview(ext, SPENDING_PIPELINE_EXTRACT_PREVIEW),
            'column_hints_block': _pipeline_text_preview(col, SPENDING_PIPELINE_HINTS_PREVIEW),
        },
        'direction_hints': {
            'count': len(hints),
            'sample': _spending_direction_hints_json_sample(hints, SPENDING_PIPELINE_DIRECTION_HINT_ROWS),
        },
    }


def _spending_pipeline_dict_text(name: str, text: str, raw_len: int) -> dict:
    n = (name or '').lower()
    if n.endswith('.xlsx') or n.endswith('.xlsm'):
        src_fmt = 'xlsx'
        note = 'Converted Excel workbook to CSV-like text; structured headers parse locally (no LLM extract).'
    elif n.endswith('.csv'):
        src_fmt = 'csv'
        note = 'Decoded as UTF-8 text; structured CSV headers parse locally when recognized.'
    else:
        src_fmt = 'text'
        note = 'Decoded as UTF-8 text; no PDF layout or column-hint pass.'
    return {
        'source_format': src_fmt,
        'lengths': {
            'raw_bytes': raw_len,
            'extracted_text': len(text),
            'column_hints_block': 0,
            'combined_before_truncation': len(text),
        },
        'previews': {
            'extracted_text': _pipeline_text_preview(text, SPENDING_PIPELINE_EXTRACT_PREVIEW),
            'column_hints_block': _pipeline_text_preview('', SPENDING_PIPELINE_HINTS_PREVIEW),
        },
        'direction_hints': {'count': 0, 'sample': []},
        'note': note,
    }


def _spending_text_hints_pipeline_from_raw(name: str, raw: bytes) -> tuple[str, list, dict]:
    """Full spending parse: text for the model, direction hints, pipeline dict (before truncation)."""
    if name.endswith('.pdf'):
        text, pdf_parts = _build_spending_pdf_statement_text(raw)
        hints = _build_spending_direction_hints(raw)
        pipeline = _spending_pipeline_dict_pdf(pdf_parts, hints, len(raw), text)
        return text, hints, pipeline
    if _is_spreadsheet_filename(name):
        text = _extract_spreadsheet_text(raw, name)
        return text, [], _spending_pipeline_dict_text(name, text, len(raw))
    text = raw.decode('utf-8', errors='replace')
    return text, [], _spending_pipeline_dict_text(name, text, len(raw))


def _finalize_spending_upload(text: str, hints: list, pipeline: dict | None) -> tuple[tuple | None, str | None]:
    if not text or len(re.sub(r'\s+', '', text)) < 40:
        return None, 'Extracted text is empty or too short. Try exporting CSV or Excel (.xlsx) from your bank or another PDF.'
    truncated = False
    if len(text) > STATEMENT_MAX_CHARS_FOR_LLM:
        text = text[:STATEMENT_MAX_CHARS_FOR_LLM]
        truncated = True
    if pipeline is not None:
        pipeline['truncated_for_llm'] = truncated
        pipeline['llm_char_limit'] = STATEMENT_MAX_CHARS_FOR_LLM
        pipeline['lengths']['sent_to_llm'] = len(text)
    return (text, truncated, hints, pipeline), None


def _iter_prepare_spending_raw(name: str, raw: bytes):
    """Yield progress dicts; then yield prep_complete with prep tuple, or error dict."""
    hints: list = []
    pipeline: dict | None = None
    text = ''
    try:
        if name.endswith('.pdf'):
            yield {'type': 'progress', 'step': 'pdf_text', 'message': 'Extracting PDF text (layout-aware)…'}
            pdf_meta: dict = {}
            base = _extract_pdf_text(raw, meta_out=pdf_meta)
            yield {'type': 'progress', 'step': 'pdf_column_hints', 'message': 'Deriving column hints from statement layout…'}
            hints_str = _pdfplumber_spending_column_hints(raw)
            if hints_str:
                text = base + SPENDING_PDF_COLUMN_HINTS_BANNER + hints_str
            else:
                text = base
            pdf_parts = {
                'extracted_text': base,
                'column_hints_block': hints_str,
                'pdf_text_engine': pdf_meta.get('engine'),
                'pdf_engines_considered': pdf_meta.get('engines') or [],
            }
            yield {'type': 'progress', 'step': 'pdf_direction_hints', 'message': 'Deriving credit/debit hints from geometry…'}
            hints = _build_spending_direction_hints(raw)
            pipeline = _spending_pipeline_dict_pdf(pdf_parts, hints, len(raw), text)
        elif _is_spreadsheet_filename(name):
            yield {'type': 'progress', 'step': 'decode_spreadsheet', 'message': 'Converting Excel spreadsheet…'}
            text = _extract_spreadsheet_text(raw, name)
            pipeline = _spending_pipeline_dict_text(name, text, len(raw))
        else:
            yield {'type': 'progress', 'step': 'decode_text', 'message': 'Decoding text or CSV…'}
            text = raw.decode('utf-8', errors='replace')
            pipeline = _spending_pipeline_dict_text(name, text, len(raw))
    except Exception as e:
        logger.exception('statement read failed')
        yield {'type': 'error', 'message': f'Could not read file: {e}', 'http_status': 400}
        return

    yield {'type': 'progress', 'step': 'validate', 'message': 'Validating text and applying model size limit…'}
    fin, err = _finalize_spending_upload(text, hints, pipeline)
    if err:
        yield {'type': 'error', 'message': err, 'http_status': 400}
        return
    yield {'type': 'prep_complete', 'prep': fin}


def _spending_statement_preview_finalize(
    data: dict,
    spending: dict,
    period: dict,
    direction_hints: list,
    pipeline: dict,
    truncated_text: bool,
    raw_rows: list,
    extraction_meta: dict | None = None,
) -> dict:
    rows = _normalize_spending_transactions(raw_rows)
    before_ct = len(rows)
    rows, dropped_outside, boundary_count = _filter_spending_rows_by_period(
        rows, period['period_start_date'], period['period_end_date']
    )
    direction_stats = _reconcile_spending_directions(rows, direction_hints)
    cache_changed = _apply_outgoing_classification(rows, spending)
    if cache_changed:
        save_data(data)

    rm = period.get('report_month') or ''
    for r in rows:
        r['report_month'] = rm
    tr_preview = simulate_spending_transfer_reconciliation_preview(spending, rm, rows)
    for r in rows:
        rid = str(r.get('id') or '')
        r['reconciliation'] = (tr_preview.get('pairing') or {}).get(rid) or {
            'paired': False,
            'peer_id': None,
            'peer_is_from_ledger': False,
            'peer_description': None,
        }

    d_led, d_up, dup_ledger_fps = _apply_spending_preview_duplicate_marks(rm, rows, spending)
    missed_n = sum(1 for r in rows if r.get('preview_review_reason') == 'missed')
    expected_bill_n = sum(1 for r in rows if r.get('preview_review_reason') == 'expected_bill')

    incoming_total = round(sum(r['amount'] for r in rows if r.get('direction') == 'incoming'), 2)
    outgoing_total = round(sum(r['amount'] for r in rows if r.get('direction') == 'outgoing'), 2)
    months = sorted({r.get('month') for r in rows if r.get('month')}, reverse=True)
    summary = {
        'report_month': period['report_month'],
        'period_start': period['period_start'],
        'period_end': period['period_end'],
        'total_rows': len(rows),
        'raw_extraction_count': before_ct,
        'filtered_out_count': dropped_outside,
        'date_boundary_count': boundary_count,
        'incoming_total': incoming_total,
        'outgoing_total': outgoing_total,
        'net': round(incoming_total - outgoing_total, 2),
        'months': months,
        'direction_reconciliation': direction_stats,
        'transfer_reconciliation_preview': {
            'reconciliation': tr_preview.get('reconciliation'),
            'ledger_row_count_in_month': tr_preview.get('ledger_row_count_in_month', 0),
            'auto_applied_pairs_in_simulation': tr_preview.get('auto_applied_pairs_in_simulation', 0),
        },
        'preview_duplicate_ledger': d_led,
        'preview_duplicate_upload': d_up,
        'preview_missed_manual': missed_n,
        'preview_expected_bill': expected_bill_n,
        'duplicate_ledger_fingerprints': dup_ledger_fps,
    }
    if extraction_meta:
        summary['extraction'] = extraction_meta
    return {
        'transactions': rows,
        'truncated': truncated_text,
        'pipeline': pipeline,
        'summary': summary,
    }


def _spending_statement_preview_payload(
    text: str,
    truncated_text: bool,
    direction_hints: list,
    pipeline: dict,
    period: dict,
) -> dict:
    """Build the same JSON body as spending_statement_preview after file prep."""
    data = load_data()
    spending, changed = _ensure_user_spending(data, session['username'])
    if changed:
        save_data(data)

    period_hint = (
        f'Prefer transactions whose completed/settled date (or sole statement date) falls on or between '
        f'{period["period_start"]} and {period["period_end"]} inclusive. '
        f'When a row has both started and completed dates, always set date=completed_date even if started_date '
        f'is outside that range — still include the row. When in doubt, include the row.'
    )
    raw_rows = None
    extraction_meta = None
    for ev in iter_spending_transaction_extraction(text, period_hint):
        if ev.get('type') == 'result':
            raw_rows = ev.get('rows')
            extraction_meta = ev.get('meta')
    if raw_rows is None:
        raw_rows = []
    return _spending_statement_preview_finalize(
        data,
        spending,
        period,
        direction_hints,
        pipeline,
        truncated_text,
        raw_rows,
        extraction_meta,
    )


def _prepare_statement_text_from_upload(*, for_spending: bool = False):
    """
    Read multipart file from request.
    Returns ((text, truncated_bool, hints_list, pipeline_dict), None) when for_spending=True,
    else ((text, truncated_bool), None). On error returns (None, message).
    hints_list is direction hints from PDF geometry (empty for non-PDF).
    pipeline_dict describes intermediate representations for the spending UI (PDF stages, previews).
    """
    f = request.files.get('file')
    if not f or not getattr(f, 'filename', None):
        return None, 'No file uploaded'
    raw = f.read()
    if len(raw) > STATEMENT_MAX_BYTES:
        return None, 'File too large (max 2 MB).'
    name = (f.filename or '').lower()
    hints: list = []
    pipeline: dict | None = None
    try:
        if for_spending:
            text, hints, pipeline = _spending_text_hints_pipeline_from_raw(name, raw)
        elif name.endswith('.pdf'):
            text = _extract_pdf_text(raw)
        elif _is_spreadsheet_filename(name):
            text = _extract_spreadsheet_text(raw, name)
        else:
            text = raw.decode('utf-8', errors='replace')
    except Exception as e:
        logger.exception('statement read failed')
        return None, f'Could not read file: {e}'
    if for_spending:
        fin, ferr = _finalize_spending_upload(text, hints, pipeline)
        if ferr:
            return None, ferr
        return fin, None
    if not text or len(re.sub(r'\s+', '', text)) < 40:
        return None, 'Extracted text is empty or too short. Try exporting CSV or Excel (.xlsx) from your bank or another PDF.'
    truncated = False
    if len(text) > STATEMENT_MAX_CHARS_FOR_LLM:
        text = text[:STATEMENT_MAX_CHARS_FOR_LLM]
        truncated = True
    return (text, truncated), None


def _extract_baseline_statement_llm(statement_text: str) -> tuple[list, str | None]:
    """
    Broader extraction for expected recurring bills + statement period_end.
    Returns (transactions_list, period_end_iso_or_none).
    """
    client = _get_openai_client()
    if not client:
        raise RuntimeError('OPENAI_API_KEY is not set')

    model = os.getenv('OPENAI_MODEL', 'gpt-4o-mini').strip() or 'gpt-4o-mini'
    truncated = False
    text = statement_text
    if len(text) > STATEMENT_MAX_CHARS_FOR_LLM:
        text = text[:STATEMENT_MAX_CHARS_FOR_LLM]
        truncated = True

    system = (
        'You analyse UK bank statement text. Return ONLY valid JSON with keys: '
        '"transactions" (array) and "period_end" (string or null). '
        '"transactions": outgoing debits that could plausibly be recurring monthly household bills '
        '(council tax, water, electricity, gas, broadband, rent DD, insurance, phone, TV licence, service charges, etc.). '
        'Be more inclusive than a strict filter but EXCLUDE: supermarkets, restaurants, pubs, coffee, '
        'ATM cash, transfers to people, salary, benefits, gambling, general shopping unless clearly a utility DD. '
        'Each item: statement_date (YYYY-MM-DD), amount (positive debit), description (string), category (string or null). '
        '"period_end": the statement period end date as YYYY-MM-DD if visible in the text, else null.'
    )
    user_msg = 'Bank statement text:\n\n' + text
    if truncated:
        user_msg += '\n\n(Note: text was truncated.)'

    completion = client.chat.completions.create(
        model=model,
        messages=[
            {'role': 'system', 'content': system},
            {'role': 'user', 'content': user_msg},
        ],
        response_format={'type': 'json_object'},
        temperature=0.2,
    )
    raw = completion.choices[0].message.content or '{}'
    data, jerr = _parse_llm_json_object(raw, context='baseline_statement')
    if jerr:
        logger.warning('baseline_statement LLM JSON parse failed: %s', jerr)
        return [], None
    txs = data.get('transactions')
    if txs is None:
        txs = data.get('items') or []
    if not isinstance(txs, list):
        txs = []
    pe = data.get('period_end')
    if pe is not None and not isinstance(pe, str):
        pe = str(pe) if pe else None
    if pe:
        d = _parse_iso_date(pe)
        pe = d.strftime('%Y-%m-%d') if d else None
    return txs, pe


def _default_reminder_day_from_period_end(period_end: str | None) -> int:
    d = _parse_iso_date(period_end or '')
    if d is None:
        return 1
    nxt = d + timedelta(days=1)
    return nxt.day


def _amounts_close_for_compare(a: float, b: float) -> bool:
    if abs(a - b) <= 0.05:
        return True
    m = max(abs(a), abs(b), 1e-9)
    return abs(a - b) / m <= 0.02


def _compare_to_baseline(candidates: list, baseline: list | None) -> dict | None:
    if not baseline:
        return None
    missing = []
    amount_changed = []
    used_c = set()

    for b in baseline:
        best_j = None
        best_sim = -1.0
        bd = str(b.get('description', ''))
        for j, c in enumerate(candidates):
            if j in used_c:
                continue
            sim = _description_similarity(bd, c.get('description', ''))
            if sim > best_sim:
                best_sim = sim
                best_j = j
        if best_j is None or best_sim < BASELINE_SIM_THRESHOLD:
            missing.append(b)
            continue
        used_c.add(best_j)
        c = candidates[best_j]
        try:
            ab = float(b.get('amount_bank', 0))
            cb = float(c.get('amount_bank', 0))
        except (TypeError, ValueError):
            missing.append(b)
            used_c.discard(best_j)
            continue
        if not _amounts_close_for_compare(ab, cb):
            amount_changed.append({
                'baseline': b,
                'found': c,
                'expected_amount_bank': round(ab, 2),
                'actual_amount_bank': round(cb, 2),
            })

    new = [candidates[j] for j in range(len(candidates)) if j not in used_c]
    return {'missing': missing, 'new': new, 'amount_changed': amount_changed}


def _get_smtp_config():
    host = os.getenv('SMTP_HOST', '').strip()
    if not host:
        return None
    return {
        'host': host,
        'port': int(os.getenv('SMTP_PORT', '587')),
        'user': os.getenv('SMTP_USER', '').strip(),
        'password': os.getenv('SMTP_PASSWORD', ''),
        'use_tls': _env_bool('SMTP_USE_TLS', True),
        'mail_from': os.getenv('MAIL_FROM', '').strip(),
    }


def _recipient_for_loan(loan: dict) -> str:
    return (loan.get('bill_reminder_email') or '').strip() or os.getenv(
        'FINANCE_TRACKER_NOTIFY_EMAIL', ''
    ).strip()


def _public_app_base_url():
    base = os.getenv('PUBLIC_BASE_URL', '').strip().rstrip('/')
    if base:
        return base
    return request.url_root.rstrip('/') if request else ''


def _send_bill_reminder_email(loan_id: str, loan: dict) -> bool:
    cfg = _get_smtp_config()
    to_addr = _recipient_for_loan(loan)
    if not cfg or not cfg['mail_from'] or not to_addr:
        logger.warning('bill reminder skipped: SMTP or recipient not configured')
        return False
    name = loan.get('name', 'Loan')
    base = os.getenv('PUBLIC_BASE_URL', '').strip().rstrip('/')
    link = f'{base}/loan/{loan_id}' if base else f'/loan/{loan_id}'
    msg = EmailMessage()
    msg['Subject'] = f'Finance tracker: upload statement for "{name}"'
    msg['From'] = cfg['mail_from']
    msg['To'] = to_addr
    msg.set_content(
        f"""Time to upload your latest bank statement for "{name}" and import shared bill repayments.

Open the loan: {link}

If you use expected bills, compare the import preview to your baseline for any changes.
"""
    )
    try:
        with smtplib.SMTP(cfg['host'], cfg['port'], timeout=30) as smtp:
            if cfg['use_tls']:
                smtp.starttls()
            if cfg['user']:
                smtp.login(cfg['user'], cfg['password'])
            smtp.send_message(msg)
        return True
    except Exception:
        logger.exception('SMTP send failed for loan %s', loan_id)
        return False


def tick_bill_reminders():
    """Scheduled daily: send at most one reminder per loan per calendar month on the configured day."""
    data = load_data()
    today = date.today()
    this_month = today.strftime('%Y-%m')
    changed = False
    for loan_id, loan in data['loans'].items():
        if loan.get('deleted'):
            continue
        if not loan.get('bill_baseline'):
            continue
        br = loan.get('bill_reminder') or {}
        day = int(br.get('day_of_month', 1))
        if not 1 <= day <= 31:
            day = 1
        last_day = monthrange(today.year, today.month)[1]
        effective = min(day, last_day)
        if today.day != effective:
            continue
        if br.get('last_sent_month') == this_month:
            continue
        if not _send_bill_reminder_email(loan_id, loan):
            continue
        br['last_sent_month'] = this_month
        loan['bill_reminder'] = br
        changed = True
    if changed:
        save_data(data)


def schedule_bill_reminders_job():
    job_id = 'bill_reminders_global'
    for job in scheduler.get_jobs():
        if job.id == job_id:
            scheduler.remove_job(job_id)
    scheduler.add_job(
        tick_bill_reminders,
        CronTrigger(hour=BILL_REMINDER_HOUR, minute=0),
        id=job_id,
        replace_existing=True,
    )


def _login_form_credentials():
    """Normal form POST; also accept JSON for non-browser clients."""
    username = (request.form.get('username') or '').strip()
    password = (request.form.get('password') or '').strip()
    if username or password:
        return username, password
    if request.is_json:
        data = request.get_json(silent=True) or {}
        return (
            str(data.get('username', '') or '').strip(),
            str(data.get('password', '') or '').strip(),
        )
    return username, password


@app.route('/login', methods=['GET', 'POST'])
def login():
    if request.method == 'POST':
        username, password = _login_form_credentials()
        if username in USERS and USERS[username] == password:
            session.permanent = True
            session['username'] = username
            return redirect(url_for('index'))
        logger.warning(
            'login failed: user_in_map=%s form_keys=%s content_type=%s',
            username in USERS,
            list(request.form.keys()),
            request.content_type,
        )
        return render_template('login.html', error='Invalid username or password')

    return render_template('login.html')

@app.route('/logout')
def logout():
    session.pop('username', None)
    return redirect(url_for('login'))

@app.route('/')
@login_required
def index():
    data = load_data()
    spending, changed = _ensure_user_spending(data, session['username'])
    if changed:
        save_data(data)

    daily = _daily_budget_status(spending)
    plan_figures = daily.get('plan') or {}
    daily_has_plan = float(plan_figures.get('income_monthly') or 0) > 0

    insight_map = spending.get('monthly_insights') or {}
    months = sorted(insight_map.keys(), reverse=True)
    latest_month = months[0] if months else None
    latest_insight = insight_map.get(latest_month) if latest_month else None
    latest_net = 0.0
    latest_savings_rate = None
    if isinstance(latest_insight, dict):
        try:
            latest_net = float(latest_insight.get('net') or 0)
        except (TypeError, ValueError):
            latest_net = 0.0
        try:
            sr = latest_insight.get('savings_rate')
            latest_savings_rate = float(sr) if sr is not None else None
        except (TypeError, ValueError):
            latest_savings_rate = None

    loans = data.get('loans') or {}
    active_loans = [loan for loan in loans.values() if not loan.get('deleted')]
    loan_count = len(active_loans)
    loan_total = 0.0
    for loan in active_loans:
        try:
            loan_total += float(loan.get('loan_amount') or 0)
        except (TypeError, ValueError):
            pass
    loan_total = round(loan_total, 2)

    statements = list(spending.get('statements') or [])
    statements.sort(key=lambda s: str(s.get('uploaded_at') or ''), reverse=True)
    recent_statements = statements[:8]
    known_bank_sources = _collect_bank_sources(spending)

    return render_template(
        'home.html',
        username=session.get('username'),
        spending_categories=SPENDING_ALLOWED_CATEGORIES,
        daily_has_plan=daily_has_plan,
        daily_remaining=float(daily.get('remaining_today') or 0),
        daily_limit=float(daily.get('daily_limit') or 0),
        daily_spent=float(daily.get('spent_today') or 0),
        latest_month=latest_month,
        latest_net=latest_net,
        latest_savings_rate=latest_savings_rate,
        loan_count=loan_count,
        loan_total=loan_total,
        recent_statements=recent_statements,
        known_bank_sources=known_bank_sources,
    )


@app.route('/loans')
@login_required
def loans_list():
    data = load_data()
    loans_summary = []
    loan_total = 0.0
    active_count = 0
    for loan_id, loan in data['loans'].items():
        total_loan_quantity, total_paid = calculate_loan_stats(loan['transactions'])
        deleted = loan.get('deleted', False)
        balance = float(loan['loan_amount'] or 0)
        if not deleted:
            active_count += 1
            loan_total += balance
        loans_summary.append({
            'id': loan_id,
            'name': loan['name'],
            'current_balance': loan['loan_amount'],
            'total_loan_quantity': total_loan_quantity,
            'total_paid': total_paid,
            'deleted': deleted
        })
    return render_template(
        'loans.html',
        loans=loans_summary,
        username=session.get('username'),
        loan_count=active_count,
        loan_total=round(loan_total, 2),
    )


@app.route('/spending')
@login_required
def spending_tab():
    data = load_data()
    spending, changed = _ensure_user_spending(data, session['username'])
    if changed:
        save_data(data)
    months = sorted((spending.get('monthly_insights') or {}).keys(), reverse=True)
    return render_template(
        'spending.html',
        username=session.get('username'),
        months=months,
        spending_categories=SPENDING_ALLOWED_CATEGORIES,
    )


@app.route('/spending/daily')
@login_required
def daily_budget_tab():
    data = load_data()
    spending, changed = _ensure_user_spending(data, session['username'])
    if changed:
        save_data(data)
    months = sorted((spending.get('monthly_insights') or {}).keys(), reverse=True)
    return render_template(
        'daily_budget.html',
        username=session.get('username'),
        months=months,
        spending_categories=DAILY_ENTRY_CATEGORIES,
        daily_modes=sorted(DAILY_BUDGET_MODES),
    )


@app.route('/spending/search')
@login_required
def spending_search_tab():
    """Global transaction search across all months (manual + statement lines)."""
    data = load_data()
    spending, changed = _ensure_user_spending(data, session['username'])
    if changed:
        save_data(data)
    return render_template(
        'spending_search.html',
        username=session.get('username'),
        spending_categories=SPENDING_ALLOWED_CATEGORIES,
        known_bank_sources=_collect_bank_sources(spending),
    )


@app.route('/loan/<loan_id>')
@login_required
def loan_details(loan_id):
    data = load_data()
    if loan_id not in data['loans']:
        return redirect(url_for('loans_list'))
    
    loan = data['loans'][loan_id]
    total_loan_quantity, total_paid = calculate_loan_stats(loan['transactions'])
    loan_deleted = loan.get('deleted', False)
    recurring_payments = loan.get('recurring_payments', [])
    bill_baseline = loan.get('bill_baseline') or []
    bill_reminder = loan.get('bill_reminder')
    bill_reminder_email = loan.get('bill_reminder_email') or ''
    smtp_configured = _get_smtp_config() is not None
    default_notify_email = os.getenv('FINANCE_TRACKER_NOTIFY_EMAIL', '').strip()

    user_bucket = (data.get('users') or {}).get(session['username']) or {}
    user_spending = user_bucket.get('spending') or {}
    outgoing_spending = [
        t for t in (user_spending.get('transactions') or [])
        if t.get('direction') == 'outgoing'
    ]
    spending_months = sorted(_spending_report_months_with_data(outgoing_spending), reverse=True)

    return render_template('index.html',
                         loan_id=loan_id,
                         spending_months=spending_months,
                         username=session.get('username'),
                         loan_name=loan['name'],
                         loan_amount=loan['loan_amount'],
                         interest_rate=loan['interest_rate'],
                         interest_day=loan['interest_day'],
                         transactions=loan['transactions'],
                         total_loan_quantity=total_loan_quantity,
                         total_paid=total_paid,
                         loan_deleted=loan_deleted,
                         recurring_payments=recurring_payments,
                         bill_baseline=bill_baseline,
                         bill_reminder=bill_reminder,
                         bill_reminder_email=bill_reminder_email,
                         smtp_configured=smtp_configured,
                         default_notify_email=default_notify_email)

@app.route('/api/loan', methods=['POST'])
@login_required
def create_loan():
    data = load_data()
    loan_data = request.get_json()
    
    # Generate a unique ID for the new loan
    loan_id = f"loan_{len(data['loans']) + 1}"
    
    new_loan = {
        'name': loan_data['name'],
        'loan_amount': float(loan_data['loan_amount']),
        'interest_rate': float(loan_data['interest_rate']),
        'interest_day': 1,  # Default to 1st of month
        'transactions': [{
            'date': datetime.now().strftime('%Y-%m-%d'),
            'type': 'initial',
            'amount': float(loan_data['loan_amount']),
            'description': 'Initial loan amount',
            'user': session['username']
        }]
    }
    
    data['loans'][loan_id] = new_loan
    save_data(data)
    schedule_interest_task(loan_id, new_loan['interest_day'])
    
    return jsonify({'id': loan_id, **new_loan})

@app.route('/api/loan/<loan_id>/update_interest_day', methods=['POST'])
@login_required
def update_interest_day(loan_id):
    data = load_data()
    if loan_id not in data['loans']:
        return jsonify({'error': 'Loan not found'}), 404
    
    loan = data['loans'][loan_id]
    new_day = int(request.get_json()['interest_day'])
    old_day = loan['interest_day']
    
    if not 1 <= new_day <= 31:
        return jsonify({'error': 'Day must be between 1 and 31'}), 400
    
    if new_day != old_day:
        # Add the change to transaction history
        new_transaction = {
            'date': datetime.now().strftime('%Y-%m-%d'),
            'type': 'settings',
            'amount': 0,
            'description': f'Interest day changed from {old_day} to {new_day}',
            'user': session['username']
        }
        loan['transactions'].append(new_transaction)
        
        # Update the interest day
        loan['interest_day'] = new_day
        save_data(data)
        schedule_interest_task(loan_id, new_day)
    
    return jsonify(loan)

@app.route('/api/loan/<loan_id>/update_interest_rate', methods=['POST'])
@login_required
def update_interest_rate(loan_id):
    data = load_data()
    if loan_id not in data['loans']:
        return jsonify({'error': 'Loan not found'}), 404
    
    loan = data['loans'][loan_id]
    new_rate = float(request.get_json()['interest_rate'])
    old_rate = loan['interest_rate']
    
    if new_rate < 0 or new_rate > 100:
        return jsonify({'error': 'Interest rate must be between 0 and 100'}), 400
    
    if new_rate != old_rate:
        # Add the change to transaction history
        new_transaction = {
            'date': datetime.now().strftime('%Y-%m-%d'),
            'type': 'settings',
            'amount': 0,
            'description': f'Interest rate changed from {old_rate:.2f}% to {new_rate:.2f}%',
            'user': session['username']
        }
        loan['transactions'].append(new_transaction)
        
        # Update the interest rate
        loan['interest_rate'] = new_rate
        save_data(data)
    
    return jsonify(loan)

@app.route('/api/loan/<loan_id>/transaction', methods=['POST'])
@login_required
def add_transaction(loan_id):
    data = load_data()
    if loan_id not in data['loans']:
        return jsonify({'error': 'Loan not found'}), 404
    
    loan = data['loans'][loan_id]
    transaction = request.get_json()
    
    amount = float(transaction['amount'])
    transaction_type = transaction['type']
    
    if transaction_type == 'repayment':
        amount = -amount  # Repayments reduce the loan amount
    
    new_transaction = {
        'date': datetime.now().strftime('%Y-%m-%d'),
        'type': transaction_type,
        'amount': amount,
        'description': transaction['description'],
        'user': session['username']
    }
    
    loan['transactions'].append(new_transaction)
    loan['loan_amount'] += amount
    
    save_data(data)
    return jsonify(loan)

@app.route('/api/loan/<loan_id>/apply_interest', methods=['POST'])
@login_required
def apply_interest(loan_id):
    data = load_data()
    if loan_id not in data['loans']:
        return jsonify({'error': 'Loan not found'}), 404

    loan = data['loans'][loan_id]
    if loan.get('deleted'):
        return jsonify({'error': 'Loan is deleted'}), 400

    body = request.get_json(silent=True) or {}
    year = body.get('year')
    month = body.get('month')
    rate = loan['interest_rate']
    username = session['username']

    if year is not None or month is not None:
        if year is None or month is None:
            return jsonify({'error': 'Both year and month are required'}), 400
        try:
            year = int(year)
            month = int(month)
            as_of = resolve_interest_as_of_date(year, month, loan.get('interest_day', 1))
        except (TypeError, ValueError) as e:
            return jsonify({'error': str(e) or 'Invalid year or month'}), 400

        if interest_applied_in_month(loan.get('transactions'), year, month):
            return jsonify({
                'error': f'Interest has already been applied for {year:04d}-{month:02d}'
            }), 400

        balance = loan_balance_as_of_end_of_day(loan.get('transactions'), as_of)
        monthly_interest = compute_monthly_interest_amount(balance, rate)
        as_of_str = as_of.strftime('%Y-%m-%d')
        new_transaction = {
            'date': as_of_str,
            'type': 'interest',
            'amount': monthly_interest,
            'description': f'Monthly interest at {rate}% APR (as of {as_of_str})',
            'user': username,
        }
    else:
        today = date.today()
        if interest_applied_in_month(loan.get('transactions'), today.year, today.month):
            return jsonify({
                'error': f'Interest has already been applied for {today.year:04d}-{today.month:02d}'
            }), 400

        monthly_interest = compute_monthly_interest_amount(loan['loan_amount'], rate)
        new_transaction = {
            'date': today.strftime('%Y-%m-%d'),
            'type': 'interest',
            'amount': monthly_interest,
            'description': f'Monthly interest at {rate}% APR',
            'user': username,
        }

    loan['transactions'].append(new_transaction)
    loan['loan_amount'] += monthly_interest

    save_data(data)
    return jsonify(loan)

@app.route('/api/loan/<loan_id>/update_name', methods=['POST'])
@login_required
def update_loan_name(loan_id):
    data = load_data()
    if loan_id not in data['loans']:
        return jsonify({'error': 'Loan not found'}), 404
    
    loan = data['loans'][loan_id]
    new_name = request.get_json()['name']
    old_name = loan['name']
    
    if not new_name or new_name.strip() == '':
        return jsonify({'error': 'Loan name cannot be empty'}), 400
    
    if new_name != old_name:
        # Add the change to transaction history
        new_transaction = {
            'date': datetime.now().strftime('%Y-%m-%d'),
            'type': 'settings',
            'amount': 0,
            'description': f'Loan name changed from "{old_name}" to "{new_name}"',
            'user': session['username']
        }
        loan['transactions'].append(new_transaction)
        
        # Update the loan name
        loan['name'] = new_name
        save_data(data)
    
    return jsonify(loan)

def update_loan_status(loan_id, deleted_status, action_description):
    """Helper function to update loan status and add transaction history entry."""
    data = load_data()
    if loan_id not in data['loans']:
        return None, 'Loan not found', 404
    
    # Update loan status
    data['loans'][loan_id]['deleted'] = deleted_status
    
    # Add status change event to transaction history
    new_transaction = {
        'date': datetime.now().strftime('%Y-%m-%d'),
        'type': 'settings',
        'amount': 0,
        'description': action_description,
        'user': session['username']
    }
    data['loans'][loan_id]['transactions'].append(new_transaction)
    
    save_data(data)
    return data['loans'][loan_id], f'Loan {action_description.lower()}', 200

@app.route('/api/loan/<loan_id>/<action>', methods=['POST'])
def update_loan_state(loan_id, action):
    try:
        if action not in ['delete', 'recover']:
            return jsonify({'error': 'Invalid action'}), 400
            
        loan_data, message, status_code = update_loan_status(
            loan_id, 
            action == 'delete',  # True for delete, False for recover
            f'Loan {"marked as deleted" if action == "delete" else "recovered from deleted state"}'
        )
        
        if status_code != 200:
            return jsonify({'error': message}), status_code
            
        return jsonify({'message': message})
    except Exception as e:
        return jsonify({'error': str(e)}), 500

def process_recurring_payments(loan_id):
    """Process recurring payments that are due today."""
    data = load_data()
    if loan_id not in data['loans']:
        return
        
    loan = data['loans'][loan_id]
    if 'recurring_payments' not in loan:
        return
        
    today = date.today()
    for payment in loan['recurring_payments']:
        next_payment_date = datetime.strptime(payment['next_payment_date'], '%Y-%m-%d').date()
        
        if today >= next_payment_date:
            # Add the payment as a transaction
            new_transaction = {
                'date': today.strftime('%Y-%m-%d'),
                'type': 'repayment',
                'amount': -float(payment['amount']),  # Negative for repayments
                'description': f'Recurring payment ({payment["schedule"]})',
                'user': 'system'  # Automated payment
            }
            loan['transactions'].append(new_transaction)
            loan['loan_amount'] += new_transaction['amount']  # Add negative amount to reduce balance
            
            # Calculate next payment date
            if payment['schedule'] == 'monthly':
                next_payment_date = next_payment_date + relativedelta(months=1)
            elif payment['schedule'] == 'bi-weekly':
                next_payment_date = next_payment_date + relativedelta(weeks=2)
            elif payment['schedule'] == 'weekly':
                next_payment_date = next_payment_date + relativedelta(weeks=1)
                
            # Update next payment date
            payment['next_payment_date'] = next_payment_date.strftime('%Y-%m-%d')
    
    save_data(data)

def schedule_recurring_payments(loan_id):
    """Schedule recurring payment processing for this loan."""
    # Remove existing recurring payment jobs for this loan
    job_id = f'recurring_payments_{loan_id}'
    for job in scheduler.get_jobs():
        if job.id == job_id:
            scheduler.remove_job(job_id)
    
    # Schedule new recurring payment job to run daily
    scheduler.add_job(
        lambda: process_recurring_payments(loan_id),
        CronTrigger(hour=0, minute=0),  # Run at midnight
        id=job_id,
        replace_existing=True
    )

@app.route('/api/loan/<loan_id>/recurring_payment', methods=['POST'])
def add_recurring_payment(loan_id):
    try:
        data = load_data()
        if loan_id not in data['loans']:
            return jsonify({'error': 'Loan not found'}), 404
            
        payment_data = request.get_json()
        amount = float(payment_data['amount'])
        schedule = payment_data['schedule']
        start_date = datetime.strptime(payment_data['start_date'], '%Y-%m-%d').date()
        today = date.today()
        
        # Calculate next payment date based on schedule
        next_payment_date = start_date
        if schedule == 'monthly':
            next_payment_date = start_date + relativedelta(months=1)
        elif schedule == 'bi-weekly':
            next_payment_date = start_date + relativedelta(weeks=2)
        elif schedule == 'weekly':
            next_payment_date = start_date + relativedelta(weeks=1)
        else:
            return jsonify({'error': 'Invalid schedule'}), 400
            
        # Create new recurring payment
        new_payment = {
            'amount': amount,
            'schedule': schedule,
            'start_date': start_date.strftime('%Y-%m-%d'),
            'next_payment_date': next_payment_date.strftime('%Y-%m-%d')
        }
        
        # Initialize recurring_payments list if it doesn't exist
        if 'recurring_payments' not in data['loans'][loan_id]:
            data['loans'][loan_id]['recurring_payments'] = []
            
        data['loans'][loan_id]['recurring_payments'].append(new_payment)
        
        # If start date is today, process the payment immediately
        if start_date == today:
            # Add the payment as a transaction
            new_transaction = {
                'date': today.strftime('%Y-%m-%d'),
                'type': 'repayment',
                'amount': -amount,  # Negative for repayments
                'description': f'Recurring payment ({schedule})',
                'user': 'system'
            }
            data['loans'][loan_id]['transactions'].append(new_transaction)
            data['loans'][loan_id]['loan_amount'] += new_transaction['amount']  # Add negative amount to reduce balance
        
        save_data(data)
        
        # Schedule recurring payment processing
        schedule_recurring_payments(loan_id)
        
        return jsonify({'message': 'Recurring payment added successfully', 'payment': new_payment})
    except Exception as e:
        return jsonify({'error': str(e)}), 500

@app.route('/api/loan/<loan_id>/recurring_payment/<int:payment_index>', methods=['DELETE'])
def delete_recurring_payment(loan_id, payment_index):
    try:
        data = load_data()
        if loan_id not in data['loans']:
            return jsonify({'error': 'Loan not found'}), 404
            
        if 'recurring_payments' not in data['loans'][loan_id]:
            return jsonify({'error': 'No recurring payments found'}), 404
            
        if payment_index < 0 or payment_index >= len(data['loans'][loan_id]['recurring_payments']):
            return jsonify({'error': 'Invalid payment index'}), 400
            
        # Remove the payment
        deleted_payment = data['loans'][loan_id]['recurring_payments'].pop(payment_index)
        
        # Add a transaction to record the deletion
        new_transaction = {
            'date': datetime.now().strftime('%Y-%m-%d'),
            'type': 'settings',
            'amount': 0,
            'description': f'Recurring payment deleted (£{deleted_payment["amount"]:.2f} {deleted_payment["schedule"]})',
            'user': session['username']
        }
        data['loans'][loan_id]['transactions'].append(new_transaction)
        
        save_data(data)
        
        return jsonify({'message': 'Recurring payment deleted successfully', 'payment': deleted_payment})
    except Exception as e:
        return jsonify({'error': str(e)}), 500


@app.route('/api/loan/<loan_id>/statement/preview', methods=['POST'])
@login_required
def statement_preview(loan_id):
    """
    Upload a bank statement (PDF, CSV, Excel .xlsx, or plain text). Text is sent to OpenAI for extraction.
    Env: OPENAI_API_KEY (required), OPENAI_MODEL (default gpt-4o-mini), OPENAI_BASE_URL (optional).
    """
    data = load_data()
    if loan_id not in data['loans']:
        return jsonify({'error': 'Loan not found'}), 404
    loan = data['loans'][loan_id]
    if loan.get('deleted'):
        return jsonify({'error': 'Loan is deleted'}), 400

    if not _get_openai_client():
        return jsonify({'error': 'Statement import is not configured (set OPENAI_API_KEY).'}), 503

    prep, err = _prepare_statement_text_from_upload()
    if err:
        return jsonify({'error': err}), 400
    text, truncated_text = prep

    try:
        raw_txs = _extract_transactions_llm(text)
    except Exception as e:
        logger.exception('LLM extraction failed')
        return jsonify({'error': str(e)}), 500

    candidates = _normalize_candidates(raw_txs)
    _mark_duplicates(candidates, loan)
    out = {'candidates': candidates, 'truncated': truncated_text}
    bl = loan.get('bill_baseline')
    if bl:
        diff = _compare_to_baseline(candidates, bl)
        if diff is not None:
            out['baseline_diff'] = diff
    return jsonify(out)


@app.route('/api/loan/<loan_id>/statement/import', methods=['POST'])
@login_required
def statement_import(loan_id):
    """Bulk-append repayments with historical dates. Body: {\"transactions\":[{\"date\",\"amount\",\"description\"}]} — amount is positive (your share)."""
    data = load_data()
    if loan_id not in data['loans']:
        return jsonify({'error': 'Loan not found'}), 404
    loan = data['loans'][loan_id]
    if loan.get('deleted'):
        return jsonify({'error': 'Loan is deleted'}), 400

    payload = request.get_json(silent=True) or {}
    rows = payload.get('transactions')
    if not isinstance(rows, list) or not rows:
        return jsonify({'error': 'transactions[] is required and must be non-empty'}), 400

    username = session['username']
    earliest = date(1990, 1, 1)
    latest = date.today() + timedelta(days=1)

    validated = []
    for r in rows:
        if not isinstance(r, dict):
            return jsonify({'error': 'Invalid transaction entry'}), 400
        d = _parse_iso_date(str(r.get('date') or ''))
        if d is None:
            return jsonify({'error': f'Invalid date: {r.get("date")!r}'}), 400
        if d < earliest or d > latest:
            return jsonify({'error': f'Date out of range: {d.isoformat()}'}), 400
        try:
            amt = float(r.get('amount'))
        except (TypeError, ValueError):
            return jsonify({'error': 'Invalid amount'}), 400
        if amt <= 0:
            return jsonify({'error': 'Amount must be positive'}), 400
        desc = str(r.get('description') or 'Imported repayment').strip()[:500]
        if not desc:
            desc = 'Imported repayment'
        validated.append((d.strftime('%Y-%m-%d'), round(amt, 2), desc))

    for d_str, amt, desc in validated:
        stored_amt = -amt
        new_transaction = {
            'date': d_str,
            'type': 'repayment',
            'amount': stored_amt,
            'description': desc,
            'user': username,
        }
        loan['transactions'].append(new_transaction)
        loan['loan_amount'] += stored_amt

    save_data(data)
    return jsonify(loan)


@app.route('/api/loan/<loan_id>/statement/from-spending', methods=['POST'])
@login_required
def statement_from_spending(loan_id):
    """Reuse transactions already imported on the home screen as loan-repayment
    candidates. Instead of re-uploading a file, this rebuilds statement text from
    the outgoing lines already extracted for the selected month(s) and re-runs the
    household-bill LLM filter so the review table matches the file-upload flow.
    Body: {"report_month": "YYYY-MM" | "all"}."""
    data = load_data()
    if loan_id not in data['loans']:
        return jsonify({'error': 'Loan not found'}), 404
    loan = data['loans'][loan_id]
    if loan.get('deleted'):
        return jsonify({'error': 'Loan is deleted'}), 400

    if not _get_openai_client():
        return jsonify({'error': 'Statement import is not configured (set OPENAI_API_KEY).'}), 503

    spending, changed = _ensure_user_spending(data, session['username'])
    if changed:
        save_data(data)

    payload = request.get_json(silent=True) or {}
    month_raw = str(payload.get('report_month') or '').strip()
    outgoing = [
        t for t in (spending.get('transactions') or [])
        if t.get('direction') == 'outgoing'
    ]
    if month_raw and month_raw.lower() != 'all':
        month = _normalize_spending_month_key(month_raw)
        if not month:
            return jsonify({'error': 'Invalid report_month (use YYYY-MM).'}), 400
        outgoing = [t for t in outgoing if _report_month_for_spending_tx(t) == month]

    lines = []
    for t in sorted(outgoing, key=lambda x: str(x.get('date') or '')):
        try:
            amt = abs(float(t.get('amount') or 0))
        except (TypeError, ValueError):
            continue
        if amt <= 0:
            continue
        d = str(t.get('date') or '')
        desc = str(t.get('description') or '').strip()
        lines.append(f'{d}  {desc}  -\u00a3{amt:.2f}')

    if not lines:
        return jsonify({
            'candidates': [],
            'truncated': False,
            'source_count': 0,
            'message': 'No imported outgoing transactions found for the selected period. '
                       'Import a statement on the home screen first.',
        })

    pseudo_text = '\n'.join(lines)

    try:
        raw_txs = _extract_transactions_llm(pseudo_text)
    except Exception as e:
        logger.exception('LLM extraction from imported spending failed')
        return jsonify({'error': str(e)}), 500

    candidates = _normalize_candidates(raw_txs)
    _mark_duplicates(candidates, loan)
    out = {
        'candidates': candidates,
        'truncated': len(pseudo_text) > STATEMENT_MAX_CHARS_FOR_LLM,
        'source_count': len(lines),
    }
    bl = loan.get('bill_baseline')
    if bl:
        diff = _compare_to_baseline(candidates, bl)
        if diff is not None:
            out['baseline_diff'] = diff
    return jsonify(out)


@app.route('/api/loan/<loan_id>/statement/baseline-preview', methods=['POST'])
@login_required
def statement_baseline_preview(loan_id):
    data = load_data()
    if loan_id not in data['loans']:
        return jsonify({'error': 'Loan not found'}), 404
    loan = data['loans'][loan_id]
    if loan.get('deleted'):
        return jsonify({'error': 'Loan is deleted'}), 400

    if not _get_openai_client():
        return jsonify({'error': 'Statement import is not configured (set OPENAI_API_KEY).'}), 503

    prep, err = _prepare_statement_text_from_upload()
    if err:
        return jsonify({'error': err}), 400
    text, truncated_text = prep

    try:
        raw_txs, period_end = _extract_baseline_statement_llm(text)
    except Exception as e:
        logger.exception('baseline LLM extraction failed')
        return jsonify({'error': str(e)}), 500

    candidates = _normalize_candidates(raw_txs)
    inferred_day = _default_reminder_day_from_period_end(period_end)
    return jsonify({
        'candidates': candidates,
        'truncated': truncated_text,
        'inferred_period_end': period_end,
        'inferred_reminder_day': inferred_day,
    })


@app.route('/api/loan/<loan_id>/statement/baseline-save', methods=['POST'])
@login_required
def statement_baseline_save(loan_id):
    data = load_data()
    if loan_id not in data['loans']:
        return jsonify({'error': 'Loan not found'}), 404
    loan = data['loans'][loan_id]
    if loan.get('deleted'):
        return jsonify({'error': 'Loan is deleted'}), 400

    payload = request.get_json(silent=True) or {}
    items = payload.get('items')
    if not isinstance(items, list) or not items:
        return jsonify({'error': 'items[] is required and must be non-empty'}), 400

    day = int(payload.get('day_of_month', 1))
    if not 1 <= day <= 31:
        return jsonify({'error': 'day_of_month must be 1–31'}), 400

    reminder_email = str(payload.get('reminder_email') or '').strip()
    inferred_period_end = payload.get('inferred_period_end')
    if inferred_period_end is not None and inferred_period_end != '':
        d = _parse_iso_date(str(inferred_period_end))
        inferred_period_end = d.strftime('%Y-%m-%d') if d else None
    else:
        inferred_period_end = None

    baseline = []
    for row in items:
        p = _parse_baseline_item_row(row)
        if p:
            baseline.append(p)

    if not baseline:
        return jsonify({'error': 'No valid baseline items'}), 400

    loan['bill_baseline'] = baseline
    loan['bill_reminder_email'] = reminder_email or None
    loan['bill_reminder'] = {
        'day_of_month': day,
        'last_sent_month': (loan.get('bill_reminder') or {}).get('last_sent_month'),
        'inferred_period_end': inferred_period_end,
    }
    save_data(data)
    schedule_bill_reminders_job()
    return jsonify({
        'ok': True,
        'bill_baseline': baseline,
        'bill_reminder': loan['bill_reminder'],
        'bill_reminder_email': loan.get('bill_reminder_email'),
    })


@app.route('/api/loan/<loan_id>/bill-baseline', methods=['GET'])
@login_required
def get_bill_baseline(loan_id):
    data = load_data()
    if loan_id not in data['loans']:
        return jsonify({'error': 'Loan not found'}), 404
    loan = data['loans'][loan_id]
    return jsonify({
        'bill_baseline': loan.get('bill_baseline') or [],
        'bill_reminder': loan.get('bill_reminder'),
        'bill_reminder_email': loan.get('bill_reminder_email'),
        'smtp_configured': _get_smtp_config() is not None,
        'default_notify_email': os.getenv('FINANCE_TRACKER_NOTIFY_EMAIL', '').strip(),
    })


@app.route('/api/loan/<loan_id>/bill-baseline', methods=['PATCH'])
@login_required
def patch_bill_baseline_reminder(loan_id):
    """Update reminder email (and optionally reminder day) without replacing baseline rows."""
    data = load_data()
    if loan_id not in data['loans']:
        return jsonify({'error': 'Loan not found'}), 404
    loan = data['loans'][loan_id]
    if loan.get('deleted'):
        return jsonify({'error': 'Loan is deleted'}), 400

    payload = request.get_json(silent=True) or {}
    if not payload:
        return jsonify({'error': 'JSON body required'}), 400

    if 'reminder_email' in payload:
        re = str(payload.get('reminder_email') or '').strip()
        loan['bill_reminder_email'] = re or None
    if 'day_of_month' in payload:
        day = int(payload.get('day_of_month', 1))
        if not 1 <= day <= 31:
            return jsonify({'error': 'day_of_month must be 1–31'}), 400
        br = dict(loan.get('bill_reminder') or {})
        br['day_of_month'] = day
        loan['bill_reminder'] = br

    if 'reminder_email' not in payload and 'day_of_month' not in payload:
        return jsonify({'error': 'Provide reminder_email and/or day_of_month'}), 400

    save_data(data)
    schedule_bill_reminders_job()
    return jsonify({
        'ok': True,
        'bill_reminder_email': loan.get('bill_reminder_email'),
        'bill_reminder': loan.get('bill_reminder'),
    })


@app.route('/api/loan/<loan_id>/bill-baseline', methods=['PUT'])
@login_required
def put_bill_baseline(loan_id):
    """Replace all baseline items. Optional: day_of_month, reminder_email, inferred_period_end."""
    data = load_data()
    if loan_id not in data['loans']:
        return jsonify({'error': 'Loan not found'}), 404
    loan = data['loans'][loan_id]
    if loan.get('deleted'):
        return jsonify({'error': 'Loan is deleted'}), 400

    payload = request.get_json(silent=True) or {}
    items = payload.get('items')
    if not isinstance(items, list):
        return jsonify({'error': 'items must be an array'}), 400

    baseline = []
    for i, row in enumerate(items):
        p = _parse_baseline_item_row(row)
        if p is None:
            return jsonify({'error': f'Invalid baseline item at index {i}'}), 400
        baseline.append(p)

    loan['bill_baseline'] = baseline

    if 'day_of_month' in payload:
        day = int(payload.get('day_of_month', 1))
        if not 1 <= day <= 31:
            return jsonify({'error': 'day_of_month must be 1–31'}), 400
        br = dict(loan.get('bill_reminder') or {})
        br['day_of_month'] = day
        loan['bill_reminder'] = br
    if 'reminder_email' in payload:
        re = str(payload.get('reminder_email') or '').strip()
        loan['bill_reminder_email'] = re or None
    if 'inferred_period_end' in payload:
        ipe = payload.get('inferred_period_end')
        if ipe is not None and ipe != '':
            d = _parse_iso_date(str(ipe))
            ipe = d.strftime('%Y-%m-%d') if d else None
        else:
            ipe = None
        br = dict(loan.get('bill_reminder') or {})
        br['inferred_period_end'] = ipe
        loan['bill_reminder'] = br

    save_data(data)
    schedule_bill_reminders_job()
    return jsonify({
        'ok': True,
        'bill_baseline': baseline,
        'bill_reminder': loan.get('bill_reminder'),
        'bill_reminder_email': loan.get('bill_reminder_email'),
    })


@app.route('/api/loan/<loan_id>/bill-baseline/item/<item_id>', methods=['DELETE'])
@login_required
def delete_bill_baseline_item(loan_id, item_id):
    data = load_data()
    if loan_id not in data['loans']:
        return jsonify({'error': 'Loan not found'}), 404
    loan = data['loans'][loan_id]
    if loan.get('deleted'):
        return jsonify({'error': 'Loan is deleted'}), 400

    bl = loan.get('bill_baseline') or []
    new_bl = [b for b in bl if str(b.get('id')) != str(item_id)]
    if len(new_bl) == len(bl):
        return jsonify({'error': 'Item not found'}), 404

    loan['bill_baseline'] = new_bl
    save_data(data)
    return jsonify({'ok': True, 'bill_baseline': new_bl})


@app.route('/api/loan/<loan_id>/bill-baseline/item/<item_id>', methods=['PATCH'])
@login_required
def patch_bill_baseline_item(loan_id, item_id):
    data = load_data()
    if loan_id not in data['loans']:
        return jsonify({'error': 'Loan not found'}), 404
    loan = data['loans'][loan_id]
    if loan.get('deleted'):
        return jsonify({'error': 'Loan is deleted'}), 400

    bl = loan.get('bill_baseline') or []
    idx = next((i for i, b in enumerate(bl) if str(b.get('id')) == str(item_id)), None)
    if idx is None:
        return jsonify({'error': 'Item not found'}), 404

    payload = request.get_json(silent=True) or {}
    cur = dict(bl[idx])
    if 'description' in payload:
        d = str(payload.get('description') or '').strip()[:500]
        cur['description'] = d if d else 'Bill'
    if 'amount_bank' in payload:
        try:
            ab = float(payload.get('amount_bank'))
        except (TypeError, ValueError):
            return jsonify({'error': 'Invalid amount_bank'}), 400
        if ab <= 0:
            return jsonify({'error': 'amount_bank must be positive'}), 400
        cur['amount_bank'] = round(ab, 2)
    if 'amount_share' in payload:
        try:
            sh = float(payload.get('amount_share'))
        except (TypeError, ValueError):
            return jsonify({'error': 'Invalid amount_share'}), 400
        if sh <= 0:
            return jsonify({'error': 'amount_share must be positive'}), 400
        cur['amount_share'] = round(sh, 2)
    if 'category' in payload:
        c = payload.get('category')
        cur['category'] = str(c)[:120] if c is not None and str(c).strip() else None
    if 'note' in payload:
        cur['note'] = _sanitize_note(payload.get('note'))

    merged = _parse_baseline_item_row(cur)
    if merged is None:
        return jsonify({'error': 'Invalid item after update'}), 400
    merged['id'] = str(cur.get('id'))

    bl[idx] = merged
    loan['bill_baseline'] = bl
    save_data(data)
    return jsonify({'ok': True, 'item': merged, 'bill_baseline': bl})


@app.route('/api/loan/<loan_id>/bill-baseline/merge-candidates', methods=['POST'])
@login_required
def merge_baseline_candidates(loan_id):
    """POST JSON {\"candidates\": [...]} — same shape as normalized statement candidates. Returns merged baseline-shaped rows."""
    data = load_data()
    if loan_id not in data['loans']:
        return jsonify({'error': 'Loan not found'}), 404
    loan = data['loans'][loan_id]
    if loan.get('deleted'):
        return jsonify({'error': 'Loan is deleted'}), 400

    payload = request.get_json(silent=True) or {}
    candidates = payload.get('candidates')
    if not isinstance(candidates, list):
        return jsonify({'error': 'candidates must be an array'}), 400

    existing = loan.get('bill_baseline') or []
    items = merge_baseline_with_candidates(existing, candidates)
    return jsonify({'items': items})


@app.route('/api/spending/statement/preview', methods=['POST'])
@login_required
def spending_statement_preview():
    prep, err = _prepare_statement_text_from_upload(for_spending=True)
    if err:
        return jsonify({'error': err}), 400
    text, truncated_text, direction_hints, pipeline = prep

    # Alias-mapped CSV/XLSX can proceed without OpenAI; unfamiliar headers need a
    # tiny header LLM call (or full extract), so require the API key then.
    if (
        _try_parse_tabular_spending_transactions(text, allow_llm=False) is None
        and not _get_openai_client()
    ):
        return jsonify({'error': 'Statement analysis is not configured (set OPENAI_API_KEY).'}), 503

    period, perr = _parse_spending_period_from_values(
        request.form.get('report_month'),
        request.form.get('period_start'),
        request.form.get('period_end'),
    )
    if perr:
        return jsonify({'error': perr}), 400

    try:
        body = _spending_statement_preview_payload(
            text, truncated_text, direction_hints, pipeline, period
        )
    except RuntimeError as e:
        if 'OPENAI_API_KEY' in str(e):
            return jsonify({'error': 'Statement analysis is not configured (set OPENAI_API_KEY).'}), 503
        logger.exception('spending statement extraction failed')
        return jsonify({'error': str(e)}), 500
    except Exception as e:
        logger.exception('spending statement extraction failed')
        return jsonify({'error': str(e)}), 500
    return jsonify(body)


@app.route('/api/spending/statement/preview-stream', methods=['POST'])
@login_required
def spending_statement_preview_stream():
    """NDJSON stream of progress events; ends with {\"type\":\"complete\",\"payload\":{...}}."""
    period, perr = _parse_spending_period_from_values(
        request.form.get('report_month'),
        request.form.get('period_start'),
        request.form.get('period_end'),
    )
    if perr:

        def err400():
            yield json.dumps({'type': 'error', 'message': perr, 'http_status': 400, 'elapsed_ms': 0}) + '\n'

        return Response(stream_with_context(err400()), mimetype='application/x-ndjson', status=400)

    f = request.files.get('file')
    if not f or not getattr(f, 'filename', None):

        def err_nf():
            yield json.dumps({'type': 'error', 'message': 'No file uploaded', 'http_status': 400, 'elapsed_ms': 0}) + '\n'

        return Response(stream_with_context(err_nf()), mimetype='application/x-ndjson', status=400)

    name = (f.filename or '').lower()
    # Buffer the full upload before streaming: Werkzeug closes the input stream once the
    # response body is iterated, so f.read() inside the generator raises "I/O on closed file".
    raw = f.read()
    if len(raw) > STATEMENT_MAX_BYTES:

        def err_sz():
            yield json.dumps({
                'type': 'error',
                'message': 'File too large (max 2 MB).',
                'http_status': 400,
                'elapsed_ms': 0,
            }) + '\n'

        return Response(stream_with_context(err_sz()), mimetype='application/x-ndjson', status=400)

    def ndjson_gen():
        t0 = time.perf_counter()

        def elapsed_ms() -> int:
            return int((time.perf_counter() - t0) * 1000)

        def emit(obj: dict) -> str:
            d = dict(obj)
            d.setdefault('elapsed_ms', elapsed_ms())
            return json.dumps(d, default=str) + '\n'

        yield emit({
            'type': 'progress',
            'step': 'file_loaded',
            'message': f'Received file ({len(raw):,} bytes). Parsing statement…',
        })

        prep = None
        for event in _iter_prepare_spending_raw(name, raw):
            et = event.get('type')
            if et == 'progress':
                yield emit(event)
            elif et == 'error':
                yield emit(event)
                return
            elif et == 'prep_complete':
                prep = event.get('prep')
            else:
                logger.warning('unknown prep event: %s', event)

        if not prep:
            yield emit({'type': 'error', 'message': 'Statement preparation failed.', 'http_status': 500})
            return

        text, truncated_text, direction_hints, pipeline = prep

        if (
            _try_parse_tabular_spending_transactions(text, allow_llm=False) is None
            and not _get_openai_client()
        ):
            yield emit({
                'type': 'error',
                'message': 'Statement analysis is not configured (set OPENAI_API_KEY).',
                'http_status': 503,
            })
            return

        period_hint = (
            f'Prefer transactions whose completed/settled date (or sole statement date) falls on or between '
            f'{period["period_start"]} and {period["period_end"]} inclusive. '
            f'When a row has both started and completed dates, always set date=completed_date even if started_date '
            f'is outside that range — still include the row. When in doubt, include the row.'
        )

        yield emit({'type': 'progress', 'step': 'load_profile', 'message': 'Loading your spending profile…'})
        data = load_data()
        spending, changed = _ensure_user_spending(data, session['username'])
        if changed:
            save_data(data)

        raw_rows = None
        extraction_meta = None
        try:
            for ev in iter_spending_transaction_extraction(text, period_hint):
                et = ev.get('type')
                if et == 'progress':
                    yield emit(ev)
                elif et == 'result':
                    raw_rows = ev.get('rows')
                    extraction_meta = ev.get('meta')
        except RuntimeError as e:
            if 'OPENAI_API_KEY' in str(e):
                yield emit({
                    'type': 'error',
                    'message': 'Statement analysis is not configured (set OPENAI_API_KEY).',
                    'http_status': 503,
                })
                return
            logger.exception('spending statement stream extraction failed')
            yield emit({'type': 'error', 'message': str(e), 'http_status': 500})
            return
        except Exception as e:
            logger.exception('spending statement stream extraction failed')
            yield emit({'type': 'error', 'message': str(e), 'http_status': 500})
            return

        if raw_rows is None:
            raw_rows = []

        yield emit({
            'type': 'progress',
            'step': 'postprocess',
            'message': 'Normalising rows, filtering dates, reconciling directions, applying categories…',
        })
        try:
            body = _spending_statement_preview_finalize(
                data,
                spending,
                period,
                direction_hints,
                pipeline,
                truncated_text,
                raw_rows,
                extraction_meta,
            )
        except Exception as e:
            logger.exception('spending statement postprocess failed')
            yield emit({'type': 'error', 'message': str(e), 'http_status': 500})
            return

        yield emit({'type': 'progress', 'step': 'done', 'message': 'Preview ready.'})
        yield emit({'type': 'complete', 'payload': body})

    return Response(
        stream_with_context(ndjson_gen()),
        mimetype='application/x-ndjson',
        headers={'Cache-Control': 'no-store', 'X-Accel-Buffering': 'no'},
    )


@app.route('/api/spending/statement/import', methods=['POST'])
@login_required
def spending_statement_import():
    payload = request.get_json(silent=True) or {}
    rows = payload.get('transactions')
    if not isinstance(rows, list) or not rows:
        return jsonify({'error': 'transactions[] is required and must be non-empty'}), 400

    period, perr = _parse_spending_period_from_values(
        payload.get('report_month'),
        payload.get('period_start'),
        payload.get('period_end'),
    )
    if perr:
        return jsonify({'error': perr}), 400
    report_month = period['report_month']
    ps_d = period['period_start_date']
    pe_d = period['period_end_date']
    # Optional bank/account the statement came from (UI: "Source").
    # Accept top-level `source` or `bank_source`; do not confuse with tx.origin `source`.
    bank_source = _normalize_bank_source(payload.get('bank_source') or payload.get('source'))

    data = load_data()
    spending, changed = _ensure_user_spending(data, session['username'])
    if changed:
        save_data(data)

    validated = []
    for row in rows:
        if not isinstance(row, dict):
            return jsonify({'error': 'Invalid transaction entry'}), 400
        d = _parse_bank_statement_date(str(row.get('date') or ''))
        if d is None:
            return jsonify({'error': f'Invalid date: {row.get("date")!r}'}), 400
        if d < ps_d or d > pe_d:
            return jsonify({'error': f'Transaction date {d.isoformat()} is outside the reporting period.'}), 400
        desc = str(row.get('description') or '').strip()[:500] or 'Bank transaction'
        col_dir, col_amt = _direction_and_amount_from_money_columns(row)
        if col_dir is not None and col_amt is not None:
            amount = round(col_amt, 2)
            direction = col_dir
        else:
            try:
                amount = float(row.get('amount'))
            except (TypeError, ValueError):
                return jsonify({'error': 'Invalid amount'}), 400
            if amount <= 0:
                return jsonify({'error': 'Amount must be positive'}), 400
            amount = round(abs(amount), 2)
            direction = _normalize_spending_direction(row.get('direction'), desc, amount)
        category = row.get('category')
        if direction == 'outgoing':
            category = str(category or 'unclassified').strip().lower()
            if category not in SPENDING_CATEGORY_SET:
                category = 'unclassified'
        else:
            category = None
        rationale = str(row.get('rationale') or '').strip()[:240]
        confidence = row.get('confidence')
        if confidence is not None:
            try:
                confidence = max(0.0, min(1.0, float(confidence)))
            except (TypeError, ValueError):
                confidence = None
        validated.append({
            'date': d.strftime('%Y-%m-%d'),
            'month': d.strftime('%Y-%m'),
            'report_month': report_month,
            'amount': round(amount, 2),
            'description': desc,
            'direction': direction,
            'category': category,
            'confidence': confidence,
            'rationale': rationale,
        })

    tx_store = spending.setdefault('transactions', [])
    statement_store = spending.setdefault('statements', [])
    existing_fingerprints = {
        str(t.get('fingerprint'))
        for t in tx_store
        if t.get('fingerprint')
    }
    statement_id = str(uuid.uuid4())
    now_iso = datetime.utcnow().isoformat() + 'Z'
    inserted = 0
    skipped = 0
    affected_months = set()
    claimed_manual_ids: set[str] = set()

    for row in validated:
        fp = _spending_fingerprint(
            report_month, row['date'], row['amount'], row['direction'], row['description']
        )
        if fp in existing_fingerprints:
            skipped += 1
            continue
        manual_match = _daily_budget_fuzzy_match_manual(
            spending,
            date_str=row['date'],
            amount=row['amount'],
            description=row['description'],
            direction=row['direction'],
            exclude_ids=claimed_manual_ids,
        )
        if manual_match is not None:
            _daily_budget_claim_manual_match(manual_match, row, statement_id, fp)
            mid = str(manual_match.get('id') or '')
            if mid:
                claimed_manual_ids.add(mid)
            existing_fingerprints.add(fp)
            skipped += 1
            affected_months.add(report_month)
            continue
        existing_fingerprints.add(fp)
        tx = {
            'id': str(uuid.uuid4()),
            'date': row['date'],
            'month': row['month'],
            'report_month': report_month,
            'description': row['description'],
            'amount': row['amount'],
            'direction': row['direction'],
            'category': row['category'],
            'confidence': row['confidence'],
            'rationale': row['rationale'],
            'source_statement_id': statement_id,
            'source': 'statement',
            'created_at': now_iso,
            'fingerprint': fp,
        }
        if bank_source:
            tx['bank_source'] = bank_source
        tx_store.append(tx)
        inserted += 1
        affected_months.add(report_month)
        next_month = _month_next(report_month)
        if next_month:
            affected_months.add(next_month)

    statement_record = {
        'id': statement_id,
        'uploaded_at': now_iso,
        'file_name': str(payload.get('file_name') or '').strip()[:200],
        'report_month': report_month,
        'period_start': period['period_start'],
        'period_end': period['period_end'],
        'original_row_count': len(validated),
        'imported_count': inserted,
        'skipped_duplicates': skipped,
        'months': sorted({report_month}),
    }
    if bank_source:
        statement_record['bank_source'] = bank_source
    statement_store.append(statement_record)
    reconciliation = apply_auto_transfer_pairing_for_month(spending, report_month)
    _recompute_monthly_insights(spending, affected_months if affected_months else None)
    save_data(data)

    return jsonify({
        'ok': True,
        'statement_id': statement_id,
        'imported_count': inserted,
        'skipped_duplicates': skipped,
        'report_month': report_month,
        'period_start': period['period_start'],
        'period_end': period['period_end'],
        'months': sorted({report_month}),
        'reconciliation': reconciliation,
        'statement': statement_record,
        'known_bank_sources': _collect_bank_sources(spending),
    })


def _normalize_spending_month_key(raw: str | None) -> str | None:
    s = (raw or '').strip()[:7]
    if len(s) != 7 or s[4] != '-':
        return None
    try:
        datetime.strptime(s + '-01', '%Y-%m-%d')
    except ValueError:
        return None
    return s


@app.route('/api/spending/months', methods=['GET'])
@login_required
def spending_months():
    data = load_data()
    spending, changed = _ensure_user_spending(data, session['username'])
    if changed:
        save_data(data)
    months = sorted((spending.get('monthly_insights') or {}).keys(), reverse=True)
    return jsonify({'months': months})


@app.route('/api/spending/transactions/search', methods=['GET'])
@login_required
def spending_transactions_search():
    """
    Search the user's full spending ledger (all months).
    Query params: q, date_from, date_to, category, direction, source, min_amount, max_amount,
    limit (default 100, max 500), offset (default 0).
    Requires at least one of q / date_from / date_to / category / direction / source /
    min_amount / max_amount — otherwise returns an empty result set (prompt to search).
    """
    data = load_data()
    spending, changed = _ensure_user_spending(data, session['username'])
    if changed:
        save_data(data)

    q = str(request.args.get('q') or '').strip()
    date_from = str(request.args.get('date_from') or request.args.get('from') or '').strip()[:10]
    date_to = str(request.args.get('date_to') or request.args.get('to') or '').strip()[:10]
    category = str(request.args.get('category') or '').strip().lower()
    direction = str(request.args.get('direction') or '').strip().lower()
    bank_source = str(request.args.get('source') or request.args.get('bank_source') or '').strip()
    min_amount = _parse_optional_amount(request.args.get('min_amount'))
    max_amount = _parse_optional_amount(request.args.get('max_amount'))

    try:
        limit = int(request.args.get('limit') or 100)
    except (TypeError, ValueError):
        limit = 100
    limit = max(1, min(limit, 500))
    try:
        offset = int(request.args.get('offset') or 0)
    except (TypeError, ValueError):
        offset = 0
    offset = max(0, offset)

    has_criterion = bool(
        q
        or date_from
        or date_to
        or category
        or direction in ('incoming', 'outgoing')
        or bank_source
        or min_amount is not None
        or max_amount is not None
    )
    filters = {
        'q': q,
        'date_from': date_from or None,
        'date_to': date_to or None,
        'category': category or None,
        'direction': direction if direction in ('incoming', 'outgoing') else None,
        'source': bank_source or None,
        'min_amount': min_amount,
        'max_amount': max_amount,
    }
    if not has_criterion:
        return jsonify({
            'transactions': [],
            'total': 0,
            'limit': limit,
            'offset': offset,
            'searched': False,
            'filters': filters,
            'known_bank_sources': _collect_bank_sources(spending),
        })

    matched = _search_spending_transactions(
        spending.get('transactions') or [],
        q=q,
        date_from=date_from or None,
        date_to=date_to or None,
        category=category or None,
        direction=direction or None,
        bank_source=bank_source or None,
        min_amount=min_amount,
        max_amount=max_amount,
    )
    total = len(matched)
    page = matched[offset : offset + limit]
    return jsonify({
        'transactions': page,
        'total': total,
        'limit': limit,
        'offset': offset,
        'searched': True,
        'filters': filters,
        'known_bank_sources': _collect_bank_sources(spending),
    })


@app.route('/api/spending/metrics-trend', methods=['GET'])
@login_required
def spending_metrics_trend():
    """Chronological series of income, outgoing, and net from stored monthly insights (for charts)."""
    data = load_data()
    spending, changed = _ensure_user_spending(data, session['username'])
    if changed:
        save_data(data)
    insight_map = spending.get('monthly_insights') or {}
    months = sorted(insight_map.keys())
    points = []
    for m in months:
        ins = insight_map.get(m)
        if not isinstance(ins, dict):
            continue
        try:
            income = float(ins.get('income_total', 0) or 0)
            out = float(ins.get('outgoing_total', 0) or 0)
            net = float(ins.get('net', 0) or 0)
        except (TypeError, ValueError):
            continue
        outgoing_by_category: dict[str, float] = {}
        for row in ins.get('category_breakdown') or []:
            if not isinstance(row, dict):
                continue
            c = str(row.get('category', 'unclassified') or 'unclassified')
            try:
                raw_amt = float(row.get('amount', 0) or 0)
            except (TypeError, ValueError):
                continue
            amt = round(raw_amt, 2)
            if amt <= 0:
                continue
            outgoing_by_category[c] = amt
        points.append({
            'month': m,
            'income': round(income, 2),
            'outgoing': round(out, 2),
            'net': round(net, 2),
            'outgoing_by_category': outgoing_by_category,
        })
    return jsonify({'points': points})


@app.route('/api/spending/month/<month>', methods=['DELETE'])
@login_required
def spending_delete_month(month):
    mk = _normalize_spending_month_key(month)
    if not mk:
        return jsonify({'error': 'Invalid month (use YYYY-MM)'}), 400

    data = load_data()
    spending, changed = _ensure_user_spending(data, session['username'])
    if changed:
        save_data(data)

    txs = spending.setdefault('transactions', [])
    before = len(txs)
    txs[:] = [t for t in txs if _report_month_for_spending_tx(t) != mk]
    removed = before - len(txs)

    statements = spending.setdefault('statements', [])
    spending['statements'] = [
        s for s in statements
        if isinstance(s, dict) and str(s.get('report_month') or '').strip()[:7] != mk
    ]

    _recompute_monthly_insights(spending, None)
    save_data(data)

    months = sorted((spending.get('monthly_insights') or {}).keys(), reverse=True)
    return jsonify({
        'ok': True,
        'month': mk,
        'removed_transactions': removed,
        'months': months,
    })


@app.route('/api/spending/insights', methods=['GET'])
@login_required
def spending_insights():
    data = load_data()
    spending, changed = _ensure_user_spending(data, session['username'])
    if changed:
        save_data(data)
    requested_month = (request.args.get('month') or '').strip()
    months = sorted((spending.get('monthly_insights') or {}).keys(), reverse=True)
    if not months:
        return jsonify({
            'month': None,
            'insight': None,
            'available_months': [],
            'known_bank_sources': _collect_bank_sources(spending),
        })
    target_month = requested_month or months[0]
    insight_map = spending.get('monthly_insights') or {}
    need_recompute = target_month not in insight_map
    if not need_recompute:
        cur = insight_map.get(target_month)
        if isinstance(cur, dict) and 'largest_outgoing' not in cur:
            need_recompute = True
        if isinstance(cur, dict) and 'budget_action_items' not in cur:
            need_recompute = True
    if need_recompute:
        _recompute_monthly_insights(spending, {target_month})
        save_data(data)
    insight = (spending.get('monthly_insights') or {}).get(target_month)
    month_rows = [t for t in (spending.get('transactions') or []) if _report_month_for_spending_tx(t) == target_month]
    month_rows.sort(key=lambda r: (r.get('date', ''), r.get('description', '')))
    return jsonify({
        'month': target_month,
        'insight': insight,
        'available_months': sorted((spending.get('monthly_insights') or {}).keys(), reverse=True),
        'transactions': month_rows,
        'known_bank_sources': _collect_bank_sources(spending),
    })


@app.route('/api/spending/insights/recompute', methods=['POST'])
@login_required
def spending_insights_recompute():
    """
    Rebuild all stored monthly insight blobs from current transactions.
    Use after app changes or to refresh KPIs, trends, and anomalies.
    """
    data = load_data()
    spending, changed = _ensure_user_spending(data, session['username'])
    if changed:
        save_data(data)
    _recompute_monthly_insights(spending, None)
    save_data(data)
    months = sorted((spending.get('monthly_insights') or {}).keys(), reverse=True)
    return jsonify({'ok': True, 'months': months, 'recomputed': len(months)})


@app.route('/api/spending/savings-advice', methods=['POST'])
@login_required
def spending_savings_advice():
    """
    Data-grounded savings suggestions via LLM. Sends focal insight, capped per-line transactions for
    the focal month (and a small per-line sample for the two most recent prior months), plus trend_series.
    Requires OPENAI_API_KEY. Third party receives this JSON.
    """
    payload = request.get_json(silent=True) or {}
    data = load_data()
    spending, changed = _ensure_user_spending(data, session['username'])
    if changed:
        save_data(data)
    available = sorted((spending.get('monthly_insights') or {}).keys(), reverse=True)
    if not available:
        return jsonify({'error': 'No spending data'}), 404
    month = str(payload.get('month') or '').strip() or available[0]
    if month not in available:
        return jsonify({'error': 'Unknown month or no data for that month.'}), 400
    insight_map = spending.get('monthly_insights') or {}
    need = (month not in insight_map) or (not isinstance(insight_map.get(month), dict)) or (
        isinstance(insight_map.get(month), dict) and 'budget_action_items' not in (insight_map.get(month) or {})
    )
    if need:
        _recompute_monthly_insights(spending, {month})
        save_data(data)
    insight_map = spending.get('monthly_insights') or {}
    if not isinstance(insight_map.get(month), dict):
        return jsonify({'error': 'No insight for that month.'}), 400
    if not _get_openai_client():
        return jsonify({
            'error': 'Savings ideas require OPENAI_API_KEY to be set on the server.',
        }), 503
    context, n_trend, berr, tx_meta = _build_savings_advice_context(spending, month)
    if berr or not context:
        return jsonify({'error': 'Could not build savings context.'}), 400
    advice, llm_err, model_used = _generate_savings_advice_with_llm(context)
    if not advice or llm_err:
        return jsonify({
            'error': llm_err or 'Could not generate savings ideas. Try again.',
        }), 500
    out: dict = {
        'month': month,
        'advice': advice,
        'model': model_used,
        'trend_months_included': n_trend,
    }
    if tx_meta and isinstance(tx_meta, dict):
        ft = tx_meta.get('focal_transactions') or {}
        if isinstance(ft, dict):
            out['focal_transaction_count_in_month'] = ft.get('count_in_month', 0)
            out['focal_transaction_count_sent'] = ft.get('count_sent', 0)
            out['focal_transactions_truncated'] = bool(ft.get('truncated'))
        out['prior_transaction_sample_months'] = tx_meta.get('prior_transaction_months_included', 0)
    return jsonify(out)


@app.route('/api/spending/transaction/<tx_id>/category', methods=['PATCH'])
@login_required
def spending_reclassify(tx_id):
    payload = request.get_json(silent=True) or {}
    category = str(payload.get('category') or '').strip().lower()
    if category not in SPENDING_CATEGORY_SET:
        return jsonify({'error': 'Invalid category'}), 400

    data = load_data()
    spending, changed = _ensure_user_spending(data, session['username'])
    if changed:
        save_data(data)

    txs = spending.setdefault('transactions', [])
    target = next((t for t in txs if str(t.get('id')) == str(tx_id)), None)
    if not target:
        return jsonify({'error': 'Transaction not found'}), 404
    if target.get('direction') != 'outgoing':
        return jsonify({'error': 'Only outgoing transactions can be reclassified'}), 400

    target['category'] = category
    target['confidence'] = 1.0
    target['rationale'] = 'manual override'
    label = _normalize_label(target.get('description', ''))
    if label:
        spending.setdefault('classification_overrides', {})[label] = category

    bucket = target.get('report_month') or target.get('month')
    months_to_recompute = {bucket}
    nxt = _month_next(str(bucket or ''))
    if nxt:
        months_to_recompute.add(nxt)
    _recompute_monthly_insights(spending, {m for m in months_to_recompute if m})
    save_data(data)

    return jsonify({
        'ok': True,
        'transaction': target,
        'insight': (spending.get('monthly_insights') or {}).get(bucket),
    })


@app.route('/api/spending/transaction/<tx_id>/insights', methods=['PATCH'])
@login_required
def spending_transaction_insights_exclude(tx_id):
    """Set ``insights_excluded`` for a non-paired transaction (same effect as linked rows on KPIs)."""
    payload = request.get_json(silent=True) or {}
    raw = payload.get('excluded')
    if raw is None:
        raw = payload.get('insights_excluded')
    if raw is None:
        return jsonify({'error': 'excluded (boolean) is required'}), 400
    if not isinstance(raw, bool):
        return jsonify({'error': 'excluded must be a boolean'}), 400
    excluded = raw

    data = load_data()
    spending, changed = _ensure_user_spending(data, session['username'])
    if changed:
        save_data(data)

    txs = spending.setdefault('transactions', [])
    target = next((t for t in txs if str(t.get('id')) == str(tx_id)), None)
    if not target:
        return jsonify({'error': 'Transaction not found'}), 404
    if _spending_is_paired_leg(target):
        return jsonify({
            'error': 'Linked internal transfers are already excluded; use Unlink if you need them in totals.',
        }), 400

    target['insights_excluded'] = excluded
    bucket = str(_report_month_for_spending_tx(target) or '')
    months_to_recompute = {bucket}
    nxt = _month_next(bucket)
    if nxt:
        months_to_recompute.add(nxt)
    _recompute_monthly_insights(spending, {m for m in months_to_recompute if m})
    save_data(data)

    return jsonify({
        'ok': True,
        'transaction': target,
        'insight': (spending.get('monthly_insights') or {}).get(bucket),
    })


@app.route('/api/spending/transaction/<tx_id>/pair', methods=['PATCH'])
@login_required
def spending_pair_transaction(tx_id):
    payload = request.get_json(silent=True) or {}
    if payload.get('unlink') is True:
        data = load_data()
        spending, changed = _ensure_user_spending(data, session['username'])
        if changed:
            save_data(data)
        txs = spending.setdefault('transactions', [])
        target = next((t for t in txs if str(t.get('id')) == str(tx_id)), None)
        if not target:
            return jsonify({'error': 'Transaction not found'}), 404
        if not target.get('transfer_pair_id'):
            return jsonify({'error': 'Transaction is not paired'}), 400
        if not _unlink_spending_pair(txs, target):
            return jsonify({'error': 'Failed to unlink pair'}), 400
        bucket = str(_report_month_for_spending_tx(target) or '')
        _recompute_monthly_insights(spending, {bucket} if bucket else None)
        save_data(data)
        return jsonify({
            'ok': True,
            'insight': (spending.get('monthly_insights') or {}).get(bucket),
            'reconciliation': _spending_transfer_reconciliation_for_month(
                [t for t in txs if _report_month_for_spending_tx(t) == bucket]
            ),
        })

    peer_id = str(payload.get('peer_id') or '').strip()
    if not peer_id:
        return jsonify({'error': 'peer_id is required, or set unlink: true'}), 400

    data = load_data()
    spending, changed = _ensure_user_spending(data, session['username'])
    if changed:
        save_data(data)
    txs = spending.setdefault('transactions', [])
    a = next((t for t in txs if str(t.get('id')) == str(tx_id)), None)
    b = next((t for t in txs if str(t.get('id')) == peer_id), None)
    err = _spending_manual_pair_error(a, b)
    if err:
        return jsonify({'error': err}), 400
    _apply_spending_pair_to_rows(a, b, 'manual')
    bucket = str(_report_month_for_spending_tx(a) or '')
    _recompute_monthly_insights(spending, {bucket} if bucket else None)
    save_data(data)
    return jsonify({
        'ok': True,
        'transactions': [a, b],
        'insight': (spending.get('monthly_insights') or {}).get(bucket),
        'reconciliation': _spending_transfer_reconciliation_for_month(
            [t for t in txs if _report_month_for_spending_tx(t) == bucket]
        ),
    })


@app.route('/api/spending/recategorize', methods=['POST'])
@login_required
def spending_recategorize_month():
    """Re-run AI categorisation for outgoing transactions in a reporting month (manual overrides kept)."""
    payload = request.get_json(silent=True) or {}
    report_month = str(payload.get('report_month') or '').strip()[:7]
    data = load_data()
    spending, changed = _ensure_user_spending(data, session['username'])
    if changed:
        save_data(data)
    result = _rerun_spending_categorization_for_month(spending, report_month)
    if not result.get('ok'):
        return jsonify({'error': result.get('error', 'Recategorisation failed')}), 400
    save_data(data)
    target_month = report_month
    month_rows = [
        t for t in (spending.get('transactions') or [])
        if _report_month_for_spending_tx(t) == target_month
    ]
    month_rows.sort(key=lambda r: (r.get('date', ''), r.get('description', '')))
    return jsonify({
        'ok': True,
        'report_month': target_month,
        'labels_refreshed': result.get('labels_refreshed', 0),
        'insight': (spending.get('monthly_insights') or {}).get(target_month),
        'transactions': month_rows,
        'available_months': sorted((spending.get('monthly_insights') or {}).keys(), reverse=True),
    })


@app.route('/api/spending/pair/apply', methods=['POST'])
@login_required
def spending_pair_reapply():
    """Re-run auto internal-transfer matching for a reporting month."""
    payload = request.get_json(silent=True) or {}
    report_month = str(payload.get('report_month') or '').strip()[:7]
    if len(report_month) != 7 or report_month[4] != '-':
        return jsonify({'error': 'report_month (YYYY-MM) is required'}), 400
    data = load_data()
    spending, changed = _ensure_user_spending(data, session['username'])
    if changed:
        save_data(data)
    reconciliation = apply_auto_transfer_pairing_for_month(spending, report_month)
    _recompute_monthly_insights(spending, {report_month})
    save_data(data)
    return jsonify({
        'ok': True,
        'reconciliation': reconciliation,
        'insight': (spending.get('monthly_insights') or {}).get(report_month),
    })


@app.route('/api/spending/daily/status', methods=['GET'])
@login_required
def spending_daily_status():
    data = load_data()
    spending, changed = _ensure_user_spending(data, session['username'])
    as_of_raw = (request.args.get('date') or '').strip()
    as_of = _parse_iso_date(as_of_raw) if as_of_raw else _daily_budget_today()
    if as_of is None:
        return jsonify({'error': 'Invalid date'}), 400
    debt_changed = _daily_budget_sync_overspend_state(spending, as_of)
    if changed or debt_changed:
        save_data(data)
    status = _daily_budget_status(spending, as_of=as_of)
    return jsonify({'ok': True, **status})


@app.route('/api/spending/daily/entry', methods=['POST'])
@login_required
def spending_daily_entry_create():
    payload = request.get_json(silent=True) or {}
    try:
        amount = float(payload.get('amount'))
    except (TypeError, ValueError):
        return jsonify({'error': 'Valid amount is required'}), 400
    if amount <= 0:
        return jsonify({'error': 'Amount must be positive'}), 400
    amount = round(amount, 2)
    title = str(payload.get('title') or payload.get('description') or '').strip()[:500]
    if not title:
        return jsonify({'error': 'Title is required'}), 400
    category = str(payload.get('category') or 'other').strip().lower()
    if category not in SPENDING_CATEGORY_SET or category == 'unclassified':
        category = 'other'
    date_raw = str(payload.get('date') or '').strip()
    today = _daily_budget_today()
    d = _parse_iso_date(date_raw) if date_raw else today
    if d is None:
        return jsonify({'error': 'Invalid date'}), 400
    if d > today:
        return jsonify({'error': 'Cannot log spend for a future date'}), 400

    data = load_data()
    spending, changed = _ensure_user_spending(data, session['username'])
    if changed:
        save_data(data)

    d_str = d.strftime('%Y-%m-%d')
    month_key = d.strftime('%Y-%m')
    tx_id = str(uuid.uuid4())
    base_fp = _spending_fingerprint(month_key, d_str, amount, 'outgoing', title)
    fp = f'{base_fp}|manual|{tx_id}'

    now_iso = datetime.utcnow().isoformat() + 'Z'
    tx = {
        'id': tx_id,
        'date': d_str,
        'month': month_key,
        'report_month': month_key,
        'description': title,
        'amount': amount,
        'direction': 'outgoing',
        'category': category,
        'confidence': 1.0,
        'rationale': 'manual daily entry',
        'source': 'manual',
        'created_at': now_iso,
        'fingerprint': fp,
    }
    spending.setdefault('transactions', []).append(tx)
    _recompute_monthly_insights(spending, {month_key})
    save_data(data)
    status = _daily_budget_status(spending, as_of=d)
    return jsonify({'ok': True, 'transaction': tx, 'status': status})


@app.route('/api/spending/daily/entry/<tx_id>', methods=['DELETE'])
@login_required
def spending_daily_entry_delete(tx_id):
    data = load_data()
    spending, changed = _ensure_user_spending(data, session['username'])
    if changed:
        save_data(data)
    txs = spending.get('transactions') or []
    target = None
    for t in txs:
        if str(t.get('id')) == str(tx_id):
            target = t
            break
    if target is None:
        return jsonify({'error': 'Transaction not found'}), 404
    if str(target.get('source') or '') != 'manual':
        return jsonify({'error': 'Only manual daily entries can be deleted here'}), 400
    month_key = _report_month_for_spending_tx(target) or str(target.get('month') or '')
    spending['transactions'] = [t for t in txs if str(t.get('id')) != str(tx_id)]
    if month_key:
        _recompute_monthly_insights(spending, {month_key})
    save_data(data)
    status = _daily_budget_status(spending)
    return jsonify({'ok': True, 'status': status})


@app.route('/api/spending/daily/plan', methods=['GET', 'PUT'])
@login_required
def spending_daily_plan():
    data = load_data()
    spending, changed = _ensure_user_spending(data, session['username'])
    if changed:
        save_data(data)
    bucket, _ = _ensure_daily_budget(spending)
    plan = bucket['plan']

    if request.method == 'GET':
        return jsonify({'ok': True, 'plan': _serialize_daily_budget_plan(plan)})

    payload = request.get_json(silent=True) or {}
    first_save = not plan.get('updated_at')
    if 'income_monthly' in payload:
        try:
            plan['income_monthly'] = round(max(0.0, float(payload.get('income_monthly'))), 2)
        except (TypeError, ValueError):
            return jsonify({'error': 'Invalid income_monthly'}), 400
    if 'savings_percent' in payload:
        try:
            pct = float(payload.get('savings_percent'))
        except (TypeError, ValueError):
            return jsonify({'error': 'Invalid savings_percent'}), 400
        plan['savings_percent'] = round(max(0.0, min(100.0, pct)), 2)
    if 'daily_mode' in payload:
        mode = str(payload.get('daily_mode') or '').strip()
        if mode not in DAILY_BUDGET_MODES:
            return jsonify({'error': f'daily_mode must be one of {sorted(DAILY_BUDGET_MODES)}'}), 400
        plan['daily_mode'] = mode
    if 'underspend_priority' in payload:
        priority = str(payload.get('underspend_priority') or '').strip()
        if priority not in DAILY_BUDGET_UNDERSPEND_PRIORITIES:
            return jsonify({
                'error': f'underspend_priority must be one of {sorted(DAILY_BUDGET_UNDERSPEND_PRIORITIES)}',
            }), 400
        plan['underspend_priority'] = priority
    if 'bill_items' in payload:
        plan['bill_items'] = _normalize_daily_bill_items(payload.get('bill_items'))
        plan['bills_monthly'] = round(
            sum(float(b['amount']) for b in plan['bill_items'] if b.get('included', True)),
            2,
        )
    elif 'bills_monthly' in payload:
        try:
            plan['bills_monthly'] = round(max(0.0, float(payload.get('bills_monthly'))), 2)
        except (TypeError, ValueError):
            return jsonify({'error': 'Invalid bills_monthly'}), 400
    if 'source_month' in payload:
        sm = payload.get('source_month')
        if sm in (None, ''):
            plan['source_month'] = None
        else:
            plan['source_month'] = _normalize_spending_month_key(str(sm))
    if 'pay_day' in payload:
        try:
            pay_day = int(payload.get('pay_day'))
        except (TypeError, ValueError):
            return jsonify({'error': 'Invalid pay_day (use 1–31)'}), 400
        if pay_day < 1 or pay_day > 31:
            return jsonify({'error': 'pay_day must be between 1 and 31'}), 400
        plan['pay_day'] = pay_day
    if 'tracking_from' in payload:
        raw_tf = payload.get('tracking_from')
        if raw_tf in (None, ''):
            plan['tracking_from'] = None
        else:
            tf = _daily_budget_parse_tracking_from(raw_tf)
            if tf is None:
                return jsonify({'error': 'Invalid tracking_from (use YYYY-MM-DD)'}), 400
            plan['tracking_from'] = tf.isoformat()
    elif first_save and not plan.get('tracking_from'):
        # Mid-period onboarding: pace from today so empty earlier days aren't £0.
        plan['tracking_from'] = _daily_budget_today().isoformat()
    plan['updated_at'] = datetime.utcnow().isoformat() + 'Z'
    save_data(data)
    status = _daily_budget_status(spending)
    return jsonify({'ok': True, 'plan': _serialize_daily_budget_plan(plan), 'status': status})


@app.route('/api/spending/daily/plan/from-statements', methods=['POST'])
@login_required
def spending_daily_plan_from_statements():
    payload = request.get_json(silent=True) or {}
    data = load_data()
    spending, changed = _ensure_user_spending(data, session['username'])
    if changed:
        save_data(data)
    months = sorted((spending.get('monthly_insights') or {}).keys(), reverse=True)
    month = _normalize_spending_month_key(str(payload.get('month') or ''))
    if not month:
        month = months[0] if months else None
    if not month:
        return jsonify({
            'error': 'No statement months available. Import a statement in Monthly Spending first.',
            'months': [],
        }), 400
    use_llm = payload.get('use_llm', True)
    if isinstance(use_llm, str):
        use_llm = use_llm.strip().lower() not in ('0', 'false', 'no')
    estimate = _build_hybrid_bill_estimate(spending, month, use_llm=bool(use_llm))
    if estimate.get('error'):
        return jsonify({'error': estimate['error']}), 400

    apply = payload.get('apply', False)
    if apply:
        bucket, _ = _ensure_daily_budget(spending)
        plan = bucket['plan']
        plan['income_monthly'] = float(estimate['income_monthly'])
        plan['bill_items'] = estimate['bill_items']
        plan['bills_monthly'] = float(estimate['bills_monthly'])
        plan['source_month'] = month
        plan['updated_at'] = datetime.utcnow().isoformat() + 'Z'
        save_data(data)
        status = _daily_budget_status(spending)
        return jsonify({
            'ok': True,
            'estimate': estimate,
            'plan': _serialize_daily_budget_plan(plan),
            'status': status,
            'available_months': months,
        })

    return jsonify({'ok': True, 'estimate': estimate, 'available_months': months})


@app.route('/api/spending/daily/goals', methods=['GET', 'POST'])
@login_required
def spending_daily_goals():
    data = load_data()
    spending, changed = _ensure_user_spending(data, session['username'])
    if changed:
        save_data(data)
    bucket, _ = _ensure_daily_budget(spending)
    goals = bucket.setdefault('goals', [])

    if request.method == 'GET':
        status = _daily_budget_status(spending)
        return jsonify({
            'ok': True,
            'goals': goals,
            'underspend_saved': status.get('underspend_saved', 0),
        })

    payload = request.get_json(silent=True) or {}
    name = str(payload.get('name') or '').strip()[:120]
    if not name:
        return jsonify({'error': 'Goal name is required'}), 400
    try:
        target = round(max(0.01, float(payload.get('target_amount'))), 2)
    except (TypeError, ValueError):
        return jsonify({'error': 'Valid target_amount is required'}), 400
    goal = {
        'id': str(uuid.uuid4()),
        'name': name,
        'target_amount': target,
        'created_at': datetime.utcnow().isoformat() + 'Z',
    }
    goals.append(goal)
    save_data(data)
    status = _daily_budget_status(spending)
    return jsonify({'ok': True, 'goal': goal, 'goals': goals, 'underspend_saved': status.get('underspend_saved', 0)})


@app.route('/api/spending/daily/goals/<goal_id>', methods=['PATCH', 'DELETE'])
@login_required
def spending_daily_goal_item(goal_id):
    data = load_data()
    spending, changed = _ensure_user_spending(data, session['username'])
    if changed:
        save_data(data)
    bucket, _ = _ensure_daily_budget(spending)
    goals = bucket.setdefault('goals', [])
    idx = next((i for i, g in enumerate(goals) if str(g.get('id')) == str(goal_id)), None)
    if idx is None:
        return jsonify({'error': 'Goal not found'}), 404

    if request.method == 'DELETE':
        goals.pop(idx)
        save_data(data)
        return jsonify({'ok': True, 'goals': goals})

    payload = request.get_json(silent=True) or {}
    goal = goals[idx]
    if 'name' in payload:
        name = str(payload.get('name') or '').strip()[:120]
        if not name:
            return jsonify({'error': 'Goal name is required'}), 400
        goal['name'] = name
    if 'target_amount' in payload:
        try:
            goal['target_amount'] = round(max(0.01, float(payload.get('target_amount'))), 2)
        except (TypeError, ValueError):
            return jsonify({'error': 'Invalid target_amount'}), 400
    save_data(data)
    status = _daily_budget_status(spending)
    return jsonify({'ok': True, 'goal': goal, 'goals': goals, 'underspend_saved': status.get('underspend_saved', 0)})


@app.route('/api/spending/daily/overspend/decision', methods=['POST'])
@login_required
def spending_daily_overspend_decision():
    """Opt in/out of carrying a completed period's net overspend as repay debt."""
    payload = request.get_json(silent=True) or {}
    decision = str(payload.get('decision') or '').strip().lower()
    if decision in ('accept', 'accepted', 'carry'):
        decision = 'accept'
    elif decision in ('decline', 'declined', 'skip', 'not_this_time'):
        decision = 'decline'
    else:
        return jsonify({'error': 'decision must be accept or decline'}), 400

    period_start = _parse_iso_date(str(payload.get('period_start') or '').strip())
    period_end = _parse_iso_date(str(payload.get('period_end') or '').strip())
    if period_start is None or period_end is None or period_end < period_start:
        return jsonify({'error': 'period_start and period_end are required as YYYY-MM-DD'}), 400

    data = load_data()
    spending, changed = _ensure_user_spending(data, session['username'])
    as_of = _daily_budget_today()
    debt_changed = _daily_budget_sync_overspend_state(spending, as_of)
    bucket, _ = _ensure_daily_budget(spending)
    plan = bucket.get('plan') or {}
    figures = _daily_budget_plan_figures(plan)
    pay_day = _daily_budget_parse_pay_day(plan.get('pay_day'))
    current_start, _current_end = _daily_budget_pay_period(as_of, pay_day)
    prev_start, prev_end = _daily_budget_previous_pay_period(current_start, pay_day)
    if period_start != prev_start or period_end != prev_end:
        return jsonify({'error': 'Can only decide for the immediately previous pay period'}), 400
    if as_of <= prev_end:
        return jsonify({'error': 'That pay period has not finished yet'}), 400

    key = _daily_budget_period_key(period_start, period_end)
    decisions = bucket.get('overspend_decisions') if isinstance(bucket.get('overspend_decisions'), dict) else {}
    if key in decisions:
        return jsonify({'error': 'Already decided for this pay period'}), 400

    summary = _daily_budget_period_net_overspend(
        spending, plan, figures, period_start, period_end,
    )
    net = float(summary.get('net_overspend') or 0)
    if net <= 0:
        return jsonify({'error': 'No net overspend for that pay period'}), 400

    status_before = _daily_budget_status(spending, as_of=as_of)
    current_bal = 0.0
    if isinstance(status_before.get('overspend_debt'), dict):
        current_bal = float(status_before['overspend_debt'].get('balance') or 0)

    if decision == 'accept':
        _daily_budget_accept_overspend(
            bucket,
            period_start=period_start,
            period_end=period_end,
            net_overspend=net,
            current_period_start=current_start,
            current_display_balance=current_bal,
        )
    else:
        _daily_budget_decline_overspend(
            bucket,
            period_start=period_start,
            period_end=period_end,
            net_overspend=net,
        )

    save_data(data)
    status = _daily_budget_status(spending, as_of=as_of)
    return jsonify({
        'ok': True,
        'decision': decision,
        'net_overspend': net,
        'status': status,
        'changed': changed or debt_changed,
    })


@app.route('/api/spending/daily/overspend/write-off', methods=['POST'])
@login_required
def spending_daily_overspend_write_off():
    data = load_data()
    spending, changed = _ensure_user_spending(data, session['username'])
    as_of = _daily_budget_today()
    debt_changed = _daily_budget_sync_overspend_state(spending, as_of)
    bucket, _ = _ensure_daily_budget(spending)
    debt = bucket.get('overspend_debt') if isinstance(bucket.get('overspend_debt'), dict) else None
    if debt is None or _daily_budget_debt_open_balance(debt) <= 0:
        # Also allow write-off when only in-period display balance remains.
        status_check = _daily_budget_status(spending, as_of=as_of)
        od = status_check.get('overspend_debt')
        if not isinstance(od, dict) or float(od.get('balance') or 0) <= 0:
            return jsonify({'error': 'No overspend debt to write off'}), 400

    wiped = _daily_budget_write_off_debt(bucket)
    save_data(data)
    status = _daily_budget_status(spending, as_of=as_of)
    return jsonify({
        'ok': True,
        'written_off': wiped,
        'status': status,
        'changed': changed or debt_changed,
    })


if __name__ == '__main__':
    # Schedule initial interest tasks and recurring payment processing for all loans
    data = load_data()
    for loan_id, loan in data['loans'].items():
        schedule_interest_task(loan_id, loan['interest_day'])
        if 'recurring_payments' in loan and loan['recurring_payments']:
            schedule_recurring_payments(loan_id)
    schedule_bill_reminders_job()
    host = os.getenv('HOST', '0.0.0.0').strip() or '0.0.0.0'
    try:
        port = int(os.getenv('PORT', '5000') or '5000')
    except ValueError:
        port = 5000
    app.run(host=host, port=port, debug=False, use_reloader=False)
